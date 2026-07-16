import logging
import io
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Count, Sum, Q, Max
from django.http import HttpResponse
import pandas as pd

from ..models import (
    MSME, Cohort, ProgrammeGroup, MSMEGrowthSnapshot,
    BusinessGrowthExpert, BGEGroup, MSMEReport, GroupReport,
)
from ..serializers import (
    CohortSerializer, MSMESerializer, MSMEGrowthSnapshotSerializer,
)
from .mixins import (
    ViewerReadOnlyMixin, _managed_groups, _is_viewer, _is_programme_manager,
)

logger = logging.getLogger(__name__)


class CohortViewSet(ViewerReadOnlyMixin, viewsets.ModelViewSet):
    queryset = Cohort.objects.all()
    serializer_class = CohortSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Annotate msme_count at DB level so CohortSerializer.get_msme_count
        # can avoid a per-row COUNT query (N+1 fix).
        return Cohort.objects.annotate(_msme_count=Count('msmes', distinct=True))

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff and not request.user.is_superuser:
            raise PermissionDenied("Only admins can delete cohorts.")
        return super().destroy(request, *args, **kwargs)


class ProgrammeGroupViewSet(viewsets.ModelViewSet):
    """Cross-cutting labels that can be applied to MSMEs (e.g. Green MSMEs, Agroprocessors).
    Read-only for non-admins; create/update/delete restricted to admins."""
    queryset = ProgrammeGroup.objects.all()
    serializer_class = None  # defined below via import
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Annotate msme_count at DB level so ProgrammeGroupSerializer.get_msme_count
        # can avoid a per-row COUNT query (N+1 fix).
        return ProgrammeGroup.objects.annotate(_msme_count=Count('msmes', distinct=True))

    def get_serializer_class(self):
        from ..serializers import ProgrammeGroupSerializer
        return ProgrammeGroupSerializer

    def create(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can create programme groups.")
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can edit programme groups.")
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can delete programme groups.")
        return super().destroy(request, *args, **kwargs)


class MSMEGrowthSnapshotViewSet(ViewerReadOnlyMixin, viewsets.ModelViewSet):
    """
    Growth snapshots for a single MSME.

    List/retrieve all snapshots for a given MSME:
        GET /api/growth-snapshots/?msme=<id>

    Create a new BGE-visit snapshot:
        POST /api/growth-snapshots/
        { msme, snapshot_date, source, annual_turnover, ... }
    """
    serializer_class   = MSMEGrowthSnapshotSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = MSMEGrowthSnapshot.objects.select_related('msme', 'collected_by').order_by('snapshot_date')
        msme_id = self.request.query_params.get('msme')
        bge_id  = self.request.query_params.get('bge')
        if msme_id:
            qs = qs.filter(msme_id=msme_id)
        if bge_id:
            # Security: non-admin users can only query their own BGE data (IDOR fix)
            user = self.request.user
            if not (user.is_staff or user.is_superuser or _managed_groups(user) is not None):
                own_bge = getattr(user, 'bge_profile', None)
                if not own_bge or str(own_bge.id) != str(bge_id):
                    raise PermissionDenied("You can only access your own data.")
            # All snapshots for MSMEs directly assigned to this BGE or in their groups
            from ..models import BusinessGrowthExpert, BGEGroup
            try:
                bge = BusinessGrowthExpert.objects.get(pk=bge_id)
                group_msme_ids = MSME.objects.filter(assigned_group__in=bge.bge_groups.all()).values_list('id', flat=True)
                direct_msme_ids = MSME.objects.filter(assigned_bge=bge).values_list('id', flat=True)
                co_msme_ids = MSME.objects.filter(co_assigned_bges=bge).values_list('id', flat=True)
                all_ids = set(list(direct_msme_ids) + list(group_msme_ids) + list(co_msme_ids))
                qs = qs.filter(msme_id__in=all_ids)
            except BusinessGrowthExpert.DoesNotExist:
                qs = qs.none()
        return qs

    def perform_create(self, serializer):
        """Auto-set collected_by from the logged-in user's identity.

        Resolution order (first match wins):
        1. Value already in the request — respect it (admin attribution on behalf of BGE).
        2. user.bge_profile OneToOne — the clean path once accounts are linked.
        3. Username match: BGE whose name slug equals the Django username (e.g. jimmy.ouni → Jimmy Ouni).
        4. Email match: BGE whose email local-part equals the username.
        Falls through silently so admin submissions (no bge_profile) still save.
        """
        if serializer.validated_data.get('collected_by'):
            serializer.save()
            return

        user = self.request.user

        # Path 2 — linked profile
        try:
            serializer.save(collected_by=user.bge_profile)
            return
        except Exception:
            pass

        # Path 3 — match BGE name slug to username  (e.g. "jimmy.ouni" → "Jimmy Ouni")
        from ..models import BusinessGrowthExpert
        import re

        def _slug(text):
            s = re.sub(r'[^a-z0-9]+', '.', (text or '').lower()).strip('.')
            return re.sub(r'\.+', '.', s)

        uname = user.username.lower()
        for bge in BusinessGrowthExpert.objects.filter(user__isnull=True):
            if _slug(bge.name) == uname:
                # Auto-link for next time so we don't need to re-scan
                bge.user = user
                bge.save(update_fields=['user'])
                serializer.save(collected_by=bge)
                return

        # Path 4 — email local-part match
        for bge in BusinessGrowthExpert.objects.filter(user__isnull=True):
            if bge.email and bge.email.split('@')[0].lower() == uname:
                bge.user = user
                bge.save(update_fields=['user'])
                serializer.save(collected_by=bge)
                return

        serializer.save()

    def destroy(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can delete snapshots.")
        return super().destroy(request, *args, **kwargs)


class MSMEViewSet(ViewerReadOnlyMixin, viewsets.ModelViewSet):
    queryset = MSME.objects.filter(is_active=True)
    serializer_class = MSMESerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = MSME.objects.filter(is_active=True)

        user = self.request.user
        group_ids = _managed_groups(user)
        if group_ids is not None:
            # Programme manager — scope to MSMEs in their managed programme groups
            qs = qs.filter(programme_groups__in=group_ids).distinct()
        elif _is_viewer(user):
            pass  # viewers see all MSMEs read-only (ViewerReadOnlyMixin blocks writes)
        elif not (user.is_staff or user.is_superuser):
            # BGE users only see their own assigned MSMEs — directly OR via any group they belong to.
            # Exception: ?training=1 returns all MSMEs so a facilitator can record attendance
            # for participants who are not personally assigned to them.
            training_context = self.request.query_params.get('training') == '1'
            if not training_context:
                try:
                    bge = user.bge_profile
                    from django.db.models import Q
                    qs = qs.filter(
                        Q(assigned_bge=bge) |
                        Q(assigned_group__members=bge) |
                        Q(co_assigned_bges=bge)
                    ).distinct()
                except Exception:
                    qs = qs.none()

        search = self.request.query_params.get('search')
        if search:
            # Combine search clauses with Q() inside ONE .filter(), not via
            # queryset union (`qs.filter(...) | qs.filter(...)`) — the union
            # form silently drops the BGE tenant scoping that was applied
            # above, leaking every other BGE's MSMEs into search results.
            from django.db.models import Q
            qs = qs.filter(
                Q(business_name__icontains=search) |
                Q(owner_name__icontains=search)    |
                Q(sector__icontains=search)        |
                Q(msme_code__icontains=search)
            )

        business_type = self.request.query_params.get('business_type')
        if business_type:
            qs = qs.filter(business_type=business_type)

        sector = self.request.query_params.get('sector')
        if sector:
            qs = qs.filter(sector=sector)

        cohort = self.request.query_params.get('cohort')
        if cohort:
            qs = qs.filter(cohort_id=cohort)

        city = self.request.query_params.get('city')
        if city:
            qs = qs.filter(city__iexact=city)

        return (
            qs.select_related('cohort', 'assigned_bge', 'assigned_group')
            .prefetch_related('programme_groups')
            # Annotate counts and latest dates at the DB level to eliminate the
            # N+1 queries that MSMESerializer.get_total_reports / get_last_support_date
            # would otherwise fire (one round-trip per row).
            .annotate(
                _reports_count=Count('reports', distinct=True),
                _group_reports_count=Count('group_reports', distinct=True),
                _last_individual_date=Max('reports__visit_date'),
                _last_group_date=Max('group_reports__visit_date'),
            )
            .order_by('-created_at')
        )

    def _is_admin_or_cohort_admin(self, request):
        u = request.user
        if u.is_staff or u.is_superuser:
            return True
        return _managed_groups(u) is not None

    def destroy(self, request, *args, **kwargs):
        if not self._is_admin_or_cohort_admin(request):
            raise PermissionDenied("Only admins can delete MSMEs.")
        return super().destroy(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        if not self._is_admin_or_cohort_admin(request):
            raise PermissionDenied("Only admins can create MSMEs.")
        return super().create(request, *args, **kwargs)

    @action(detail=True, methods=['patch'])
    def assign_bge(self, request, pk=None):
        if not self._is_admin_or_cohort_admin(request):
            raise PermissionDenied("Only admins can assign BGEs.")
        msme = self.get_object()
        # Programme managers may only assign BGEs to MSMEs within their managed groups
        if _is_programme_manager(request.user):
            managed = _managed_groups(request.user) or []
            if not msme.programme_groups.filter(id__in=managed).exists():
                raise PermissionDenied("You can only assign BGEs to MSMEs in your managed programme groups.")
        bge_id = request.data.get('bge_id')
        objectives = (request.data.get('objectives') or '').strip()
        assignment_date = request.data.get('assignment_date') or None
        bge = None  # initialise so the notification block below is always safe
        if bge_id:
            try:
                bge = BusinessGrowthExpert.objects.get(pk=bge_id)
                # Same primary BGE re-submitted — just update objectives/date
                if msme.assigned_bge_id and msme.assigned_bge_id == bge.id:
                    msme.assignment_objectives = objectives
                    msme.assignment_date = assignment_date
                    msme.save()
                    return Response(MSMESerializer(msme).data)
                if msme.assigned_bge_id and msme.assigned_bge_id != bge.id:
                    # MSME already has a primary BGE — add the new BGE as co-assigned
                    # so BOTH BGEs keep the MSME in their list (joint deployment).
                    existing_primary = msme.assigned_bge  # capture before save
                    msme.co_assigned_bges.add(bge)
                    msme.save()
                    # Notify the new (co-assigned) BGE via push
                    from .bge import _notify_bge, _send_co_assignment_alert
                    _notify_bge(
                        bge,
                        title='Joint MSME Assignment',
                        body=f'You have been co-assigned to {msme.business_name} alongside {existing_primary.name}. Check your dashboard for details.',
                        url='/bge'
                    )
                    # Notify the existing primary BGE via push + email
                    _notify_bge(
                        existing_primary,
                        title='Joint Deployment Notice',
                        body=f'{bge.name} has also been assigned to visit {msme.business_name}. Check your dashboard for details.',
                        url='/bge'
                    )
                    _send_co_assignment_alert(existing_primary, bge, msme)
                    return Response(MSMESerializer(msme).data)
                # No existing primary — set this BGE as primary
                msme.assigned_bge = bge
            except BusinessGrowthExpert.DoesNotExist:
                return Response({'error': 'BGE not found'}, status=status.HTTP_404_NOT_FOUND)
        else:
            # Unassign: clear primary and all co-assignees
            msme.assigned_bge = None
            msme.co_assigned_bges.clear()
        msme.assignment_objectives = objectives
        msme.assignment_date = assignment_date
        msme.save()
        # Notify the BGE about the new assignment
        if bge_id and bge:
            from .bge import _notify_bge
            _notify_bge(
                bge,
                title='New MSME Assignment',
                body=f'You have been assigned to {msme.business_name}. Check your dashboard for details.',
                url='/bge'
            )
        return Response(MSMESerializer(msme).data)

    @action(detail=True, methods=['patch'])
    def remove_co_assigned(self, request, pk=None):
        if not self._is_admin_or_cohort_admin(request):
            raise PermissionDenied("Only admins can modify BGE assignments.")
        msme = self.get_object()
        bge_id = request.data.get('bge_id')
        if not bge_id:
            return Response({'error': 'bge_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            bge = BusinessGrowthExpert.objects.get(pk=bge_id)
            msme.co_assigned_bges.remove(bge)
        except BusinessGrowthExpert.DoesNotExist:
            return Response({'error': 'BGE not found'}, status=status.HTTP_404_NOT_FOUND)
        return Response(MSMESerializer(msme).data)

    @action(detail=True, methods=['patch'])
    def assign_cohort(self, request, pk=None):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can assign cohorts.")
        msme = self.get_object()
        cohort_id = request.data.get('cohort_id')
        if cohort_id:
            try:
                cohort = Cohort.objects.get(pk=cohort_id)
                msme.cohort = cohort
            except Cohort.DoesNotExist:
                return Response({'error': 'Cohort not found'}, status=status.HTTP_404_NOT_FOUND)
        else:
            msme.cohort = None
        msme.save()
        return Response(MSMESerializer(msme).data)

    @action(detail=True, methods=['patch'], url_path='set-groups')
    def set_groups(self, request, pk=None):
        """Add or remove this MSME from programme groups.

        Body: { "group_ids": [1, 2, ...] }   — replaces the full set.
        Body: { "add": [1], "remove": [2] }   — incremental add/remove.
        """
        if not self._is_admin_or_cohort_admin(request):
            raise PermissionDenied("Only admins can modify programme group membership.")
        msme = self.get_object()
        if 'group_ids' in request.data:
            ids = request.data['group_ids'] or []
            msme.programme_groups.set(ids)
        else:
            add_ids    = request.data.get('add', [])
            remove_ids = request.data.get('remove', [])
            if add_ids:
                msme.programme_groups.add(*add_ids)
            if remove_ids:
                msme.programme_groups.remove(*remove_ids)
        return Response(MSMESerializer(msme, context={'request': request}).data)

    @action(detail=False, methods=['get'], url_path='upload-template', permission_classes=[])
    def upload_template(self, request):
        """Download a blank MSME upload template (.xlsx) with the unified schema."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "MSMEs"

        columns = [
            ('Business Name',     32, 'Required. The legal/trading name of the MSME.'),
            ('Owner Name',        24, 'Primary contact / owner name.'),
            ('Sex',               10, 'Male or Female.'),
            ('Phone',             18, 'Owner/contact phone number.'),
            ('Email',             28, 'Owner/contact email.'),
            ('Business Email',    28, 'Business email (if separate from owner email).'),
            ('Business Type',     20, 'Micro / Small / Medium / Sole proprietorship / Company / Partnership / Cooperative.'),
            ('Sector',            22, 'Optional. Manufacturing / Services / Trade / Agriculture / Technology / Healthcare / Education / Construction / Other. Auto-inferred from Business Type if blank.'),
            ('District',          18, 'District of operation.'),
            ('Town',              16, 'Town or city.'),
            ('Physical Location', 32, 'Street / landmark address (optional).'),
            ('Role',              18, 'Role of contact person (Director, Manager, etc.).'),
            ('Cohort',            14, 'e.g. Cohort 1, Cohort 2 (optional — can also be set in upload form).'),
        ]

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', start_color='1A2F4B')
        header_align = Alignment(horizontal='center', vertical='center')

        for i, (name, width, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=i, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[1].height = 28
        ws.freeze_panes = 'A2'

        # Add an example row so users see the expected shape
        example = [
            'Acholi Shea Cooperative Ltd', 'Anena Sharon', 'Female', '256775779335',
            'anena.sharon@example.com', 'info@acholishea.org', 'Cooperative',
            'Agriculture', 'Gulu', 'Gulu', 'Plot 12, Gulu town', 'Director', 'Cohort 1',
        ]
        for i, val in enumerate(example, start=1):
            c = ws.cell(row=2, column=i, value=val)
            c.font = Font(italic=True, color='888888')

        # Field-guidance row at the bottom
        guide_row = 4
        ws.cell(row=guide_row, column=1, value='Notes:').font = Font(bold=True, color='C8102E')
        for i, (_, _, note) in enumerate(columns, start=1):
            ws.cell(row=guide_row + 1, column=i, value=note).font = Font(size=9, color='666666', italic=True)
            ws.cell(row=guide_row + 1, column=i).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[guide_row + 1].height = 60

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="MSME_Upload_Template.xlsx"'
        return resp

    @action(detail=False, methods=['post'], url_path='upload')
    @transaction.atomic
    def upload(self, request):
        """Upload MSME list from Cohort 1 (CSV/Excel) or Cohort 2 (Survey Excel).
        Auto-detects format from column names.

        The whole import runs inside one DB transaction. Per-row exceptions are
        captured into the `skipped` list (so one bad row doesn't kill the rest),
        but if anything escapes that try/except — or the function returns with
        an HTTP error after rows were inserted — the transaction is rolled back
        and the table is left in its pre-import state.
        Accepts optional form field: cohort_name (string).
        """
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can upload MSME data.")

        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        # update_existing=true (default) → update duplicates; false → skip them
        update_existing = request.data.get('update_existing', 'true').lower() != 'false'
        cohort_name = request.data.get('cohort_name', '').strip()

        # Read file (CSV or Excel)
        try:
            raw_bytes = file.read()
            if file.name.endswith('.csv'):
                df = pd.read_csv(io.BytesIO(raw_bytes))
            elif file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(io.BytesIO(raw_bytes))
                # If the first row appears to be blank/unnamed, the real header may be in row 1
                unnamed = sum(1 for c in df.columns if str(c).startswith('Unnamed:'))
                if unnamed > len(df.columns) / 2:
                    df = pd.read_excel(io.BytesIO(raw_bytes), header=1)
            else:
                return Response({'error': 'Please upload a CSV or Excel file (.csv, .xlsx, .xls).'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': f'Could not read file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        # Normalise column names for detection (strip whitespace, keep original for lookup)
        cols = set(df.columns.tolist())
        cols_stripped = {c.strip() for c in cols}

        # Unified template — preferred format used by the cleaning script.
        # Required: Business Name, Owner Name, Phone, Email, District, Town, Sex, Business Type
        is_unified = ('Business Name' in cols_stripped and 'Owner Name' in cols_stripped)
        # Cohort 2 — numbered survey format (1.1. Business Name: …)
        is_cohort2 = '1.1.  Business Name:' in cols or any('Business Name' in c for c in cols if '1.1' in c)
        # Cohort 1 — PRUDEV II Cohort1 format (Name, District, Town, Name of contact person …)
        is_cohort1 = (not is_unified) and ('Name' in cols and 'Name of contact person' in cols)
        # Cohort 2 simple — survey export with plain headers (Business Name:, Name of Business Owner: …)
        is_cohort2_simple = (
            not is_unified and not is_cohort1 and not is_cohort2 and
            any('Business Name' in c for c in cols_stripped) and
            any('Business Owner' in c for c in cols_stripped)
        )

        if not is_unified and not is_cohort1 and not is_cohort2 and not is_cohort2_simple:
            return Response({
                'error': (
                    'Unrecognised file format. Use the unified upload template '
                    '(columns: Business Name, Owner Name, Sex, Phone, Email, Business Email, '
                    'Business Type, District, Town, Physical Location, Role).'
                )
            }, status=status.HTTP_400_BAD_REQUEST)

        # Resolve or create cohort
        cohort_obj = None
        if cohort_name:
            cohort_obj, _ = Cohort.objects.get_or_create(name=cohort_name)

        def clean_str(val, default=''):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return default
            s = str(val).strip()
            return '' if s == 'nan' else s

        def clean_phone(val):
            raw = clean_str(val)
            if not raw:
                return ''
            try:
                return str(int(float(raw)))
            except (ValueError, TypeError):
                return raw

        def map_gender(val):
            v = clean_str(val).upper()
            if v in ('M', 'MALE'):
                return 'MALE'
            if v in ('F', 'FEMALE'):
                return 'FEMALE'
            return ''

        def map_business_type(val):
            v = clean_str(val).upper()
            if 'MEDIUM' in v:
                return 'MEDIUM'
            if 'SMALL' in v or 'COMPANY' in v or 'SMC' in v or 'PARTNERSHIP' in v:
                return 'SMALL'
            return 'MICRO'  # sole proprietorship, cooperative, default

        def map_sector(val):
            v = clean_str(val).upper()
            if any(x in v for x in ('AGRO', 'FARM', 'AGRICULTURE', 'AGRI')):
                return 'AGRICULTURE'
            if any(x in v for x in ('MANUFACTUR', 'PROCESSING', 'MILLER', 'MILL')):
                return 'MANUFACTURING'
            if any(x in v for x in ('TRADE', 'BUYER', 'SHOP', 'INPUT', 'VET')):
                return 'TRADE'
            if any(x in v for x in ('SERVICE', 'BDS', 'DEVELOPMENT', 'FINANCIAL', 'PROVIDER', 'FINANCE')):
                return 'SERVICES'
            if 'TECH' in v or 'INNOVATOR' in v:
                return 'TECHNOLOGY'
            if 'HEALTH' in v:
                return 'HEALTHCARE'
            if 'EDUCATION' in v:
                return 'EDUCATION'
            if 'CONSTRUCT' in v:
                return 'CONSTRUCTION'
            return 'OTHER'

        created, updated, skipped = 0, 0, []

        def find_col(cols, prefix, keyword=None):
            """Match column whose stripped name starts with prefix+ space/dot, optionally containing keyword."""
            import re
            pat = re.compile(r'^\s*' + re.escape(prefix) + r'[\s\.]')
            for c in cols:
                if pat.match(c):
                    if keyword is None or keyword.lower() in c.lower():
                        return c
            return None

        # Pre-compute Cohort 2 simple column mapping once (outside the row loop)
        # Strip all column names to a lookup dict: stripped_name → original_name
        col_map = {c.strip(): c for c in cols}

        def get_col(*keywords):
            """Return first column whose stripped name contains all keywords (case-insensitive)."""
            for c_stripped, c_orig in col_map.items():
                c_lower = c_stripped.lower()
                if all(k.lower() in c_lower for k in keywords):
                    return c_orig
            return None

        if is_unified:
            u_bname    = col_map.get('Business Name')
            u_owner    = col_map.get('Owner Name')
            u_sex      = col_map.get('Sex')
            u_phone    = col_map.get('Phone')
            u_email    = col_map.get('Email')
            u_bemail   = col_map.get('Business Email')
            u_type     = col_map.get('Business Type')
            u_sector   = col_map.get('Sector')   # optional — falls back to map_sector(Business Type)
            u_district = col_map.get('District')
            u_town     = col_map.get('Town')
            u_address  = col_map.get('Physical Location') or col_map.get('Address')

        if is_cohort2_simple:
            # Column names for Cohort 2 simple format (e.g. "Business Name:", "Name of Business Owner:", …)
            # Use exact strip-match first, then fallback to keyword search
            s2_bname    = col_map.get('Business Name:') or col_map.get('Business Name') or get_col('business name')
            s2_owner    = col_map.get('Name of Business Owner:') or col_map.get('Name of Business Owner') or get_col('name of business owner')
            # Phone: prefer dedicated business phone, fall back to owner contacts
            s2_phone    = (col_map.get('Business Phone Number(s):') or col_map.get('Business Phone Number')
                           or get_col('business phone') or get_col('business owner contact'))
            s2_email    = col_map.get("Business Owners Email:") or col_map.get("Business Owner's Email") or get_col('owner', 'email')
            s2_bemail   = col_map.get('Business email address') or col_map.get('Business email address ') or get_col('business email')
            s2_sex      = col_map.get('Sex') or col_map.get('Sex ') or get_col('sex')
            s2_type     = col_map.get('Type of Business: ') or col_map.get('Type of Business:') or col_map.get('Type of Business') or get_col('type of business')
            s2_district = col_map.get('District') or col_map.get('district')
            s2_town     = col_map.get('Town') or col_map.get('town')
            s2_sector   = None  # not present in this format

        # Pre-compute Cohort 2 numbered column mapping once (outside the row loop)
        if is_cohort2:
            c2_bname  = find_col(cols, '1.1',  'Business Name')
            c2_brn    = find_col(cols, '1.2',  'Registration')
            c2_owner  = find_col(cols, '1.4',  'Owner')
            c2_phone1 = find_col(cols, '1.5')   # owner contacts
            c2_sex    = find_col(cols, '1.6',  'Sex')
            c2_email  = find_col(cols, '1.7',  'Email')
            c2_type   = find_col(cols, '1.10', 'Type')
            c2_phone2 = find_col(cols, '1.12') # business phone (preferred)
            c2_bemail = find_col(cols, '1.13')
            c2_sector = find_col(cols, '2.1',  'core business')

        blank_rows = 0  # track silently-blank rows so user knows why count differs
        for i, row in df.iterrows():
            try:
                if is_unified:
                    business_name = clean_str(row.get(u_bname, '')) if u_bname else ''
                    if not business_name:
                        if all((pd.isna(v) or str(v).strip() in ('', 'nan')) for v in row.values):
                            blank_rows += 1
                        else:
                            skipped.append({'row': i + 2, 'error': 'Missing Business Name'})
                        continue
                    record = {
                        'business_name': business_name,
                        'owner_name':    clean_str(row.get(u_owner, '')) if u_owner else '',
                        'gender':        map_gender(row.get(u_sex, '')) if u_sex else '',
                        'phone':         clean_phone(row.get(u_phone, '')) if u_phone else '',
                        'email':         clean_str(row.get(u_email, '')) if u_email else '',
                        'business_email': clean_str(row.get(u_bemail, '')) if u_bemail else '',
                        'business_type': map_business_type(row.get(u_type, '')) if u_type else 'MICRO',
                        # Sector: prefer an explicit Sector column, otherwise infer from Business Type / sector keywords
                        'sector':        (
                            map_sector(row.get(u_sector, '')) if u_sector
                            else (map_sector(row.get(u_type, '')) if u_type else 'OTHER')
                        ),
                        'state':         clean_str(row.get(u_district, '')) if u_district else '',
                        'city':          clean_str(row.get(u_town, '')) if u_town else '',
                        'address':       clean_str(row.get(u_address, '')) if u_address else '',
                        'country':       'Uganda',
                        'source_file':   file.name,
                        'cohort':        cohort_obj,
                        'is_active':     True,
                    }
                elif is_cohort2_simple:
                    business_name = clean_str(row.get(s2_bname, '')) if s2_bname else ''
                    if not business_name:
                        # check if the whole row is blank vs just missing business name
                        if all((pd.isna(v) or str(v).strip() in ('', 'nan')) for v in row.values):
                            blank_rows += 1
                        else:
                            skipped.append({'row': i + 2, 'error': 'Missing Business Name'})
                        continue

                    record = {
                        'business_name': business_name,
                        'owner_name': clean_str(row.get(s2_owner, '')) if s2_owner else '',
                        'phone': clean_phone(row.get(s2_phone, '')) if s2_phone else '',
                        'email': clean_str(row.get(s2_email, '')) if s2_email else '',
                        'business_email': clean_str(row.get(s2_bemail, '')) if s2_bemail else '',
                        'gender': map_gender(row.get(s2_sex, '')) if s2_sex else '',
                        'business_type': map_business_type(row.get(s2_type, '')) if s2_type else 'MICRO',
                        'sector': 'OTHER',
                        'state': clean_str(row.get(s2_district, '')) if s2_district else '',
                        'city': clean_str(row.get(s2_town, '')) if s2_town else '',
                        'country': 'Uganda',
                        'source_file': file.name,
                        'cohort': cohort_obj,
                        'is_active': True,
                    }

                elif is_cohort2:
                    bname_col  = c2_bname
                    owner_col  = c2_owner
                    sex_col    = c2_sex
                    email_col  = c2_email
                    type_col   = c2_type
                    phone_col  = c2_phone2 or c2_phone1
                    bemail_col = c2_bemail
                    sector_col = c2_sector
                    brn_col    = c2_brn

                    business_name = clean_str(row.get(bname_col, '')) if bname_col else ''
                    if not business_name:
                        if all((pd.isna(v) or str(v).strip() in ('', 'nan')) for v in row.values):
                            blank_rows += 1
                        else:
                            skipped.append({'row': i + 2, 'error': 'Missing Business Name'})
                        continue

                    record = {
                        'business_name': business_name,
                        'registration_number': clean_str(row.get(brn_col, '')) if brn_col else '',
                        'owner_name': clean_str(row.get(owner_col, '')) if owner_col else '',
                        'gender': map_gender(row.get(sex_col, '')) if sex_col else '',
                        'email': clean_str(row.get(email_col, '')) if email_col else '',
                        'phone': clean_phone(row.get(phone_col, '')) if phone_col else '',
                        'business_email': clean_str(row.get(bemail_col, '')) if bemail_col else '',
                        'business_type': map_business_type(row.get(type_col, '')) if type_col else 'MICRO',
                        'sector': map_sector(row.get(sector_col, '')) if sector_col else 'OTHER',
                        'state': clean_str(row.get('District', '')),
                        'city': clean_str(row.get('Town/City', '')),
                        'country': 'Uganda',
                        'source_file': file.name,
                        'cohort': cohort_obj,
                        'is_active': True,
                    }
                else:  # Cohort 1
                    business_name = clean_str(row.get('Name', ''))
                    if not business_name:
                        if all((pd.isna(v) or str(v).strip() in ('', 'nan')) for v in row.values):
                            blank_rows += 1
                        else:
                            skipped.append({'row': i + 2, 'error': 'Missing Name (business)'})
                        continue

                    gender = map_gender(row.get('Sex of founder', row.get('Gender of Key contact person', '')))

                    record = {
                        'business_name': business_name,
                        'owner_name': clean_str(row.get('Name of contact person', '')),
                        'gender': gender,
                        'phone': clean_phone(row.get('Mobile phone numbers ', row.get('Mobile phone numbers', ''))),
                        'email': clean_str(row.get('Email Address of contact person', '')),
                        'business_email': clean_str(row.get('Business Email Address', '')),
                        'address': clean_str(row.get('Physical location', '')),
                        'state': clean_str(row.get('District', '')),
                        'city': clean_str(row.get('Town', '')),
                        'business_type': 'MICRO',
                        'sector': 'OTHER',
                        'country': 'Uganda',
                        'source_file': file.name,
                        'cohort': cohort_obj,
                        'is_active': True,
                    }

                # Update-or-create by business name + owner (avoid duplicates on re-upload).
                # Pop both keys out of `record` so they aren't passed twice to create().
                business_name = record.pop('business_name')
                owner_name    = record.pop('owner_name', '')
                lookup = {'business_name': business_name}
                if owner_name:
                    lookup['owner_name'] = owner_name

                # Wrap each row's DB write in a savepoint so an IntegrityError
                # on row N doesn't poison the outer transaction for rows N+1..end.
                with transaction.atomic():
                    existing = MSME.objects.filter(**lookup).first()
                    if existing:
                        if update_existing:
                            # Restore owner_name into the update payload so existing rows can be enriched
                            if owner_name and not existing.owner_name:
                                existing.owner_name = owner_name
                            for k, v in record.items():
                                setattr(existing, k, v)
                            existing.save()
                            updated += 1
                        else:
                            skipped.append({'row': i + 2, 'error': f'Duplicate skipped: {business_name}'})
                    else:
                        # `lookup` carries the unique-fields, `record` carries everything else
                        MSME.objects.create(**lookup, **record)
                        created += 1

            except Exception as e:
                skipped.append({'row': i + 2, 'error': str(e)})

        msg = f"{created} MSMEs added, {updated} updated"
        if cohort_name:
            msg += f" (assigned to cohort: {cohort_name})"
        if skipped:
            msg += f", {len(skipped)} rows skipped"
        if blank_rows:
            msg += f", {blank_rows} blank rows ignored"
        msg += "."

        return Response({
            'created': created,
            'updated': updated,
            'skipped': len(skipped),
            'blank_rows': blank_rows,
            'errors': skipped[:50],  # cap at 50 in response
            'total_rows': int(len(df)),
            'message': msg,
        })

    @action(detail=False, methods=['get'])
    def analytics(self, request):
        """Rich analytics for the dashboard. Accepts optional filters that all
        downstream aggregations honour:
          ?cohort=<id>  ?district=<name>  ?sector=<code>  ?bge=<id>
        """
        from django.db.models import Q
        from django.db.models.functions import TruncMonth

        qs = MSME.objects.filter(is_active=True)

        # Apply optional filters so the analytics page can drill down.
        cohort_id = request.query_params.get('cohort')
        if cohort_id:
            qs = qs.filter(cohort_id=cohort_id)
        district = request.query_params.get('district')
        if district:
            qs = qs.filter(state__iexact=district)
        sector = request.query_params.get('sector')
        if sector:
            qs = qs.filter(sector=sector)
        bge_id = request.query_params.get('bge')
        if bge_id:
            qs = qs.filter(Q(assigned_bge_id=bge_id) | Q(assigned_group__members__id=bge_id)).distinct()

        agg = qs.aggregate(
            total_investment_needed=Sum('investment_needed'),
            total_annual_revenue=Sum('annual_revenue'),
        )

        # Total employees from the latest growth snapshot per MSME in scope.
        # 1) Find the most recent snapshot date per MSME.
        # 2) Pull employee columns from those rows and sum them.
        msme_ids = list(qs.values_list('id', flat=True))
        latest_snapshot_ids = (
            MSMEGrowthSnapshot.objects
            .filter(msme_id__in=msme_ids)
            .order_by('msme_id', '-snapshot_date')
            .distinct('msme_id')
            .values_list('id', flat=True)
        )
        snap_emp_agg = MSMEGrowthSnapshot.objects.filter(id__in=latest_snapshot_ids).aggregate(
            ft_male   = Sum('employees_ft_male'),
            ft_female = Sum('employees_ft_female'),
            pt_male   = Sum('employees_pt_male'),
            pt_female = Sum('employees_pt_female'),
            ft_refugee = Sum('employees_ft_refugee'),
            pt_refugee = Sum('employees_pt_refugee'),
        )
        snapshot_employees = {
            'ft_male':    snap_emp_agg['ft_male']    or 0,
            'ft_female':  snap_emp_agg['ft_female']  or 0,
            'pt_male':    snap_emp_agg['pt_male']    or 0,
            'pt_female':  snap_emp_agg['pt_female']  or 0,
            'ft_refugee': snap_emp_agg['ft_refugee'] or 0,
            'pt_refugee': snap_emp_agg['pt_refugee'] or 0,
        }
        total_employees_from_snapshots = sum(snapshot_employees.values())

        # Reports / activity stats
        from ..models import MSMEReport, GroupReport
        report_status_stats = list(
            MSMEReport.objects.values('status').annotate(count=Count('id'))
        )
        group_report_status_stats = list(
            GroupReport.objects.values('status').annotate(count=Count('id'))
        )

        # BGE workload (top 15 BGEs by direct + group-assigned MSMEs)
        # Annotate everything in 2 queries instead of (1 + 3*N) per BGE.
        bge_qs = (BusinessGrowthExpert.objects
                  .filter(status='approved')
                  .annotate(
                      direct=Count(
                          'assigned_msmes',
                          filter=Q(assigned_msmes__is_active=True),
                          distinct=True,
                      ),
                      reports_count=Count('reports', distinct=True),
                  ))
        # `via_group` = MSMEs reachable through any of the BGE's groups.
        # Cheaper as a single grouped query than per-BGE.
        via_group_counts = dict(
            MSME.objects.filter(is_active=True, assigned_group__members__isnull=False)
                        .values_list('assigned_group__members')
                        .annotate(c=Count('id', distinct=True))
                        .values_list('assigned_group__members', 'c')
        )
        bge_workload = []
        for bge in bge_qs[:50]:
            via = via_group_counts.get(bge.id, 0)
            bge_workload.append({
                'bge_id':       bge.id,
                'bge_name':     bge.name,
                'direct':       bge.direct,
                'via_group':    via,
                'total':        bge.direct + via,
                'reports_count': bge.reports_count,
            })
        bge_workload.sort(key=lambda x: x['total'], reverse=True)
        bge_workload = bge_workload[:15]

        # Group performance — annotate in one query, no per-row counts.
        group_qs = (BGEGroup.objects
                    .select_related('team_lead')
                    .annotate(
                        msme_count=Count(
                            'assigned_msmes',
                            filter=Q(assigned_msmes__is_active=True),
                            distinct=True,
                        ),
                        active_member_count=Count('members', distinct=True),
                        reports_count=Count('reports', distinct=True),
                    ))
        group_stats = [
            {
                'group_id':       g.id,
                'group_name':     g.name,
                'msme_count':     g.msme_count,
                'member_count':   g.active_member_count,
                'reports_count':  g.reports_count,
                'team_lead_name': g.team_lead.name if g.team_lead else None,
            }
            for g in group_qs
        ]
        group_stats.sort(key=lambda x: x['msme_count'], reverse=True)

        # Time series — MSMEs created per month (last 18 months)
        time_series = list(
            qs.annotate(month=TruncMonth('created_at'))
              .values('month')
              .annotate(count=Count('id'))
              .order_by('month')
        )

        # Gender × Business Type cross-tab — for stacked bar
        gender_x_type = list(
            qs.values('gender', 'business_type')
              .annotate(count=Count('id'))
              .order_by('business_type', 'gender')
        )

        # ── Diagnostic analytics ──────────────────────────────────────────────
        diag_qs = qs.filter(diag_imported_at__isnull=False)
        diag_total = diag_qs.count()

        # Compliance & financial access flags
        diag_compliance = {
            'has_tin':           diag_qs.filter(diag_has_tin=True).count(),
            'has_ursb':          diag_qs.filter(diag_has_unbs=True).count(),
            'has_business_bank': diag_qs.filter(diag_has_business_bank=True).count(),
            'has_mobile_money':  diag_qs.filter(diag_has_mobile_money=True).count(),
            'total':             diag_total,
        }

        # Green business breakdown
        green_count = diag_qs.filter(diag_is_green_business=True).count()
        diag_green = {'green': green_count, 'non_green': diag_total - green_count, 'total': diag_total}

        # Turnover bands (count per band)
        diag_turnover_bands = list(
            diag_qs.exclude(diag_annual_turnover='')
                   .values('diag_annual_turnover')
                   .annotate(count=Count('id'))
                   .order_by('diag_annual_turnover')
        )

        # Total assets bands
        diag_asset_bands = list(
            diag_qs.exclude(diag_total_assets='')
                   .values('diag_total_assets')
                   .annotate(count=Count('id'))
                   .order_by('diag_total_assets')
        )

        # Employee aggregates (only rows that have values)
        from django.db.models import Sum as DSum
        emp_agg = diag_qs.aggregate(
            ft_male=DSum('diag_employees_ft_male'),
            ft_female=DSum('diag_employees_ft_female'),
            pt_male=DSum('diag_employees_pt_male'),
            pt_female=DSum('diag_employees_pt_female'),
        )
        diag_employees = {
            'ft_male':   emp_agg['ft_male']   or 0,
            'ft_female': emp_agg['ft_female'] or 0,
            'pt_male':   emp_agg['pt_male']   or 0,
            'pt_female': emp_agg['pt_female'] or 0,
        }

        # Owner sex from diagnostic
        diag_owner_sex = list(
            diag_qs.exclude(diag_owner_sex='')
                   .values('diag_owner_sex')
                   .annotate(count=Count('id'))
        )

        # Years operating distribution
        diag_years_operating = list(
            diag_qs.exclude(diag_years_operating='')
                   .values('diag_years_operating')
                   .annotate(count=Count('id'))
                   .order_by('diag_years_operating')
        )

        # Top districts from diagnostic (often more complete than `state`)
        diag_districts = list(
            diag_qs.exclude(diag_district='')
                   .values('diag_district')
                   .annotate(count=Count('id'))
                   .order_by('-count')[:12]
        )

        # MSMEs with at least one data update (growth snapshot)
        msmes_with_updates = (
            MSMEGrowthSnapshot.objects
            .filter(msme_id__in=msme_ids)
            .values('msme_id')
            .distinct()
            .count()
        )

        # BGEs active in the last 30 days (filed a snapshot or MSME report)
        from datetime import timedelta
        from django.utils import timezone
        cutoff = timezone.now() - timedelta(days=30)
        bges_via_snapshots = set(
            MSMEGrowthSnapshot.objects
            .filter(created_at__gte=cutoff, collected_by__isnull=False)
            .values_list('collected_by_id', flat=True)
        )
        from ..models import MSMEReport
        bges_via_reports = set(
            MSMEReport.objects
            .filter(created_at__gte=cutoff)
            .values_list('bge_id', flat=True)
        )
        active_bges_30d = len(bges_via_snapshots | bges_via_reports)

        return Response({
            # KPIs
            'total_msmes': qs.count(),
            'total_investment_needed': agg['total_investment_needed'] or 0,
            'total_annual_revenue':    agg['total_annual_revenue']    or 0,
            'total_employees':         total_employees_from_snapshots,
            'snapshot_employees':      snapshot_employees,
            'msmes_with_updates':      msmes_with_updates,
            'active_bges_30d':         active_bges_30d,
            'total_bges':              BusinessGrowthExpert.objects.count(),
            'total_groups':            BGEGroup.objects.count(),
            'total_reports':           MSMEReport.objects.count(),
            'total_group_reports':     GroupReport.objects.count(),

            # Distributions
            'business_type_stats': list(qs.values('business_type').annotate(count=Count('id'))),
            'sector_stats':        list(qs.values('sector').annotate(count=Count('id'))),
            'cohort_stats':        list(qs.values('cohort__name', 'cohort_id').annotate(count=Count('id')).order_by('cohort__name')),
            'gender_stats':        list(qs.values('gender').annotate(count=Count('id'))),
            'top_districts':       list(qs.values('state').exclude(state='').annotate(count=Count('id')).order_by('-count')[:10]),
            'top_cities':          list(qs.values('city').exclude(city='').annotate(count=Count('id')).order_by('-count')[:10]),

            # Cross-tabs
            'gender_x_type': gender_x_type,

            # BGE / Group performance
            'bge_status_stats':    list(BusinessGrowthExpert.objects.values('status').annotate(count=Count('id'))),
            'bge_workload':        bge_workload,
            'group_stats':         group_stats,

            # Reports
            'report_status_stats': report_status_stats,
            'group_report_status_stats': group_report_status_stats,

            # Time series
            'time_series':         time_series,

            # Diagnostic baseline analytics
            'diag_total':           diag_total,
            'diag_compliance':      diag_compliance,
            'diag_green':           diag_green,
            'diag_turnover_bands':  diag_turnover_bands,
            'diag_asset_bands':     diag_asset_bands,
            'diag_employees':       diag_employees,
            'diag_owner_sex':       diag_owner_sex,
            'diag_years_operating': diag_years_operating,
            'diag_districts':       diag_districts,
        })

    @action(detail=False, methods=['get'], url_path='inactive')
    def inactive_msmes(self, request):
        """Admin-only: list all inactive MSMEs so they can be reviewed before reactivation."""
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can view inactive MSMEs.")
        qs = MSME.objects.filter(is_active=False).order_by('business_name')
        data = list(qs.values('id', 'business_name', 'msme_code', 'owner_name',
                              'sector', 'city', 'cohort_id'))
        return Response({'count': len(data), 'msmes': data})

    @action(detail=False, methods=['post'], url_path='reactivate-all')
    def reactivate_all(self, request):
        """Admin-only: set is_active=True for every inactive MSME."""
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can reactivate MSMEs.")
        updated = MSME.objects.filter(is_active=False).update(is_active=True)
        return Response({'reactivated': updated,
                         'message': f'{updated} MSME(s) reactivated successfully.'})

    @action(detail=False, methods=['post'], url_path='import-diagnostics')
    def import_diagnostics(self, request):
        """Admin-only: upload the diagnostics Excel and run the import in-process.
        POST /api/msmes/import-diagnostics/  multipart field: file
        Returns a JSON summary of matched/unmatched counts.
        """
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can import diagnostic data.")

        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'error': 'No file provided. Send multipart field "file".'}, status=400)

        import tempfile, os, io as _io, traceback as _tb
        from django.core.management import call_command

        suffix = os.path.splitext(uploaded.name)[1] or '.xlsx'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            for chunk in uploaded.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            out = _io.StringIO()
            call_command('import_diagnostics', tmp_path, stdout=out, stderr=out)
            output = out.getvalue()
        except Exception as exc:
            os.unlink(tmp_path)
            return Response({'error': f'{type(exc).__name__}: {exc}\n\n{_tb.format_exc()}'}, status=500)
        finally:
            try:
                os.unlink(tmp_path)
            except FileNotFoundError:
                pass

        return Response({'detail': output or 'Import complete.'})
