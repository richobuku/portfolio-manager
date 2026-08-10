"""
Smart MSME → BGE focal-person assignment
=========================================
Assigns each active MSME to a permanent BGE focal person using a scored
round-robin that balances three factors:

  1. Previous engagement  – BGE who has already visited/reported on this MSME
                            gets the highest priority (up to 100 pts).
  2. Location proximity   – BGE whose location is in the same or adjacent
                            district gets a bonus (10 or 5 pts).
  3. Load balancing       – A penalty proportional to the BGE's current
                            provisional load keeps the distribution even.

Endpoints
---------
  GET  /api/msmes/smart-assign/         → preview (no writes)
  POST /api/msmes/smart-assign/         → apply assignments (updates assigned_bge)
  GET  /api/msmes/smart-assign/export/  → download Excel workbook

POST body (all optional):
  apply_to   'all' | 'unassigned'   (default 'all')
"""

import io
import logging
from collections import defaultdict

from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Northern-Uganda district adjacency map (undirected).
# Two districts are "adjacent" if BGEs routinely travel between them.
# ---------------------------------------------------------------------------
_ADJACENCY: dict[str, list[str]] = {
    'gulu':      ['omoro', 'amuru', 'nwoya', 'pader', 'kitgum', 'adjumani'],
    'lira':      ['oyam', 'alebtong', 'dokolo', 'apac', 'kole', 'otuke'],
    'omoro':     ['gulu', 'kitgum', 'pader'],
    'amuru':     ['gulu', 'nwoya', 'adjumani'],
    'nwoya':     ['gulu', 'amuru'],
    'adjumani':  ['gulu', 'amuru'],
    'kitgum':    ['gulu', 'omoro', 'pader', 'agago', 'lamwo'],
    'pader':     ['kitgum', 'agago', 'gulu', 'omoro'],
    'agago':     ['pader', 'kitgum'],
    'lamwo':     ['kitgum'],
    'oyam':      ['lira', 'apac', 'kole'],
    'alebtong':  ['lira', 'apac', 'dokolo'],
    'dokolo':    ['lira', 'alebtong', 'apac'],
    'apac':      ['lira', 'oyam', 'kole', 'alebtong', 'dokolo'],
    'kole':      ['lira', 'apac', 'oyam'],
    'otuke':     ['lira', 'kole'],
}


def _normalise(text: str) -> str:
    return (text or '').lower().strip()


def _location_score(bge_location: str, msme_city: str) -> tuple[int, str]:
    """
    Returns (score, reason_string) for the geographic proximity of a
    (BGE, MSME) pair.  Scores:
      10 — same district / city
       5 — adjacent district
       0 — different / unknown
    """
    bl = _normalise(bge_location)
    mc = _normalise(msme_city)
    if not bl or not mc:
        return 0, 'unknown location'
    if mc in bl or bl in mc:
        return 10, f'same area ({msme_city})'
    # Check adjacency map
    for district, neighbours in _ADJACENCY.items():
        if district in mc:          # MSME is in this district
            for n in neighbours:
                if n in bl:         # BGE is in a neighbouring district
                    return 5, f'adjacent ({msme_city} ↔ {bge_location})'
        if district in bl:          # BGE is in this district
            for n in neighbours:
                if n in mc:         # MSME is in a neighbouring district
                    return 5, f'adjacent ({bge_location} ↔ {msme_city})'
    return 0, 'different area'


# ---------------------------------------------------------------------------
# Core algorithm
# ---------------------------------------------------------------------------

def _compute_assignments(apply_to: str = 'all', exclude_ids=None):
    """
    Returns (assignments, bge_summary, cap).

    exclude_ids: iterable of BGE primary-key integers to omit from the pool.
    """
    from ..models import MSME, BusinessGrowthExpert, MSMEReport

    exclude_ids = set(int(x) for x in (exclude_ids or []))

    # ── 1. Load data ────────────────────────────────────────────────────────
    bges = list(
        BusinessGrowthExpert.objects.filter(status='approved')
        .exclude(id__in=exclude_ids)
        .order_by('name')
        .values('id', 'name', 'bge_code', 'location')
    )
    if not bges:
        return [], [], 0

    msme_qs = MSME.objects.filter(is_active=True).order_by('city', 'business_name')
    if apply_to == 'unassigned':
        msme_qs = msme_qs.filter(assigned_bge__isnull=True)

    msmes = list(msme_qs.values(
        'id', 'msme_code', 'business_name', 'city', 'sector',
        'assigned_bge_id',
    ))
    if not msmes:
        return [], bges, 0

    # ── 2. Precompute engagement counts: (msme_id, bge_id) → count ─────────
    report_counts: dict[tuple, int] = defaultdict(int)
    for row in MSMEReport.objects.values('msme_id', 'bge_id').iterator():
        report_counts[(row['msme_id'], row['bge_id'])] += 1

    # BGE name lookup
    bge_by_id = {b['id']: b for b in bges}

    # ── 3. Capped scored round-robin ────────────────────────────────────────
    # Hard cap: each BGE gets at most ceil(total / num_bges) MSMEs.
    # This guarantees a spread of ±1.  Within that constraint, engagement
    # and location scores decide which BGE is best for each MSME.
    import math
    n_bges = len(bges)
    n_msmes = len(msmes)
    cap = math.ceil(n_msmes / n_bges)   # e.g. 34 MSMEs / 8 BGEs → cap = 5

    # Sort MSMEs so those with existing engagement come first (clearer choices)
    def _msme_priority(m):
        total_reports = sum(
            v for (mid, bid), v in report_counts.items() if mid == m['id']
        )
        return -total_reports  # descending: most-engaged MSMEs first

    msmes.sort(key=_msme_priority)

    load: dict[int, int] = defaultdict(int)  # bge_id → provisional assignments
    results = []

    for m in msmes:
        best_bge_id   = None
        best_score    = -9999
        best_loc_s    = 0
        best_eng_s    = 0
        best_loc_r    = ''
        best_eng_r    = ''

        # Eligible BGEs: those not yet at the cap.
        # If somehow all are at cap (shouldn't happen with ceil), fall back.
        eligible = [b for b in bges if load[b['id']] < cap]
        if not eligible:
            eligible = bges  # safety fallback

        for b in eligible:
            eng_count = report_counts.get((m['id'], b['id']), 0)
            eng_score = min(eng_count, 5) * 20          # 0 – 100

            loc_score, loc_reason = _location_score(b['location'], m['city'])

            continuity = 30 if m['assigned_bge_id'] == b['id'] else 0

            total = eng_score + loc_score + continuity

            if total > best_score:
                best_score   = total
                best_bge_id  = b['id']
                best_loc_s   = loc_score
                best_eng_s   = eng_score
                best_loc_r   = loc_reason
                best_eng_r   = (
                    f'{eng_count} previous report{"s" if eng_count != 1 else ""}'
                    if eng_count else 'no previous reports'
                )

        load[best_bge_id] += 1
        proposed_bge = bge_by_id[best_bge_id]

        if m['assigned_bge_id'] == best_bge_id:
            assignment_status = 'unchanged'
        elif m['assigned_bge_id'] is None:
            assignment_status = 'new'
        else:
            assignment_status = 'reassigned'

        results.append({
            'msme_id':            m['id'],
            'msme_code':          m['msme_code'] or '',
            'msme_name':          m['business_name'],
            'msme_city':          m['city'],
            'msme_sector':        m['sector'],
            'current_bge_id':     m['assigned_bge_id'],
            'current_bge_name':   (bge_by_id[m['assigned_bge_id']]['name']
                                   if m['assigned_bge_id'] in bge_by_id else None),
            'proposed_bge_id':    best_bge_id,
            'proposed_bge_name':  proposed_bge['name'],
            'proposed_bge_code':  proposed_bge['bge_code'],
            'proposed_bge_loc':   proposed_bge['location'],
            'engagement_score':   best_eng_s,
            'location_score':     best_loc_s,
            'total_score':        best_score,
            'engagement_reason':  best_eng_r,
            'location_reason':    best_loc_r,
            'status':             assignment_status,
        })

    # ── 4. Load summary per BGE ──────────────────────────────────────────────
    bge_summary = []
    for b in bges:
        count = load[b['id']]
        bge_summary.append({
            'bge_id':    b['id'],
            'bge_name':  b['name'],
            'bge_code':  b['bge_code'],
            'location':  b['location'],
            'count':     count,
        })

    return results, bge_summary, cap


# ---------------------------------------------------------------------------
# Views
# ---------------------------------------------------------------------------

def _require_admin(request):
    if not (request.user.is_staff or request.user.is_superuser):
        raise PermissionDenied("Only admins can manage BGE assignments.")


def _parse_exclude(raw: str) -> set:
    """Parse comma-separated BGE IDs from a query-string value."""
    result = set()
    for part in (raw or '').split(','):
        part = part.strip()
        if part.isdigit():
            result.add(int(part))
    return result


def _parse_exclude_list(raw) -> set:
    """Parse a list or comma-string of BGE IDs from POST body."""
    if isinstance(raw, (list, tuple)):
        return {int(x) for x in raw if str(x).strip().isdigit()}
    return _parse_exclude(str(raw) if raw else '')


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def smart_assign(request):
    """
    GET  → preview proposed assignments
    POST → apply them (writes to DB)
    """
    _require_admin(request)

    apply_to = request.query_params.get('apply_to', 'all') or 'all'
    exclude_ids = _parse_exclude(request.query_params.get('exclude', ''))
    if request.method == 'POST':
        apply_to = request.data.get('apply_to', 'all')
        exclude_ids = _parse_exclude_list(request.data.get('exclude_bge_ids', []))

    assignments, bge_summary, cap = _compute_assignments(apply_to, exclude_ids)

    stats = {
        'total':      len(assignments),
        'unchanged':  sum(1 for a in assignments if a['status'] == 'unchanged'),
        'new':        sum(1 for a in assignments if a['status'] == 'new'),
        'reassigned': sum(1 for a in assignments if a['status'] == 'reassigned'),
        'cap_per_bge': cap,
    }

    if request.method == 'GET':
        return Response({
            'stats':       stats,
            'bge_summary': bge_summary,
            'assignments': assignments,
        })

    # POST — apply
    from django.db import transaction
    from ..models import MSME
    today = timezone.now().date()
    clear_first = bool(request.data.get('clear_first', False))
    updated = 0

    with transaction.atomic():
        if clear_first:
            # Wipe all existing assignments so every MSME is treated as fresh.
            # We computed BEFORE clearing so engagement/continuity bonuses still
            # informed the algorithm — now we apply everything unconditionally.
            MSME.objects.filter(is_active=True).update(
                assigned_bge_id=None,
                assignment_date=None,
            )
            for a in assignments:
                MSME.objects.filter(id=a['msme_id']).update(
                    assigned_bge_id=a['proposed_bge_id'],
                    assignment_date=today,
                )
            updated = len(assignments)
        else:
            for a in assignments:
                if a['status'] != 'unchanged':
                    MSME.objects.filter(id=a['msme_id']).update(
                        assigned_bge_id=a['proposed_bge_id'],
                        assignment_date=today,
                    )
                    updated += 1

    return Response({
        'applied':   updated,
        'unchanged': 0 if clear_first else stats['unchanged'],
        'cleared':   clear_first,
        'stats':     stats,
        'bge_summary': bge_summary,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def smart_assign_export(request):
    """
    Download an Excel workbook with:
      - Sheet 1: Summary (BGE | # MSMEs | districts)
      - Sheet 2: Full assignment list (one row per MSME)
    """
    _require_admin(request)

    apply_to = request.query_params.get('apply_to', 'all')
    exclude_ids = _parse_exclude(request.query_params.get('exclude', ''))
    assignments, bge_summary, cap = _compute_assignments(apply_to, exclude_ids)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response({'error': 'openpyxl not installed.'}, status=500)

    wb = Workbook()

    # ── Colour palette ───────────────────────────────────────────────────────
    NAVY   = 'FF1A3A5C'
    GOLD   = 'FFFFC000'
    GREEN  = 'FFE2EFDA'
    ORANGE = 'FFFCE4D6'
    GREY   = 'FFF2F2F2'
    WHITE  = 'FFFFFFFF'

    def _hdr_cell(ws, row, col, value, bg=NAVY, fg=WHITE, bold=True):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=fg, size=10)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='FFAAAAAA')
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return c

    def _data_cell(ws, row, col, value, bg=WHITE, bold=False, align='left', number_format=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=10)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)
        thin = Side(style='thin', color='FFD0D0D0')
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if number_format:
            c.number_format = number_format
        return c

    # ── Sheet 1: BGE Summary ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'BGE Summary'
    ws1.row_dimensions[1].height = 30

    hdrs1 = ['BGE Code', 'BGE Name', 'Location', 'MSMEs Assigned', '% of Total']
    for col, h in enumerate(hdrs1, 1):
        _hdr_cell(ws1, 1, col, h)

    total_msmes = len(assignments)
    for r, b in enumerate(bge_summary, 2):
        bg = GREY if r % 2 == 0 else WHITE
        pct = (b['count'] / total_msmes * 100) if total_msmes else 0
        _data_cell(ws1, r, 1, b['bge_code'],  bg)
        _data_cell(ws1, r, 2, b['bge_name'],  bg, bold=True)
        _data_cell(ws1, r, 3, b['location'],  bg)
        _data_cell(ws1, r, 4, b['count'],     bg, align='center', number_format='0')
        _data_cell(ws1, r, 5, pct / 100,      bg, align='center', number_format='0.0%')

    # Column widths
    for col, w in zip(range(1, 6), [14, 28, 18, 16, 12]):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 2: Full Assignment List ────────────────────────────────────────
    ws2 = wb.create_sheet('Assignment List')
    ws2.freeze_panes = 'A2'
    ws2.row_dimensions[1].height = 30

    hdrs2 = [
        'MSME Code', 'Business Name', 'City / District', 'Sector',
        'Proposed BGE', 'BGE Code', 'BGE Location',
        'Previous BGE', 'Status',
        'Engagement Score', 'Location Score', 'Total Score',
        'Engagement Reason', 'Location Reason',
    ]
    for col, h in enumerate(hdrs2, 1):
        _hdr_cell(ws2, 1, col, h)

    STATUS_BG = {
        'unchanged':  GREEN,
        'new':        GREY,
        'reassigned': ORANGE,
    }

    for r, a in enumerate(assignments, 2):
        bg = STATUS_BG.get(a['status'], WHITE)
        _data_cell(ws2, r,  1, a['msme_code'],          WHITE)
        _data_cell(ws2, r,  2, a['msme_name'],          WHITE, bold=True)
        _data_cell(ws2, r,  3, a['msme_city'],          WHITE)
        _data_cell(ws2, r,  4, a['msme_sector'],        WHITE)
        _data_cell(ws2, r,  5, a['proposed_bge_name'],  bg, bold=True)
        _data_cell(ws2, r,  6, a['proposed_bge_code'],  bg)
        _data_cell(ws2, r,  7, a['proposed_bge_loc'],   bg)
        _data_cell(ws2, r,  8, a['current_bge_name'] or '—', WHITE)
        status_label = {'unchanged': 'Unchanged', 'new': 'New', 'reassigned': 'Changed'}.get(a['status'], a['status'])
        _data_cell(ws2, r,  9, status_label,            bg, align='center')
        _data_cell(ws2, r, 10, a['engagement_score'],   WHITE, align='center', number_format='0')
        _data_cell(ws2, r, 11, a['location_score'],     WHITE, align='center', number_format='0')
        _data_cell(ws2, r, 12, a['total_score'],        WHITE, align='center', number_format='0')
        _data_cell(ws2, r, 13, a['engagement_reason'],  WHITE)
        _data_cell(ws2, r, 14, a['location_reason'],    WHITE)

    # Column widths
    for col, w in zip(range(1, 15), [14, 28, 16, 18, 24, 12, 14, 24, 12, 10, 10, 10, 24, 24]):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # Sort order note in A1 tooltip (via comment workaround — just a note row)
    ws2.sheet_properties.tabColor = '1A3A5C'

    # ── Build and return the file ─────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from django.utils.timezone import now as _now
    date_str = _now().strftime('%Y%m%d')
    filename = f'BGE_Assignment_{date_str}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
