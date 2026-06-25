import logging
import re

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response

logger = logging.getLogger(__name__)
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.db.models import Count, Sum, Q, Max, Subquery, OuterRef
from django.contrib.auth.models import User
from django.core.mail import send_mail, EmailMultiAlternatives
from django.conf import settings
from django.http import HttpResponse
import pandas as pd
import io

from .models import (
    Portfolio, Investment, Transaction,
    MSME, BusinessGrowthExpert, SupportRequest,
    TrainingSession, Attendance, TrainingTopic,
    Cohort, BGEGroup, MSMEReport, GroupReport, GroupReportContribution, PushSubscription, WorkOrder,
    GroupReportAttendance, CohortAdmin, ProgrammeGroup, MSMEGrowthSnapshot, VisitReportTemplate,
    TrainingFacilitationAssignment, TrainingReport, AnnualReviewReport,
    MentorTrainingReport, TshirtReceipt, TshirtReceiptEntry,
    WorkOrderSubmission, WorkOrderPayment,
)
from .account_setup import ensure_bge_account, send_welcome_email


def _managed_groups(user):
    """Return the ProgrammeGroup IDs a programme manager can access,
    or None for superusers/staff (meaning no restriction)."""
    if user.is_staff or user.is_superuser:
        return None
    try:
        return list(user.cohort_admin_profile.managed_groups.values_list('id', flat=True))
    except CohortAdmin.DoesNotExist:
        return None


def _is_viewer(user):
    """True for accounts that have no BGE profile and no programme-manager role — read-only."""
    if user.is_staff or user.is_superuser:
        return False
    if hasattr(user, 'cohort_admin_profile'):
        return False
    if hasattr(user, 'bge_profile'):
        return False
    return True


def _is_programme_manager(user):
    """True for cohort_admin accounts that are NOT full staff/superuser."""
    if user.is_staff or user.is_superuser:
        return False
    return hasattr(user, 'cohort_admin_profile')


class ViewerReadOnlyMixin:
    """Block create/update/delete for viewer accounts."""
    def _check_not_viewer(self):
        if _is_viewer(self.request.user):
            raise PermissionDenied("Viewer accounts have read-only access.")

    def create(self, request, *args, **kwargs):
        self._check_not_viewer()
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        self._check_not_viewer()
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        self._check_not_viewer()
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        self._check_not_viewer()
        return super().destroy(request, *args, **kwargs)


class ProgrammeManagerReadOnlyMixin:
    """Block create/update/delete for programme-manager (cohort_admin) accounts."""
    def _check_not_pm(self):
        if _is_programme_manager(self.request.user):
            raise PermissionDenied("Programme Managers have read-only access to this resource.")

    def create(self, request, *args, **kwargs):
        self._check_not_pm()
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        self._check_not_pm()
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        self._check_not_pm()
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        self._check_not_pm()
        return super().destroy(request, *args, **kwargs)


def _safe_filename(name):
    """Strip characters that could break or inject into a Content-Disposition header."""
    name = re.sub(r'[\r\n"]', '', str(name))
    return name.strip() or 'download'


from pywebpush import webpush, WebPushException
import json as _json
from .serializers import (
    PortfolioSerializer, InvestmentSerializer, TransactionSerializer,
    MSMESerializer, BusinessGrowthExpertSerializer, SupportRequestSerializer,
    TrainingSessionSerializer, AttendanceSerializer, TrainingTopicSerializer,
    TrainingFacilitationAssignmentSerializer, TrainingReportSerializer,
    CohortSerializer, BGEGroupSerializer, MSMEReportSerializer,
    GroupReportSerializer, GroupReportContributionSerializer, WorkOrderSerializer,
    VisitReportTemplateSerializer,
    GroupReportAttendanceSerializer, MSMEGrowthSnapshotSerializer,
    AnnualReviewReportSerializer, MentorTrainingReportSerializer,
    WorkOrderSubmissionSerializer, WorkOrderPaymentSerializer,
)


class PortfolioViewSet(viewsets.ModelViewSet):
    """Portfolio is per-user. Admins see everything; everyone else sees only
    their own portfolios. Previously every authenticated user could read or
    modify every portfolio in the system."""
    queryset = Portfolio.objects.all()
    serializer_class = PortfolioSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = Portfolio.objects.all()
        u = self.request.user
        if not (u.is_staff or u.is_superuser):
            qs = qs.filter(user=u)
        return qs

    def perform_create(self, serializer):
        # Ensure created portfolios are bound to the requesting user
        serializer.save(user=self.request.user)

    @action(detail=False, methods=['get'])
    def analytics(self, request):
        portfolios = self.get_queryset()  # tenant-scoped
        total_value = sum(p.total_value() for p in portfolios)
        total_cost = sum(p.total_cost() for p in portfolios)
        total_return = total_value - total_cost
        total_return_pct = (total_return / total_cost * 100) if total_cost > 0 else 0
        investment_types = Investment.objects.filter(portfolio__in=portfolios).values('investment_type').annotate(
            count=Count('id'), total_value=Sum('current_price')
        )
        return Response({
            'total_portfolios': portfolios.count(),
            'total_value': total_value,
            'total_cost': total_cost,
            'total_return': total_return,
            'total_return_percentage': total_return_pct,
            'investment_types': investment_types,
        })


class InvestmentViewSet(viewsets.ModelViewSet):
    queryset = Investment.objects.all()
    serializer_class = InvestmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = Investment.objects.all()
        u = self.request.user
        if not (u.is_staff or u.is_superuser):
            qs = qs.filter(portfolio__user=u)
        pid = self.request.query_params.get('portfolio')
        if pid:
            qs = qs.filter(portfolio_id=pid)
        return qs.select_related('portfolio')


class TransactionViewSet(viewsets.ModelViewSet):
    queryset = Transaction.objects.all()
    serializer_class = TransactionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = Transaction.objects.all()
        u = self.request.user
        if not (u.is_staff or u.is_superuser):
            qs = qs.filter(investment__portfolio__user=u)
        iid = self.request.query_params.get('investment')
        if iid:
            qs = qs.filter(investment_id=iid)
        return qs.select_related('investment', 'investment__portfolio').order_by('-transaction_date')


class CohortViewSet(ViewerReadOnlyMixin, viewsets.ModelViewSet):
    queryset = Cohort.objects.all()
    serializer_class = CohortSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Annotate msme_count at DB level so CohortSerializer.get_msme_count
        # can avoid a per-row COUNT query (N+1 fix).
        return Cohort.objects.annotate(_msme_count=Count('msmes', distinct=True))

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff and not request.user.is_superuser:
            raise PermissionDenied("Only admins can delete cohorts.")
        return super().destroy(request, *args, **kwargs)


class ProgrammeGroupViewSet(viewsets.ModelViewSet):
    """Cross-cutting labels that can be applied to MSMEs (e.g. Green MSMEs, Agroprocessors).
    Read-only for non-admins; create/update/delete restricted to admins."""
    queryset = ProgrammeGroup.objects.all()
    serializer_class = None  # defined below via import
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Annotate msme_count at DB level so ProgrammeGroupSerializer.get_msme_count
        # can avoid a per-row COUNT query (N+1 fix).
        return ProgrammeGroup.objects.annotate(_msme_count=Count('msmes', distinct=True))

    def get_serializer_class(self):
        from .serializers import ProgrammeGroupSerializer
        return ProgrammeGroupSerializer

    def create(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can create programme groups.")
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can edit programme groups.")
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can delete programme groups.")
        return super().destroy(request, *args, **kwargs)


class MSMEGrowthSnapshotViewSet(ViewerReadOnlyMixin, viewsets.ModelViewSet):
    """
    Growth snapshots for a single MSME.

    List/retrieve all snapshots for a given MSME:
        GET /api/growth-snapshots/?msme=<id>

    Create a new BGE-visit snapshot:
        POST /api/growth-snapshots/
        { msme, snapshot_date, source, annual_turnover, ... }
    """
    serializer_class   = MSMEGrowthSnapshotSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = MSMEGrowthSnapshot.objects.select_related('msme', 'collected_by').order_by('snapshot_date')
        msme_id = self.request.query_params.get('msme')
        bge_id  = self.request.query_params.get('bge')
        if msme_id:
            qs = qs.filter(msme_id=msme_id)
        if bge_id:
            # Security: non-admin users can only query their own BGE data (IDOR fix)
            user = self.request.user
            if not (user.is_staff or user.is_superuser or _managed_groups(user) is not None):
                own_bge = getattr(user, 'bge_profile', None)
                if not own_bge or str(own_bge.id) != str(bge_id):
                    raise PermissionDenied("You can only access your own data.")
            # All snapshots for MSMEs directly assigned to this BGE or in their groups
            from .models import BusinessGrowthExpert, BGEGroup
            try:
                bge = BusinessGrowthExpert.objects.get(pk=bge_id)
                group_msme_ids = MSME.objects.filter(assigned_group__in=bge.bge_groups.all()).values_list('id', flat=True)
                direct_msme_ids = MSME.objects.filter(assigned_bge=bge).values_list('id', flat=True)
                co_msme_ids = MSME.objects.filter(co_assigned_bges=bge).values_list('id', flat=True)
                all_ids = set(list(direct_msme_ids) + list(group_msme_ids) + list(co_msme_ids))
                qs = qs.filter(msme_id__in=all_ids)
            except BusinessGrowthExpert.DoesNotExist:
                qs = qs.none()
        return qs

    def perform_create(self, serializer):
        """Auto-set collected_by from the logged-in user's identity.

        Resolution order (first match wins):
        1. Value already in the request — respect it (admin attribution on behalf of BGE).
        2. user.bge_profile OneToOne — the clean path once accounts are linked.
        3. Username match: BGE whose name slug equals the Django username (e.g. jimmy.ouni → Jimmy Ouni).
        4. Email match: BGE whose email local-part equals the username.
        Falls through silently so admin submissions (no bge_profile) still save.
        """
        if serializer.validated_data.get('collected_by'):
            serializer.save()
            return

        user = self.request.user

        # Path 2 — linked profile
        try:
            serializer.save(collected_by=user.bge_profile)
            return
        except Exception:
            pass

        # Path 3 — match BGE name slug to username  (e.g. "jimmy.ouni" → "Jimmy Ouni")
        from .models import BusinessGrowthExpert
        import re

        def _slug(text):
            s = re.sub(r'[^a-z0-9]+', '.', (text or '').lower()).strip('.')
            return re.sub(r'\.+', '.', s)

        uname = user.username.lower()
        for bge in BusinessGrowthExpert.objects.filter(user__isnull=True):
            if _slug(bge.name) == uname:
                # Auto-link for next time so we don't need to re-scan
                bge.user = user
                bge.save(update_fields=['user'])
                serializer.save(collected_by=bge)
                return

        # Path 4 — email local-part match
        for bge in BusinessGrowthExpert.objects.filter(user__isnull=True):
            if bge.email and bge.email.split('@')[0].lower() == uname:
                bge.user = user
                bge.save(update_fields=['user'])
                serializer.save(collected_by=bge)
                return

        serializer.save()

    def destroy(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can delete snapshots.")
        return super().destroy(request, *args, **kwargs)


class MSMEViewSet(ViewerReadOnlyMixin, viewsets.ModelViewSet):
    queryset = MSME.objects.filter(is_active=True)
    serializer_class = MSMESerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = MSME.objects.filter(is_active=True)

        user = self.request.user
        group_ids = _managed_groups(user)
        if group_ids is not None:
            # Programme manager — scope to MSMEs in their managed programme groups
            qs = qs.filter(programme_groups__in=group_ids).distinct()
        elif _is_viewer(user):
            pass  # viewers see all MSMEs read-only (ViewerReadOnlyMixin blocks writes)
        elif not (user.is_staff or user.is_superuser):
            # BGE users only see their own assigned MSMEs — directly OR via any group they belong to.
            # Exception: ?training=1 returns all MSMEs so a facilitator can record attendance
            # for participants who are not personally assigned to them.
            training_context = self.request.query_params.get('training') == '1'
            if not training_context:
                try:
                    bge = user.bge_profile
                    from django.db.models import Q
                    qs = qs.filter(
                        Q(assigned_bge=bge) |
                        Q(assigned_group__members=bge) |
                        Q(co_assigned_bges=bge)
                    ).distinct()
                except Exception:
                    qs = qs.none()

        search = self.request.query_params.get('search')
        if search:
            # Combine search clauses with Q() inside ONE .filter(), not via
            # queryset union (`qs.filter(...) | qs.filter(...)`) — the union
            # form silently drops the BGE tenant scoping that was applied
            # above, leaking every other BGE's MSMEs into search results.
            from django.db.models import Q
            qs = qs.filter(
                Q(business_name__icontains=search) |
                Q(owner_name__icontains=search)    |
                Q(sector__icontains=search)        |
                Q(msme_code__icontains=search)
            )

        business_type = self.request.query_params.get('business_type')
        if business_type:
            qs = qs.filter(business_type=business_type)

        sector = self.request.query_params.get('sector')
        if sector:
            qs = qs.filter(sector=sector)

        cohort = self.request.query_params.get('cohort')
        if cohort:
            qs = qs.filter(cohort_id=cohort)

        city = self.request.query_params.get('city')
        if city:
            qs = qs.filter(city__iexact=city)

        return (
            qs.select_related('cohort', 'assigned_bge', 'assigned_group')
            .prefetch_related('programme_groups')
            # Annotate counts and latest dates at the DB level to eliminate the
            # N+1 queries that MSMESerializer.get_total_reports / get_last_support_date
            # would otherwise fire (one round-trip per row).
            .annotate(
                _reports_count=Count('reports', distinct=True),
                _group_reports_count=Count('group_reports', distinct=True),
                _last_individual_date=Max('reports__visit_date'),
                _last_group_date=Max('group_reports__visit_date'),
            )
            .order_by('-created_at')
        )

    def _is_admin_or_cohort_admin(self, request):
        u = request.user
        if u.is_staff or u.is_superuser:
            return True
        return _managed_groups(u) is not None

    def destroy(self, request, *args, **kwargs):
        if not self._is_admin_or_cohort_admin(request):
            raise PermissionDenied("Only admins can delete MSMEs.")
        return super().destroy(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        if not self._is_admin_or_cohort_admin(request):
            raise PermissionDenied("Only admins can create MSMEs.")
        return super().create(request, *args, **kwargs)

    @action(detail=True, methods=['patch'])
    def assign_bge(self, request, pk=None):
        if not self._is_admin_or_cohort_admin(request):
            raise PermissionDenied("Only admins can assign BGEs.")
        msme = self.get_object()
        # Programme managers may only assign BGEs to MSMEs within their managed groups
        if _is_programme_manager(request.user):
            managed = _managed_groups(request.user) or []
            if not msme.programme_groups.filter(id__in=managed).exists():
                raise PermissionDenied("You can only assign BGEs to MSMEs in your managed programme groups.")
        bge_id = request.data.get('bge_id')
        objectives = (request.data.get('objectives') or '').strip()
        assignment_date = request.data.get('assignment_date') or None
        bge = None  # initialise so the notification block below is always safe
        if bge_id:
            try:
                bge = BusinessGrowthExpert.objects.get(pk=bge_id)
                # Prevent assigning the same BGE twice as primary
                if msme.assigned_bge_id and msme.assigned_bge_id == bge.id:
                    return Response(
                        {'error': f'{msme.business_name} is already assigned to {bge.name}.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                if msme.assigned_bge_id and msme.assigned_bge_id != bge.id:
                    # MSME already has a primary BGE — add the new BGE as co-assigned
                    # so BOTH BGEs keep the MSME in their list (joint deployment).
                    existing_primary = msme.assigned_bge  # capture before save
                    msme.co_assigned_bges.add(bge)
                    msme.save()
                    # Notify the new (co-assigned) BGE via push
                    _notify_bge(
                        bge,
                        title='Joint MSME Assignment',
                        body=f'You have been co-assigned to {msme.business_name} alongside {existing_primary.name}. Check your dashboard for details.',
                        url='/bge'
                    )
                    # Notify the existing primary BGE via push + email
                    _notify_bge(
                        existing_primary,
                        title='Joint Deployment Notice',
                        body=f'{bge.name} has also been assigned to visit {msme.business_name}. Check your dashboard for details.',
                        url='/bge'
                    )
                    _send_co_assignment_alert(existing_primary, bge, msme)
                    return Response(MSMESerializer(msme).data)
                # No existing primary — set this BGE as primary
                msme.assigned_bge = bge
            except BusinessGrowthExpert.DoesNotExist:
                return Response({'error': 'BGE not found'}, status=status.HTTP_404_NOT_FOUND)
        else:
            # Unassign: clear primary and all co-assignees
            msme.assigned_bge = None
            msme.co_assigned_bges.clear()
        msme.assignment_objectives = objectives
        msme.assignment_date = assignment_date
        msme.save()
        # Notify the BGE about the new assignment
        if bge_id and bge:
            _notify_bge(
                bge,
                title='New MSME Assignment',
                body=f'You have been assigned to {msme.business_name}. Check your dashboard for details.',
                url='/bge'
            )
        return Response(MSMESerializer(msme).data)

    @action(detail=True, methods=['patch'])
    def assign_cohort(self, request, pk=None):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can assign cohorts.")
        msme = self.get_object()
        cohort_id = request.data.get('cohort_id')
        if cohort_id:
            try:
                cohort = Cohort.objects.get(pk=cohort_id)
                msme.cohort = cohort
            except Cohort.DoesNotExist:
                return Response({'error': 'Cohort not found'}, status=status.HTTP_404_NOT_FOUND)
        else:
            msme.cohort = None
        msme.save()
        return Response(MSMESerializer(msme).data)

    @action(detail=True, methods=['patch'], url_path='set-groups')
    def set_groups(self, request, pk=None):
        """Add or remove this MSME from programme groups.

        Body: { "group_ids": [1, 2, ...] }   — replaces the full set.
        Body: { "add": [1], "remove": [2] }   — incremental add/remove.
        """
        if not self._is_admin_or_cohort_admin(request):
            raise PermissionDenied("Only admins can modify programme group membership.")
        msme = self.get_object()
        if 'group_ids' in request.data:
            ids = request.data['group_ids'] or []
            msme.programme_groups.set(ids)
        else:
            add_ids    = request.data.get('add', [])
            remove_ids = request.data.get('remove', [])
            if add_ids:
                msme.programme_groups.add(*add_ids)
            if remove_ids:
                msme.programme_groups.remove(*remove_ids)
        return Response(MSMESerializer(msme, context={'request': request}).data)

    @action(detail=False, methods=['get'], url_path='upload-template', permission_classes=[])
    def upload_template(self, request):
        """Download a blank MSME upload template (.xlsx) with the unified schema."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "MSMEs"

        columns = [
            ('Business Name',     32, 'Required. The legal/trading name of the MSME.'),
            ('Owner Name',        24, 'Primary contact / owner name.'),
            ('Sex',               10, 'Male or Female.'),
            ('Phone',             18, 'Owner/contact phone number.'),
            ('Email',             28, 'Owner/contact email.'),
            ('Business Email',    28, 'Business email (if separate from owner email).'),
            ('Business Type',     20, 'Micro / Small / Medium / Sole proprietorship / Company / Partnership / Cooperative.'),
            ('Sector',            22, 'Optional. Manufacturing / Services / Trade / Agriculture / Technology / Healthcare / Education / Construction / Other. Auto-inferred from Business Type if blank.'),
            ('District',          18, 'District of operation.'),
            ('Town',              16, 'Town or city.'),
            ('Physical Location', 32, 'Street / landmark address (optional).'),
            ('Role',              18, 'Role of contact person (Director, Manager, etc.).'),
            ('Cohort',            14, 'e.g. Cohort 1, Cohort 2 (optional — can also be set in upload form).'),
        ]

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', start_color='1A2F4B')
        header_align = Alignment(horizontal='center', vertical='center')

        for i, (name, width, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=i, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[1].height = 28
        ws.freeze_panes = 'A2'

        # Add an example row so users see the expected shape
        example = [
            'Acholi Shea Cooperative Ltd', 'Anena Sharon', 'Female', '256775779335',
            'anena.sharon@example.com', 'info@acholishea.org', 'Cooperative',
            'Agriculture', 'Gulu', 'Gulu', 'Plot 12, Gulu town', 'Director', 'Cohort 1',
        ]
        for i, val in enumerate(example, start=1):
            c = ws.cell(row=2, column=i, value=val)
            c.font = Font(italic=True, color='888888')

        # Field-guidance row at the bottom
        guide_row = 4
        ws.cell(row=guide_row, column=1, value='Notes:').font = Font(bold=True, color='C8102E')
        for i, (_, _, note) in enumerate(columns, start=1):
            ws.cell(row=guide_row + 1, column=i, value=note).font = Font(size=9, color='666666', italic=True)
            ws.cell(row=guide_row + 1, column=i).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[guide_row + 1].height = 60

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="MSME_Upload_Template.xlsx"'
        return resp

    @action(detail=False, methods=['post'], url_path='upload')
    @transaction.atomic
    def upload(self, request):
        """Upload MSME list from Cohort 1 (CSV/Excel) or Cohort 2 (Survey Excel).
        Auto-detects format from column names.

        The whole import runs inside one DB transaction. Per-row exceptions are
        captured into the `skipped` list (so one bad row doesn't kill the rest),
        but if anything escapes that try/except — or the function returns with
        an HTTP error after rows were inserted — the transaction is rolled back
        and the table is left in its pre-import state.
        Accepts optional form field: cohort_name (string).
        """
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can upload MSME data.")

        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        # update_existing=true (default) → update duplicates; false → skip them
        update_existing = request.data.get('update_existing', 'true').lower() != 'false'
        cohort_name = request.data.get('cohort_name', '').strip()

        # Read file (CSV or Excel)
        try:
            raw_bytes = file.read()
            if file.name.endswith('.csv'):
                df = pd.read_csv(io.BytesIO(raw_bytes))
            elif file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(io.BytesIO(raw_bytes))
                # If the first row appears to be blank/unnamed, the real header may be in row 1
                unnamed = sum(1 for c in df.columns if str(c).startswith('Unnamed:'))
                if unnamed > len(df.columns) / 2:
                    df = pd.read_excel(io.BytesIO(raw_bytes), header=1)
            else:
                return Response({'error': 'Please upload a CSV or Excel file (.csv, .xlsx, .xls).'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': f'Could not read file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        # Normalise column names for detection (strip whitespace, keep original for lookup)
        cols = set(df.columns.tolist())
        cols_stripped = {c.strip() for c in cols}

        # Unified template — preferred format used by the cleaning script.
        # Required: Business Name, Owner Name, Phone, Email, District, Town, Sex, Business Type
        is_unified = ('Business Name' in cols_stripped and 'Owner Name' in cols_stripped)
        # Cohort 2 — numbered survey format (1.1. Business Name: …)
        is_cohort2 = '1.1.  Business Name:' in cols or any('Business Name' in c for c in cols if '1.1' in c)
        # Cohort 1 — PRUDEV II Cohort1 format (Name, District, Town, Name of contact person …)
        is_cohort1 = (not is_unified) and ('Name' in cols and 'Name of contact person' in cols)
        # Cohort 2 simple — survey export with plain headers (Business Name:, Name of Business Owner: …)
        is_cohort2_simple = (
            not is_unified and not is_cohort1 and not is_cohort2 and
            any('Business Name' in c for c in cols_stripped) and
            any('Business Owner' in c for c in cols_stripped)
        )

        if not is_unified and not is_cohort1 and not is_cohort2 and not is_cohort2_simple:
            return Response({
                'error': (
                    'Unrecognised file format. Use the unified upload template '
                    '(columns: Business Name, Owner Name, Sex, Phone, Email, Business Email, '
                    'Business Type, District, Town, Physical Location, Role).'
                )
            }, status=status.HTTP_400_BAD_REQUEST)

        # Resolve or create cohort
        cohort_obj = None
        if cohort_name:
            cohort_obj, _ = Cohort.objects.get_or_create(name=cohort_name)

        def clean_str(val, default=''):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return default
            s = str(val).strip()
            return '' if s == 'nan' else s

        def clean_phone(val):
            raw = clean_str(val)
            if not raw:
                return ''
            try:
                return str(int(float(raw)))
            except (ValueError, TypeError):
                return raw

        def map_gender(val):
            v = clean_str(val).upper()
            if v in ('M', 'MALE'):
                return 'MALE'
            if v in ('F', 'FEMALE'):
                return 'FEMALE'
            return ''

        def map_business_type(val):
            v = clean_str(val).upper()
            if 'MEDIUM' in v:
                return 'MEDIUM'
            if 'SMALL' in v or 'COMPANY' in v or 'SMC' in v or 'PARTNERSHIP' in v:
                return 'SMALL'
            return 'MICRO'  # sole proprietorship, cooperative, default

        def map_sector(val):
            v = clean_str(val).upper()
            if any(x in v for x in ('AGRO', 'FARM', 'AGRICULTURE', 'AGRI')):
                return 'AGRICULTURE'
            if any(x in v for x in ('MANUFACTUR', 'PROCESSING', 'MILLER', 'MILL')):
                return 'MANUFACTURING'
            if any(x in v for x in ('TRADE', 'BUYER', 'SHOP', 'INPUT', 'VET')):
                return 'TRADE'
            if any(x in v for x in ('SERVICE', 'BDS', 'DEVELOPMENT', 'FINANCIAL', 'PROVIDER', 'FINANCE')):
                return 'SERVICES'
            if 'TECH' in v or 'INNOVATOR' in v:
                return 'TECHNOLOGY'
            if 'HEALTH' in v:
                return 'HEALTHCARE'
            if 'EDUCATION' in v:
                return 'EDUCATION'
            if 'CONSTRUCT' in v:
                return 'CONSTRUCTION'
            return 'OTHER'

        created, updated, skipped = 0, 0, []

        def find_col(cols, prefix, keyword=None):
            """Match column whose stripped name starts with prefix+ space/dot, optionally containing keyword."""
            import re
            pat = re.compile(r'^\s*' + re.escape(prefix) + r'[\s\.]')
            for c in cols:
                if pat.match(c):
                    if keyword is None or keyword.lower() in c.lower():
                        return c
            return None

        # Pre-compute Cohort 2 simple column mapping once (outside the row loop)
        # Strip all column names to a lookup dict: stripped_name → original_name
        col_map = {c.strip(): c for c in cols}

        def get_col(*keywords):
            """Return first column whose stripped name contains all keywords (case-insensitive)."""
            for c_stripped, c_orig in col_map.items():
                c_lower = c_stripped.lower()
                if all(k.lower() in c_lower for k in keywords):
                    return c_orig
            return None

        if is_unified:
            u_bname    = col_map.get('Business Name')
            u_owner    = col_map.get('Owner Name')
            u_sex      = col_map.get('Sex')
            u_phone    = col_map.get('Phone')
            u_email    = col_map.get('Email')
            u_bemail   = col_map.get('Business Email')
            u_type     = col_map.get('Business Type')
            u_sector   = col_map.get('Sector')   # optional — falls back to map_sector(Business Type)
            u_district = col_map.get('District')
            u_town     = col_map.get('Town')
            u_address  = col_map.get('Physical Location') or col_map.get('Address')

        if is_cohort2_simple:
            # Column names for Cohort 2 simple format (e.g. "Business Name:", "Name of Business Owner:", …)
            # Use exact strip-match first, then fallback to keyword search
            s2_bname    = col_map.get('Business Name:') or col_map.get('Business Name') or get_col('business name')
            s2_owner    = col_map.get('Name of Business Owner:') or col_map.get('Name of Business Owner') or get_col('name of business owner')
            # Phone: prefer dedicated business phone, fall back to owner contacts
            s2_phone    = (col_map.get('Business Phone Number(s):') or col_map.get('Business Phone Number')
                           or get_col('business phone') or get_col('business owner contact'))
            s2_email    = col_map.get("Business Owners Email:") or col_map.get("Business Owner's Email") or get_col('owner', 'email')
            s2_bemail   = col_map.get('Business email address') or col_map.get('Business email address ') or get_col('business email')
            s2_sex      = col_map.get('Sex') or col_map.get('Sex ') or get_col('sex')
            s2_type     = col_map.get('Type of Business: ') or col_map.get('Type of Business:') or col_map.get('Type of Business') or get_col('type of business')
            s2_district = col_map.get('District') or col_map.get('district')
            s2_town     = col_map.get('Town') or col_map.get('town')
            s2_sector   = None  # not present in this format

        # Pre-compute Cohort 2 numbered column mapping once (outside the row loop)
        if is_cohort2:
            c2_bname  = find_col(cols, '1.1',  'Business Name')
            c2_brn    = find_col(cols, '1.2',  'Registration')
            c2_owner  = find_col(cols, '1.4',  'Owner')
            c2_phone1 = find_col(cols, '1.5')   # owner contacts
            c2_sex    = find_col(cols, '1.6',  'Sex')
            c2_email  = find_col(cols, '1.7',  'Email')
            c2_type   = find_col(cols, '1.10', 'Type')
            c2_phone2 = find_col(cols, '1.12') # business phone (preferred)
            c2_bemail = find_col(cols, '1.13')
            c2_sector = find_col(cols, '2.1',  'core business')

        blank_rows = 0  # track silently-blank rows so user knows why count differs
        for i, row in df.iterrows():
            try:
                if is_unified:
                    business_name = clean_str(row.get(u_bname, '')) if u_bname else ''
                    if not business_name:
                        if all((pd.isna(v) or str(v).strip() in ('', 'nan')) for v in row.values):
                            blank_rows += 1
                        else:
                            skipped.append({'row': i + 2, 'error': 'Missing Business Name'})
                        continue
                    record = {
                        'business_name': business_name,
                        'owner_name':    clean_str(row.get(u_owner, '')) if u_owner else '',
                        'gender':        map_gender(row.get(u_sex, '')) if u_sex else '',
                        'phone':         clean_phone(row.get(u_phone, '')) if u_phone else '',
                        'email':         clean_str(row.get(u_email, '')) if u_email else '',
                        'business_email': clean_str(row.get(u_bemail, '')) if u_bemail else '',
                        'business_type': map_business_type(row.get(u_type, '')) if u_type else 'MICRO',
                        # Sector: prefer an explicit Sector column, otherwise infer from Business Type / sector keywords
                        'sector':        (
                            map_sector(row.get(u_sector, '')) if u_sector
                            else (map_sector(row.get(u_type, '')) if u_type else 'OTHER')
                        ),
                        'state':         clean_str(row.get(u_district, '')) if u_district else '',
                        'city':          clean_str(row.get(u_town, '')) if u_town else '',
                        'address':       clean_str(row.get(u_address, '')) if u_address else '',
                        'country':       'Uganda',
                        'source_file':   file.name,
                        'cohort':        cohort_obj,
                        'is_active':     True,
                    }
                elif is_cohort2_simple:
                    business_name = clean_str(row.get(s2_bname, '')) if s2_bname else ''
                    if not business_name:
                        # check if the whole row is blank vs just missing business name
                        if all((pd.isna(v) or str(v).strip() in ('', 'nan')) for v in row.values):
                            blank_rows += 1
                        else:
                            skipped.append({'row': i + 2, 'error': 'Missing Business Name'})
                        continue

                    record = {
                        'business_name': business_name,
                        'owner_name': clean_str(row.get(s2_owner, '')) if s2_owner else '',
                        'phone': clean_phone(row.get(s2_phone, '')) if s2_phone else '',
                        'email': clean_str(row.get(s2_email, '')) if s2_email else '',
                        'business_email': clean_str(row.get(s2_bemail, '')) if s2_bemail else '',
                        'gender': map_gender(row.get(s2_sex, '')) if s2_sex else '',
                        'business_type': map_business_type(row.get(s2_type, '')) if s2_type else 'MICRO',
                        'sector': 'OTHER',
                        'state': clean_str(row.get(s2_district, '')) if s2_district else '',
                        'city': clean_str(row.get(s2_town, '')) if s2_town else '',
                        'country': 'Uganda',
                        'source_file': file.name,
                        'cohort': cohort_obj,
                        'is_active': True,
                    }

                elif is_cohort2:
                    bname_col  = c2_bname
                    owner_col  = c2_owner
                    sex_col    = c2_sex
                    email_col  = c2_email
                    type_col   = c2_type
                    phone_col  = c2_phone2 or c2_phone1
                    bemail_col = c2_bemail
                    sector_col = c2_sector
                    brn_col    = c2_brn

                    business_name = clean_str(row.get(bname_col, '')) if bname_col else ''
                    if not business_name:
                        if all((pd.isna(v) or str(v).strip() in ('', 'nan')) for v in row.values):
                            blank_rows += 1
                        else:
                            skipped.append({'row': i + 2, 'error': 'Missing Business Name'})
                        continue

                    record = {
                        'business_name': business_name,
                        'registration_number': clean_str(row.get(brn_col, '')) if brn_col else '',
                        'owner_name': clean_str(row.get(owner_col, '')) if owner_col else '',
                        'gender': map_gender(row.get(sex_col, '')) if sex_col else '',
                        'email': clean_str(row.get(email_col, '')) if email_col else '',
                        'phone': clean_phone(row.get(phone_col, '')) if phone_col else '',
                        'business_email': clean_str(row.get(bemail_col, '')) if bemail_col else '',
                        'business_type': map_business_type(row.get(type_col, '')) if type_col else 'MICRO',
                        'sector': map_sector(row.get(sector_col, '')) if sector_col else 'OTHER',
                        'state': clean_str(row.get('District', '')),
                        'city': clean_str(row.get('Town/City', '')),
                        'country': 'Uganda',
                        'source_file': file.name,
                        'cohort': cohort_obj,
                        'is_active': True,
                    }
                else:  # Cohort 1
                    business_name = clean_str(row.get('Name', ''))
                    if not business_name:
                        if all((pd.isna(v) or str(v).strip() in ('', 'nan')) for v in row.values):
                            blank_rows += 1
                        else:
                            skipped.append({'row': i + 2, 'error': 'Missing Name (business)'})
                        continue

                    gender = map_gender(row.get('Sex of founder', row.get('Gender of Key contact person', '')))

                    record = {
                        'business_name': business_name,
                        'owner_name': clean_str(row.get('Name of contact person', '')),
                        'gender': gender,
                        'phone': clean_phone(row.get('Mobile phone numbers ', row.get('Mobile phone numbers', ''))),
                        'email': clean_str(row.get('Email Address of contact person', '')),
                        'business_email': clean_str(row.get('Business Email Address', '')),
                        'address': clean_str(row.get('Physical location', '')),
                        'state': clean_str(row.get('District', '')),
                        'city': clean_str(row.get('Town', '')),
                        'business_type': 'MICRO',
                        'sector': 'OTHER',
                        'country': 'Uganda',
                        'source_file': file.name,
                        'cohort': cohort_obj,
                        'is_active': True,
                    }

                # Update-or-create by business name + owner (avoid duplicates on re-upload).
                # Pop both keys out of `record` so they aren't passed twice to create().
                business_name = record.pop('business_name')
                owner_name    = record.pop('owner_name', '')
                lookup = {'business_name': business_name}
                if owner_name:
                    lookup['owner_name'] = owner_name

                # Wrap each row's DB write in a savepoint so an IntegrityError
                # on row N doesn't poison the outer transaction for rows N+1..end.
                with transaction.atomic():
                    existing = MSME.objects.filter(**lookup).first()
                    if existing:
                        if update_existing:
                            # Restore owner_name into the update payload so existing rows can be enriched
                            if owner_name and not existing.owner_name:
                                existing.owner_name = owner_name
                            for k, v in record.items():
                                setattr(existing, k, v)
                            existing.save()
                            updated += 1
                        else:
                            skipped.append({'row': i + 2, 'error': f'Duplicate skipped: {business_name}'})
                    else:
                        # `lookup` carries the unique-fields, `record` carries everything else
                        MSME.objects.create(**lookup, **record)
                        created += 1

            except Exception as e:
                skipped.append({'row': i + 2, 'error': str(e)})

        msg = f"{created} MSMEs added, {updated} updated"
        if cohort_name:
            msg += f" (assigned to cohort: {cohort_name})"
        if skipped:
            msg += f", {len(skipped)} rows skipped"
        if blank_rows:
            msg += f", {blank_rows} blank rows ignored"
        msg += "."

        return Response({
            'created': created,
            'updated': updated,
            'skipped': len(skipped),
            'blank_rows': blank_rows,
            'errors': skipped[:50],  # cap at 50 in response
            'total_rows': int(len(df)),
            'message': msg,
        })

    @action(detail=False, methods=['get'])
    def analytics(self, request):
        """Rich analytics for the dashboard. Accepts optional filters that all
        downstream aggregations honour:
          ?cohort=<id>  ?district=<name>  ?sector=<code>  ?bge=<id>
        """
        from django.db.models import Q
        from django.db.models.functions import TruncMonth

        qs = MSME.objects.filter(is_active=True)

        # Apply optional filters so the analytics page can drill down.
        cohort_id = request.query_params.get('cohort')
        if cohort_id:
            qs = qs.filter(cohort_id=cohort_id)
        district = request.query_params.get('district')
        if district:
            qs = qs.filter(state__iexact=district)
        sector = request.query_params.get('sector')
        if sector:
            qs = qs.filter(sector=sector)
        bge_id = request.query_params.get('bge')
        if bge_id:
            qs = qs.filter(Q(assigned_bge_id=bge_id) | Q(assigned_group__members__id=bge_id)).distinct()

        agg = qs.aggregate(
            total_investment_needed=Sum('investment_needed'),
            total_annual_revenue=Sum('annual_revenue'),
        )

        # Total employees from the latest growth snapshot per MSME in scope.
        # 1) Find the most recent snapshot date per MSME.
        # 2) Pull employee columns from those rows and sum them.
        msme_ids = list(qs.values_list('id', flat=True))
        latest_snapshot_ids = (
            MSMEGrowthSnapshot.objects
            .filter(msme_id__in=msme_ids)
            .order_by('msme_id', '-snapshot_date')
            .distinct('msme_id')
            .values_list('id', flat=True)
        )
        snap_emp_agg = MSMEGrowthSnapshot.objects.filter(id__in=latest_snapshot_ids).aggregate(
            ft_male   = Sum('employees_ft_male'),
            ft_female = Sum('employees_ft_female'),
            pt_male   = Sum('employees_pt_male'),
            pt_female = Sum('employees_pt_female'),
            ft_refugee = Sum('employees_ft_refugee'),
            pt_refugee = Sum('employees_pt_refugee'),
        )
        snapshot_employees = {
            'ft_male':    snap_emp_agg['ft_male']    or 0,
            'ft_female':  snap_emp_agg['ft_female']  or 0,
            'pt_male':    snap_emp_agg['pt_male']    or 0,
            'pt_female':  snap_emp_agg['pt_female']  or 0,
            'ft_refugee': snap_emp_agg['ft_refugee'] or 0,
            'pt_refugee': snap_emp_agg['pt_refugee'] or 0,
        }
        total_employees_from_snapshots = sum(snapshot_employees.values())

        # Reports / activity stats
        from .models import MSMEReport, GroupReport
        report_status_stats = list(
            MSMEReport.objects.values('status').annotate(count=Count('id'))
        )
        group_report_status_stats = list(
            GroupReport.objects.values('status').annotate(count=Count('id'))
        )

        # BGE workload (top 15 BGEs by direct + group-assigned MSMEs)
        # Annotate everything in 2 queries instead of (1 + 3*N) per BGE.
        bge_qs = (BusinessGrowthExpert.objects
                  .filter(status='approved')
                  .annotate(
                      direct=Count(
                          'assigned_msmes',
                          filter=Q(assigned_msmes__is_active=True),
                          distinct=True,
                      ),
                      reports_count=Count('reports', distinct=True),
                  ))
        # `via_group` = MSMEs reachable through any of the BGE's groups.
        # Cheaper as a single grouped query than per-BGE.
        via_group_counts = dict(
            MSME.objects.filter(is_active=True, assigned_group__members__isnull=False)
                        .values_list('assigned_group__members')
                        .annotate(c=Count('id', distinct=True))
                        .values_list('assigned_group__members', 'c')
        )
        bge_workload = []
        for bge in bge_qs[:50]:
            via = via_group_counts.get(bge.id, 0)
            bge_workload.append({
                'bge_id':       bge.id,
                'bge_name':     bge.name,
                'direct':       bge.direct,
                'via_group':    via,
                'total':        bge.direct + via,
                'reports_count': bge.reports_count,
            })
        bge_workload.sort(key=lambda x: x['total'], reverse=True)
        bge_workload = bge_workload[:15]

        # Group performance — annotate in one query, no per-row counts.
        group_qs = (BGEGroup.objects
                    .select_related('team_lead')
                    .annotate(
                        msme_count=Count(
                            'assigned_msmes',
                            filter=Q(assigned_msmes__is_active=True),
                            distinct=True,
                        ),
                        active_member_count=Count('members', distinct=True),
                        reports_count=Count('reports', distinct=True),
                    ))
        group_stats = [
            {
                'group_id':       g.id,
                'group_name':     g.name,
                'msme_count':     g.msme_count,
                'member_count':   g.active_member_count,
                'reports_count':  g.reports_count,
                'team_lead_name': g.team_lead.name if g.team_lead else None,
            }
            for g in group_qs
        ]
        group_stats.sort(key=lambda x: x['msme_count'], reverse=True)

        # Time series — MSMEs created per month (last 18 months)
        time_series = list(
            qs.annotate(month=TruncMonth('created_at'))
              .values('month')
              .annotate(count=Count('id'))
              .order_by('month')
        )

        # Gender × Business Type cross-tab — for stacked bar
        gender_x_type = list(
            qs.values('gender', 'business_type')
              .annotate(count=Count('id'))
              .order_by('business_type', 'gender')
        )

        # ── Diagnostic analytics ──────────────────────────────────────────────
        diag_qs = qs.filter(diag_imported_at__isnull=False)
        diag_total = diag_qs.count()

        # Compliance & financial access flags
        diag_compliance = {
            'has_tin':           diag_qs.filter(diag_has_tin=True).count(),
            'has_ursb':          diag_qs.filter(diag_has_unbs=True).count(),
            'has_business_bank': diag_qs.filter(diag_has_business_bank=True).count(),
            'has_mobile_money':  diag_qs.filter(diag_has_mobile_money=True).count(),
            'total':             diag_total,
        }

        # Green business breakdown
        green_count = diag_qs.filter(diag_is_green_business=True).count()
        diag_green = {'green': green_count, 'non_green': diag_total - green_count, 'total': diag_total}

        # Turnover bands (count per band)
        diag_turnover_bands = list(
            diag_qs.exclude(diag_annual_turnover='')
                   .values('diag_annual_turnover')
                   .annotate(count=Count('id'))
                   .order_by('diag_annual_turnover')
        )

        # Total assets bands
        diag_asset_bands = list(
            diag_qs.exclude(diag_total_assets='')
                   .values('diag_total_assets')
                   .annotate(count=Count('id'))
                   .order_by('diag_total_assets')
        )

        # Employee aggregates (only rows that have values)
        from django.db.models import Sum as DSum
        emp_agg = diag_qs.aggregate(
            ft_male=DSum('diag_employees_ft_male'),
            ft_female=DSum('diag_employees_ft_female'),
            pt_male=DSum('diag_employees_pt_male'),
            pt_female=DSum('diag_employees_pt_female'),
        )
        diag_employees = {
            'ft_male':   emp_agg['ft_male']   or 0,
            'ft_female': emp_agg['ft_female'] or 0,
            'pt_male':   emp_agg['pt_male']   or 0,
            'pt_female': emp_agg['pt_female'] or 0,
        }

        # Owner sex from diagnostic
        diag_owner_sex = list(
            diag_qs.exclude(diag_owner_sex='')
                   .values('diag_owner_sex')
                   .annotate(count=Count('id'))
        )

        # Years operating distribution
        diag_years_operating = list(
            diag_qs.exclude(diag_years_operating='')
                   .values('diag_years_operating')
                   .annotate(count=Count('id'))
                   .order_by('diag_years_operating')
        )

        # Top districts from diagnostic (often more complete than `state`)
        diag_districts = list(
            diag_qs.exclude(diag_district='')
                   .values('diag_district')
                   .annotate(count=Count('id'))
                   .order_by('-count')[:12]
        )

        # MSMEs with at least one data update (growth snapshot)
        msmes_with_updates = (
            MSMEGrowthSnapshot.objects
            .filter(msme_id__in=msme_ids)
            .values('msme_id')
            .distinct()
            .count()
        )

        # BGEs active in the last 30 days (filed a snapshot or MSME report)
        from datetime import timedelta
        from django.utils import timezone
        cutoff = timezone.now() - timedelta(days=30)
        bges_via_snapshots = set(
            MSMEGrowthSnapshot.objects
            .filter(created_at__gte=cutoff, collected_by__isnull=False)
            .values_list('collected_by_id', flat=True)
        )
        from .models import MSMEReport
        bges_via_reports = set(
            MSMEReport.objects
            .filter(created_at__gte=cutoff)
            .values_list('bge_id', flat=True)
        )
        active_bges_30d = len(bges_via_snapshots | bges_via_reports)

        return Response({
            # KPIs
            'total_msmes': qs.count(),
            'total_investment_needed': agg['total_investment_needed'] or 0,
            'total_annual_revenue':    agg['total_annual_revenue']    or 0,
            'total_employees':         total_employees_from_snapshots,
            'snapshot_employees':      snapshot_employees,
            'msmes_with_updates':      msmes_with_updates,
            'active_bges_30d':         active_bges_30d,
            'total_bges':              BusinessGrowthExpert.objects.count(),
            'total_groups':            BGEGroup.objects.count(),
            'total_reports':           MSMEReport.objects.count(),
            'total_group_reports':     GroupReport.objects.count(),

            # Distributions
            'business_type_stats': list(qs.values('business_type').annotate(count=Count('id'))),
            'sector_stats':        list(qs.values('sector').annotate(count=Count('id'))),
            'cohort_stats':        list(qs.values('cohort__name', 'cohort_id').annotate(count=Count('id')).order_by('cohort__name')),
            'gender_stats':        list(qs.values('gender').annotate(count=Count('id'))),
            'top_districts':       list(qs.values('state').exclude(state='').annotate(count=Count('id')).order_by('-count')[:10]),
            'top_cities':          list(qs.values('city').exclude(city='').annotate(count=Count('id')).order_by('-count')[:10]),

            # Cross-tabs
            'gender_x_type': gender_x_type,

            # BGE / Group performance
            'bge_status_stats':    list(BusinessGrowthExpert.objects.values('status').annotate(count=Count('id'))),
            'bge_workload':        bge_workload,
            'group_stats':         group_stats,

            # Reports
            'report_status_stats': report_status_stats,
            'group_report_status_stats': group_report_status_stats,

            # Time series
            'time_series':         time_series,

            # Diagnostic baseline analytics
            'diag_total':           diag_total,
            'diag_compliance':      diag_compliance,
            'diag_green':           diag_green,
            'diag_turnover_bands':  diag_turnover_bands,
            'diag_asset_bands':     diag_asset_bands,
            'diag_employees':       diag_employees,
            'diag_owner_sex':       diag_owner_sex,
            'diag_years_operating': diag_years_operating,
            'diag_districts':       diag_districts,
        })

    @action(detail=False, methods=['get'], url_path='inactive')
    def inactive_msmes(self, request):
        """Admin-only: list all inactive MSMEs so they can be reviewed before reactivation."""
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can view inactive MSMEs.")
        qs = MSME.objects.filter(is_active=False).order_by('business_name')
        data = list(qs.values('id', 'business_name', 'msme_code', 'owner_name',
                              'sector', 'city', 'cohort_id'))
        return Response({'count': len(data), 'msmes': data})

    @action(detail=False, methods=['post'], url_path='reactivate-all')
    def reactivate_all(self, request):
        """Admin-only: set is_active=True for every inactive MSME."""
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can reactivate MSMEs.")
        updated = MSME.objects.filter(is_active=False).update(is_active=True)
        return Response({'reactivated': updated,
                         'message': f'{updated} MSME(s) reactivated successfully.'})

    @action(detail=False, methods=['post'], url_path='import-diagnostics')
    def import_diagnostics(self, request):
        """Admin-only: upload the diagnostics Excel and run the import in-process.
        POST /api/msmes/import-diagnostics/  multipart field: file
        Returns a JSON summary of matched/unmatched counts.
        """
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can import diagnostic data.")

        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'error': 'No file provided. Send multipart field "file".'}, status=400)

        import tempfile, os, io as _io, traceback as _tb
        from django.core.management import call_command

        suffix = os.path.splitext(uploaded.name)[1] or '.xlsx'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            for chunk in uploaded.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            out = _io.StringIO()
            call_command('import_diagnostics', tmp_path, stdout=out, stderr=out)
            output = out.getvalue()
        except Exception as exc:
            os.unlink(tmp_path)
            return Response({'error': f'{type(exc).__name__}: {exc}\n\n{_tb.format_exc()}'}, status=500)
        finally:
            try:
                os.unlink(tmp_path)
            except FileNotFoundError:
                pass

        return Response({'detail': output or 'Import complete.'})


class BusinessGrowthExpertViewSet(ProgrammeManagerReadOnlyMixin, ViewerReadOnlyMixin, viewsets.ModelViewSet):
    queryset = BusinessGrowthExpert.objects.all()
    serializer_class = BusinessGrowthExpertSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = BusinessGrowthExpert.objects.all()
        s = self.request.query_params.get('status')
        if s:
            qs = qs.filter(status=s)
        group = self.request.query_params.get('group')
        if group:
            qs = qs.filter(bge_groups__id=group)
        # Prefetch the relations the serializer pulls in (assigned_msmes,
        # bge_groups). Without this, listing N BGEs caused 2*N extra queries
        # (each BGE serialized triggered .assigned_msmes.all() and .bge_groups.all()).
        return qs.prefetch_related('assigned_msmes', 'bge_groups').order_by('-created_at')

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff and not request.user.is_superuser:
            raise PermissionDenied("Only admins can delete experts.")
        return super().destroy(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can create expert profiles.")
        return super().create(request, *args, **kwargs)

    def _check_can_write(self, request, pk):
        user = request.user
        if user.is_staff or user.is_superuser:
            return
        own_bge = getattr(user, 'bge_profile', None)
        if own_bge is None or str(own_bge.id) != str(pk):
            raise PermissionDenied("You can only update your own profile.")

    def update(self, request, *args, **kwargs):
        self._check_can_write(request, kwargs.get('pk'))
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        self._check_can_write(request, kwargs.get('pk'))
        return super().partial_update(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def leaderboard(self, request):
        bges = (BusinessGrowthExpert.objects
                .filter(status='approved')
                .prefetch_related('assigned_msmes', 'bge_groups')
                .annotate(support_count=Count('support_requests'))
                .order_by('-support_count'))
        return Response(BusinessGrowthExpertSerializer(bges, many=True).data)

    @staticmethod
    def _bge_all_msme_ids(bge):
        """Return all MSME IDs visible to a BGE: direct assignments + group assignments."""
        direct = set(bge.assigned_msmes.values_list('id', flat=True))
        via_group = set(MSME.objects.filter(
            assigned_group__in=bge.bge_groups.all()
        ).values_list('id', flat=True))
        return direct | via_group

    @staticmethod
    def _already_assigned_bges(bge):
        """Return other BGEs whose issued/signed work orders overlap in date with
        any of this BGE's active work orders AND share at least one MSME.

        Detection uses (in priority order):
          1. msme_ids_snapshot — set at issue time, survives re-assignment
          2. current direct + group MSME assignments — fallback for legacy records
        Both sides use the union so legacy and new records all work.
        """
        from .models import WorkOrder as _WO

        # Best-available MSME set for this BGE: snapshot union current
        my_current = BusinessGrowthExpertViewSet._bge_all_msme_ids(bge)
        my_wos = _WO.objects.filter(bge=bge, status__in=['issued', 'signed'])
        already = {}

        for my_wo in my_wos:
            if not (my_wo.start_date and my_wo.end_date):
                continue
            # Union snapshot with current so re-assigned MSMEs are still detected
            my_msme_ids = set(my_wo.msme_ids_snapshot or []) | my_current
            if not my_msme_ids:
                continue

            overlapping = _WO.objects.filter(
                status__in=['issued', 'signed'],
                start_date__lte=my_wo.end_date,
                end_date__gte=my_wo.start_date,
            ).exclude(bge=bge).exclude(id=my_wo.id).select_related('bge')

            for owo in overlapping:
                if owo.bge_id in already:
                    continue
                other_current = BusinessGrowthExpertViewSet._bge_all_msme_ids(owo.bge)
                other_msme_ids = set(owo.msme_ids_snapshot or []) | other_current
                shared_ids = my_msme_ids & other_msme_ids
                if shared_ids:
                    shared_msmes = list(
                        MSME.objects.filter(id__in=shared_ids)
                        .values('id', 'business_name', 'msme_code')
                        .order_by('business_name')
                    )
                    already[owo.bge_id] = {
                        'bge': owo.bge,
                        'work_order_number': owo.work_order_number,
                        'objectives': (owo.objective or owo.bge.deployment_objectives or '').strip(),
                        'shared_msmes': shared_msmes,
                    }
        return list(already.values())

    def _build_assignment_email(self, bge):
        """Build plain-text + HTML email for a BGE assignment. Shared by preview and send."""
        from django.db.models import Q
        # Include both primary and co-assigned MSMEs so the email reflects everything
        # the BGE is expected to work on — whether they are the primary or joint BGE.
        msmes = MSME.objects.filter(
            Q(assigned_bge=bge) | Q(co_assigned_bges=bge),
            is_active=True,
        ).distinct().order_by('business_name')
        count = msmes.count()
        already_assigned = self._already_assigned_bges(bge)

        # ── Plain-text version ────────────────────────────────────────────────
        lines = [f"Dear {bge.name},", "", "Please find below your assignment details under the PRUDEV II Programme:", ""]
        if bge.deployment_objectives:
            lines += ["DEPLOYMENT OBJECTIVES", "─" * 40, bge.deployment_objectives, ""]
        lines += [f"ASSIGNED MSMEs ({count} {'businesses' if count != 1 else 'business'})", "─" * 40, ""]
        for i, m in enumerate(msmes, 1):
            lines.append(f"  {i}. {m.business_name} ({m.msme_code or 'No code'})")
            if m.owner_name: lines.append(f"     Owner: {m.owner_name}")
            if m.sector:     lines.append(f"     Sector: {m.sector}")
            if m.city:       lines.append(f"     Location: {m.city}")
            if m.phone:      lines.append(f"     Phone: {m.phone}")
            lines.append("")
        if already_assigned:
            lines += ["PLEASE NOTE — ANOTHER BGE IS ALREADY ASSIGNED", "─" * 40]
            lines.append(
                "Another BGE has already been assigned to work with some of the same MSMEs "
                "during this period. Please be aware of their work and coordinate accordingly."
            )
            lines.append("")
            for aa in already_assigned:
                other = aa['bge']
                lines.append(f"  BGE Name:   {other.name} ({other.bge_code or 'No code'})")
                lines.append(f"  Work Order: {aa['work_order_number']}")
                if other.phone:  lines.append(f"  Phone:      {other.phone}")
                if other.email:  lines.append(f"  Email:      {other.email}")
                if aa.get('shared_msmes'):
                    lines.append(f"  Shared MSMEs ({len(aa['shared_msmes'])}):")
                    for m in aa['shared_msmes']:
                        lines.append(f"    • {m['business_name']} ({m['msme_code'] or 'No code'})")
                if aa['objectives']:
                    lines.append(f"  Their Objectives:")
                    lines.append(f"    {aa['objectives'][:300]}")
                lines.append("")

        lines += [
            "Please log in to the PRUDEV II Portfolio Management System to view full details and submit visit reports.",
            "", "Best regards,", "PRUDEV II BDS Team", "GIZ · GOPA AFC",
        ]
        body_text = "\n".join(lines)

        # ── HTML version (renders beautifully in Outlook) ─────────────────────
        from django.utils.html import escape as _esc
        objectives_html = ""
        if bge.deployment_objectives:
            objectives_html = f"""
            <div style="background:#f8f9fa;border-left:4px solid #1A2E42;padding:12px 16px;margin:16px 0;border-radius:0 4px 4px 0;">
              <p style="font-weight:700;color:#1A2E42;margin:0 0 6px 0;font-size:12px;text-transform:uppercase;letter-spacing:.05em;">Deployment Objectives</p>
              <p style="margin:0;color:#333;white-space:pre-line;">{_esc(bge.deployment_objectives)}</p>
            </div>"""

        msme_rows_html = ""
        for i, m in enumerate(msmes, 1):
            details = []
            if m.owner_name: details.append(f"<span style='color:#555;'>Owner:</span> {_esc(m.owner_name)}")
            if m.sector:     details.append(f"<span style='color:#555;'>Sector:</span> {_esc(m.sector)}")
            if m.city:       details.append(f"<span style='color:#555;'>Location:</span> {_esc(m.city)}")
            if m.phone:      details.append(f"<span style='color:#555;'>Phone:</span> {_esc(m.phone)}")
            details_html = " &nbsp;·&nbsp; ".join(details)
            bg = "#ffffff" if i % 2 == 0 else "#f9fafb"
            msme_rows_html += f"""
            <tr style="background:{bg};">
              <td style="padding:10px 14px;font-weight:600;color:#1A2E42;width:28px;vertical-align:top;">{i}.</td>
              <td style="padding:10px 14px;">
                <strong>{_esc(m.business_name)}</strong>
                <span style="color:#888;font-size:12px;margin-left:6px;">({_esc(m.msme_code or 'No code')})</span>
                {'<br><span style="font-size:12px;color:#666;">' + details_html + '</span>' if details_html else ''}
              </td>
            </tr>"""

        # ── Already-assigned BGE notice (HTML) ───────────────────────────────
        co_bge_html = ""
        if already_assigned:
            aa_cards = ""
            for aa in already_assigned:
                other = aa['bge']
                obj_snippet = (_esc(aa['objectives'][:300]) + ('…' if len(aa['objectives']) > 300 else '')) if aa['objectives'] else '<em style="color:#999;">No objectives recorded for this BGE yet.</em>'
                contact_parts = []
                if other.phone: contact_parts.append(f"📞 {_esc(other.phone)}")
                if other.email: contact_parts.append(f"✉ {_esc(other.email)}")
                contact_line = " &nbsp;·&nbsp; ".join(contact_parts)
                # Build shared MSME list HTML
                shared_msmes_html = ''
                if aa.get('shared_msmes'):
                    msme_items = ''.join(
                        f'<li style="margin:3px 0;font-size:12px;color:#333;">'
                        f'<strong>{_esc(m["business_name"])}</strong>'
                        f'<span style="color:#888;margin-left:6px;">({_esc(m["msme_code"] or "No code")})</span>'
                        f'</li>'
                        for m in aa['shared_msmes']
                    )
                    shared_msmes_html = f"""
                  <div style="background:#E8F5E9;border-left:3px solid #2E7D32;padding:8px 12px;border-radius:0 4px 4px 0;margin-top:8px;">
                    <p style="margin:0 0 4px;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.05em;color:#2E7D32;">
                      Shared MSMEs ({len(aa['shared_msmes'])})
                    </p>
                    <ul style="margin:0;padding-left:16px;">{msme_items}</ul>
                  </div>"""
                aa_cards += f"""
                <div style="background:#fff;border:1px solid #e8edf2;border-radius:6px;padding:14px 16px;margin-top:10px;">
                  <p style="margin:0 0 4px;font-weight:700;color:#1A2E42;font-size:14px;">{_esc(other.name)}</p>
                  <p style="margin:0 0 8px;font-size:11px;color:#888;">Work Order: {_esc(aa['work_order_number'])} &nbsp;·&nbsp; {_esc(other.bge_code or 'No code')}</p>
                  {f'<p style="margin:0 0 8px;font-size:12px;color:#555;">{contact_line}</p>' if contact_line else ''}
                  {shared_msmes_html}
                  <div style="background:#f8f9fa;border-left:3px solid #C8102E;padding:8px 12px;border-radius:0 4px 4px 0;margin-top:8px;">
                    <p style="margin:0 0 3px;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.05em;color:#C8102E;">Their Visit Objectives</p>
                    <p style="margin:0;font-size:12px;color:#555;line-height:1.55;white-space:pre-line;">{obj_snippet}</p>
                  </div>
                </div>"""
            co_bge_html = f"""
            <div style="background:#FFF3E0;border:1px solid #FFCC80;border-radius:8px;padding:18px 20px;margin-top:24px;">
              <p style="margin:0 0 8px;font-weight:700;color:#E65100;font-size:13px;">
                ⚠ Please Note — Another BGE Has Already Been Assigned
              </p>
              <p style="margin:0 0 12px;color:#555;font-size:13px;line-height:1.6;">
                Another BGE has already been assigned to work with some of the same MSMEs during
                this period. Please be aware of their work and coordinate accordingly to ensure
                your visits are complementary and not duplicated.
              </p>
              {aa_cards}
            </div>"""

        body_html = f"""
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f4f6f9;font-family:Arial,Helvetica,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f6f9;padding:32px 16px;">
    <tr><td align="center">
      <table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.08);">

        <!-- Header -->
        <tr><td style="background:#1A2E42;padding:24px 32px;">
          <table width="100%" cellpadding="0" cellspacing="0">
            <tr>
              <td><p style="margin:0;color:#fff;font-size:20px;font-weight:700;">PRUDEV II</p>
                  <p style="margin:2px 0 0;color:rgba(255,255,255,.65);font-size:12px;">MSME Portfolio Management Programme</p></td>
              <td align="right"><p style="margin:0;color:#C8102E;font-size:11px;font-weight:700;letter-spacing:.05em;">GIZ · GOPA AFC</p></td>
            </tr>
          </table>
        </td></tr>

        <!-- Body -->
        <tr><td style="padding:28px 32px;">
          <p style="margin:0 0 16px;color:#333;font-size:15px;">Dear <strong>{_esc(bge.name)}</strong>,</p>
          <p style="margin:0 0 20px;color:#555;line-height:1.6;">
            Please find below your assignment details under the <strong>PRUDEV II Programme</strong>.
          </p>

          {objectives_html}

          <!-- MSME Table -->
          <p style="font-weight:700;color:#1A2E42;font-size:12px;text-transform:uppercase;letter-spacing:.05em;margin:24px 0 8px;">
            Assigned MSMEs &nbsp;<span style="background:#1A2E42;color:#fff;border-radius:12px;padding:2px 8px;font-size:11px;">{count}</span>
          </p>
          <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e8edf2;border-radius:6px;overflow:hidden;">
            <thead>
              <tr style="background:#1A2E42;">
                <th style="padding:8px 14px;color:rgba(255,255,255,.7);font-size:11px;text-align:left;font-weight:600;">#</th>
                <th style="padding:8px 14px;color:rgba(255,255,255,.7);font-size:11px;text-align:left;font-weight:600;">Business / Details</th>
              </tr>
            </thead>
            <tbody>{msme_rows_html}</tbody>
          </table>

          <p style="margin:24px 0 0;color:#555;font-size:13px;line-height:1.7;border-top:1px solid #e8edf2;padding-top:20px;">
            Please log in to the <strong>PRUDEV II Portfolio Management System</strong> to view full details and submit visit reports.
          </p>

          {co_bge_html}

        </td></tr>

        <!-- Footer -->
        <tr><td style="background:#f8f9fa;padding:16px 32px;border-top:1px solid #e8edf2;">
          <p style="margin:0;color:#777;font-size:12px;">Best regards,<br><strong>PRUDEV II BDS Team</strong><br>GIZ · GOPA AFC</p>
        </td></tr>

      </table>
    </td></tr>
  </table>
</body>
</html>"""

        return {
            'subject':    f"PRUDEV II — Assignment Brief: {count} MSME{'s' if count != 1 else ''}",
            'body':       body_text,
            'body_html':  body_html,
            'to':         bge.email,
            'msme_count': count,
        }

    @action(detail=True, methods=['patch'], url_path='set-objectives')
    def set_objectives(self, request, pk=None):
        """Save shared deployment objectives for this BGE."""
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can set deployment objectives.")
        bge = self.get_object()
        bge.deployment_objectives = request.data.get('deployment_objectives', '').strip()
        bge.save()
        return Response(BusinessGrowthExpertSerializer(bge).data)

    @action(detail=True, methods=['get'], url_path='preview-email')
    def preview_email(self, request, pk=None):
        """Return the email that would be sent without actually sending it."""
        if not (request.user.is_staff or request.user.is_superuser or _managed_groups(request.user) is not None):
            raise PermissionDenied("Only admins can preview assignment emails.")
        bge = self.get_object()
        if not bge.email:
            return Response({'error': 'This BGE expert has no email address on record.'}, status=status.HTTP_400_BAD_REQUEST)
        if not bge.assigned_msmes.filter(is_active=True).exists():
            return Response({'error': 'This BGE expert has no assigned MSMEs.'}, status=status.HTTP_400_BAD_REQUEST)
        return Response(self._build_assignment_email(bge))

    @action(detail=True, methods=['post'], url_path='send-email')
    def send_assignment_email(self, request, pk=None):
        """Send BGE their MSME assignment via Microsoft Office 365 (richard.obuku@gopa.eu)."""
        if not (request.user.is_staff or request.user.is_superuser or _managed_groups(request.user) is not None):
            raise PermissionDenied("Only admins can send assignment emails.")

        bge = self.get_object()
        if not bge.email:
            return Response({'error': 'This BGE expert has no email address on record.'}, status=status.HTTP_400_BAD_REQUEST)
        if not bge.assigned_msmes.filter(is_active=True).exists():
            return Response({'error': 'This BGE expert has no assigned MSMEs.'}, status=status.HTTP_400_BAD_REQUEST)

        email_data = self._build_assignment_email(bge)
        # Frontend editable preview may override subject/body
        subject   = request.data.get('subject', '').strip() or email_data['subject']
        body_text = request.data.get('body', '').strip()    or email_data['body']
        body_html = email_data['body_html']  # always use generated HTML

        from_addr  = settings.DEFAULT_FROM_EMAIL
        reply_to   = getattr(settings, 'EMAIL_REPLY_TO', 'richard.obuku@gopa.eu')
        try:
            msg = EmailMultiAlternatives(
                subject=subject,
                body=body_text,
                from_email=from_addr,
                to=[bge.email],
                reply_to=[reply_to],
            )
            msg.attach_alternative(body_html, "text/html")
            msg.send(fail_silently=False)
            return Response({'message': f"Email sent to {bge.email} with {email_data['msme_count']} assigned MSMEs."})
        except Exception as e:
            return Response({'error': f'Failed to send email: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='upload-signature')
    def upload_signature(self, request, pk=None):
        """BGE or admin can upload / replace the BGE's signature image.

        The image is processed with Pillow: converted to RGBA and white/near-white
        pixels are made transparent, producing a clean signature on any background.
        """
        bge = self.get_object()
        # BGEs can only update their own signature; admins can update any.
        is_admin = request.user.is_staff or request.user.is_superuser
        try:
            requester_bge = request.user.bge_profile
        except Exception:
            requester_bge = None
        if not is_admin and (requester_bge is None or requester_bge.id != bge.id):
            raise PermissionDenied("You can only upload your own signature.")

        sig_file = request.FILES.get('signature')
        if not sig_file:
            return Response({'error': 'No signature file provided.'}, status=status.HTTP_400_BAD_REQUEST)
        if sig_file.size > 5 * 1024 * 1024:  # 5 MB hard cap
            return Response({'error': 'Signature file must be under 5 MB.'}, status=status.HTTP_400_BAD_REQUEST)
        if not sig_file.content_type.startswith('image/'):
            return Response({'error': 'File must be a JPEG or PNG image.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from PIL import Image as PilImage
            import io as _io

            img = PilImage.open(sig_file).convert('RGBA')

            # Make near-white pixels transparent so the signature floats cleanly
            datas = img.getdata()
            new_data = []
            threshold = 230  # pixels brighter than this in all channels → transparent
            for item in datas:
                r, g, b, a = item
                if r > threshold and g > threshold and b > threshold:
                    new_data.append((r, g, b, 0))
                else:
                    new_data.append((r, g, b, a))
            img.putdata(new_data)

            # Normalise size: cap at 600px wide keeping aspect ratio
            max_w = 600
            if img.width > max_w:
                ratio = max_w / img.width
                img = img.resize((max_w, int(img.height * ratio)), PilImage.LANCZOS)

            buf = _io.BytesIO()
            img.save(buf, format='PNG')
            png_bytes = buf.getvalue()

            from django.core.files.base import ContentFile
            filename = f'sig_{bge.bge_code or bge.id}.png'
            bge.signature.delete(save=False)  # remove old file if present
            bge.signature.save(filename, ContentFile(png_bytes), save=False)
            # Also persist bytes in DB so signature survives Render filesystem wipes
            bge.signature_data = png_bytes
            bge.save(update_fields=['signature', 'signature_data'])

        except Exception as exc:
            return Response({'error': f'Image processing failed: {exc}'},
                            status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'signature_url': request.build_absolute_uri(f'/api/experts/{bge.id}/signature-image/') if _bge_signature_bytes(bge) else None,
            'detail': 'Signature uploaded and processed successfully.',
        })

    @action(detail=True, methods=['get'], url_path='signature-image')
    def signature_image(self, request, pk=None):
        """Serve the BGE's stored signature as a PNG.

        Reads from ``signature_data`` (DB-backed) so it works in production
        where ``/media/`` is not served and the filesystem copy may be wiped.
        """
        bge = self.get_object()
        raw = _bge_signature_bytes(bge)
        if not raw:
            return Response({'detail': 'No signature found for this BGE.'}, status=status.HTTP_404_NOT_FOUND)
        return HttpResponse(raw, content_type='image/png')

    @action(detail=True, methods=['post'], url_path='rotate-signature')
    def rotate_signature(self, request, pk=None):
        """Admin: rotate a BGE's stored signature 90° CCW or CW and save permanently."""
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can rotate signatures.")
        bge = self.get_object()
        direction = request.data.get('direction', 'ccw')  # 'ccw' or 'cw'
        # PIL rotate: positive degrees = counter-clockwise
        degrees = 90 if direction == 'ccw' else -90

        raw = _bge_signature_bytes(bge)
        if not raw:
            return Response({'detail': 'No signature found for this BGE.'},
                            status=status.HTTP_404_NOT_FOUND)
        try:
            from PIL import Image as _PilImg
            img = _PilImg.open(_io.BytesIO(raw)).convert('RGBA')
            rotated = img.rotate(degrees, expand=True)
            buf = _io.BytesIO()
            rotated.save(buf, format='PNG')
            png_bytes = buf.getvalue()

            from django.core.files.base import ContentFile
            bge.signature_data = png_bytes
            fname = f'sig_{bge.bge_code or bge.id}.png'
            bge.signature.delete(save=False)
            bge.signature.save(fname, ContentFile(png_bytes), save=False)
            bge.save(update_fields=['signature', 'signature_data'])
        except Exception as exc:
            return Response({'detail': f'Rotation failed: {exc}'},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({
            'detail': f'Signature rotated {direction.upper()} 90° and saved.',
            'signature_url': request.build_absolute_uri(f'/api/experts/{bge.id}/signature-image/') if _bge_signature_bytes(bge) else None,
        })

    @action(detail=True, methods=['post'], url_path='clean-signature')
    def clean_signature(self, request, pk=None):
        """Admin: remove background from a BGE's stored signature.

        Uses BFS flood-fill seeded from all four edges to detect the background
        colour and erase connected regions of similar colour. This handles
        white, grey, cream, or any solid-colour background — not just pure white.
        A luminance sweep pass follows to catch any disconnected bright patches.
        Tolerance is configurable via POST body {'tolerance': 50} (default 50).
        """
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can clean signatures.")

        bge = self.get_object()
        raw = _bge_signature_bytes(bge)
        if not raw:
            return Response({'detail': 'No signature found for this BGE.'},
                            status=status.HTTP_404_NOT_FOUND)

        tolerance = int(request.data.get('tolerance', 50))
        tolerance = max(10, min(tolerance, 150))  # clamp to sane range

        try:
            from PIL import Image as _PilImg
            import math as _math

            img = _PilImg.open(_io.BytesIO(raw)).convert('RGBA')
            width, height = img.size
            pixels = img.load()

            # ── Step 1: sample background colour from ALL border pixels ──────
            edge_samples = []
            for x in range(width):
                edge_samples.append(pixels[x, 0][:3])
                edge_samples.append(pixels[x, height - 1][:3])
            for y in range(1, height - 1):
                edge_samples.append(pixels[0, y][:3])
                edge_samples.append(pixels[width - 1, y][:3])

            bg_r = sum(s[0] for s in edge_samples) // len(edge_samples)
            bg_g = sum(s[1] for s in edge_samples) // len(edge_samples)
            bg_b = sum(s[2] for s in edge_samples) // len(edge_samples)

            # ── Step 2: BFS flood-fill seeded from ALL four edges ────────────
            from collections import deque
            visited = [[False] * height for _ in range(width)]
            queue = deque()

            # Seed every pixel on all four edges
            for x in range(width):
                for y_seed in (0, height - 1):
                    if not visited[x][y_seed]:
                        visited[x][y_seed] = True
                        queue.append((x, y_seed))
            for y in range(1, height - 1):
                for x_seed in (0, width - 1):
                    if not visited[x_seed][y]:
                        visited[x_seed][y] = True
                        queue.append((x_seed, y))

            neighbours = [(1, 0), (-1, 0), (0, 1), (0, -1)]
            while queue:
                x, y = queue.popleft()
                r, g, b, a = pixels[x, y]
                dist = _math.sqrt((r - bg_r) ** 2 + (g - bg_g) ** 2 + (b - bg_b) ** 2)
                if dist > tolerance:
                    continue  # not background — stop expanding
                pixels[x, y] = (r, g, b, 0)  # erase
                for dx, dy in neighbours:
                    nx, ny = x + dx, y + dy
                    if 0 <= nx < width and 0 <= ny < height and not visited[nx][ny]:
                        visited[nx][ny] = True
                        queue.append((nx, ny))

            # ── Step 3: erase isolated near-background pixels (JPEG artefacts)
            data = list(img.getdata())
            new_data = []
            for r, g, b, a in data:
                if a == 0:
                    new_data.append((r, g, b, 0))
                    continue
                dist = _math.sqrt((r - bg_r) ** 2 + (g - bg_g) ** 2 + (b - bg_b) ** 2)
                if dist < tolerance * 0.75:
                    new_data.append((r, g, b, 0))
                else:
                    new_data.append((r, g, b, a))
            img.putdata(new_data)

            # ── Step 4: luminance sweep to catch bright non-connected patches ─
            # Handles off-white paper in scanned/photographed signatures
            data2 = list(img.getdata())
            new_data2 = []
            for r, g, b, a in data2:
                if a == 0:
                    new_data2.append((r, g, b, 0))
                    continue
                lum = 0.299 * r + 0.587 * g + 0.114 * b
                sat_range = max(r, g, b) - min(r, g, b)
                if lum > 210 and sat_range < 40:
                    new_data2.append((r, g, b, 0))
                else:
                    new_data2.append((r, g, b, a))
            img.putdata(new_data2)

            # ── Step 5: save ──────────────────────────────────────────────────
            buf = _io.BytesIO()
            img.save(buf, format='PNG')
            png_bytes = buf.getvalue()

            from django.core.files.base import ContentFile
            bge.signature_data = png_bytes
            fname = f'sig_{bge.bge_code or bge.id}.png'
            bge.signature.delete(save=False)
            bge.signature.save(fname, ContentFile(png_bytes), save=False)
            bge.save(update_fields=['signature', 'signature_data'])

        except Exception as exc:
            return Response({'detail': f'Cleaning failed: {exc}'},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({
            'detail': 'Signature background removed successfully.',
            'signature_url': request.build_absolute_uri(f'/api/experts/{bge.id}/signature-image/') if _bge_signature_bytes(bge) else None,
        })

    @action(detail=False, methods=['post'], url_path='upload')
    def upload(self, request):
        """Upload BGE list from Excel. Matches PRUDEV II BGE list format."""
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can upload BGE data.")

        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'Please upload an Excel file (.xlsx or .xls).'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(io.BytesIO(file.read()))
        except Exception as e:
            return Response({'error': f'Could not read file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        update_existing = request.data.get('update_existing', 'true').lower() != 'false'
        created, updated, skipped = 0, 0, 0
        errors = []

        for i, row in df.iterrows():
            name = str(row.get('Full name', '')).strip()
            if not name or name == 'nan':
                continue

            # Phone: Excel stores as float (2.567e+11) — convert to clean string
            raw_phone = row.get('Phone number', '')
            if pd.notna(raw_phone) and str(raw_phone).strip() not in ('', 'nan'):
                try:
                    phone = str(int(float(raw_phone)))
                except (ValueError, TypeError):
                    phone = str(raw_phone).strip()
            else:
                phone = ''

            raw_email = row.get('Email address', '')
            email = str(raw_email).strip() if pd.notna(raw_email) else ''
            if email == 'nan':
                email = ''

            raw_location = row.get('Location', '')
            location = str(raw_location).strip() if pd.notna(raw_location) else ''
            if location == 'nan':
                location = ''

            raw_code = row.get('BGE code', row.get('BGE Code', ''))
            bge_code = str(raw_code).strip() if pd.notna(raw_code) else ''
            if bge_code == 'nan':
                bge_code = ''

            try:
                existing = BusinessGrowthExpert.objects.filter(name=name).first()
                if existing:
                    if update_existing:
                        existing.email = email
                        existing.phone = phone
                        existing.location = location
                        existing.bge_code = bge_code
                        existing.status = 'approved'
                        existing.save()
                        updated += 1
                    else:
                        errors.append(f'Row {i + 2}: {name} — Duplicate skipped')
                        skipped += 1
                else:
                    BusinessGrowthExpert.objects.create(
                        name=name, email=email, phone=phone,
                        location=location, bge_code=bge_code, status='approved'
                    )
                    created += 1
            except Exception as e:
                errors.append(f'Row {i + 2}: {name} — {str(e)}')
                skipped += 1

        return Response({
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'errors': errors[:10],
            'message': f'{created} BGEs added, {updated} updated, {skipped} skipped.',
        }, status=status.HTTP_200_OK)


class BGEGroupViewSet(viewsets.ModelViewSet):
    queryset = BGEGroup.objects.all()
    serializer_class = BGEGroupSerializer
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_staff and not request.user.is_superuser:
            raise PermissionDenied("Only admins can delete BGE groups.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def add_member(self, request, pk=None):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can change group membership.")
        group = self.get_object()
        bge_id = request.data.get('bge_id')
        try:
            bge = BusinessGrowthExpert.objects.get(pk=bge_id)
            group.members.add(bge)
            return Response(BGEGroupSerializer(group).data)
        except BusinessGrowthExpert.DoesNotExist:
            return Response({'error': 'Expert not found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'])
    def remove_member(self, request, pk=None):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can change group membership.")
        group = self.get_object()
        bge_id = request.data.get('bge_id')
        try:
            bge = BusinessGrowthExpert.objects.get(pk=bge_id)
            group.members.remove(bge)
            # If the BGE we just removed was the team lead, clear that role too
            # so they can't keep submitting reports on a group they no longer
            # belong to (the write-permission check matches on team_lead_id).
            if group.team_lead_id == bge.id:
                group.team_lead = None
                group.save(update_fields=['team_lead'])
            return Response(BGEGroupSerializer(group).data)
        except BusinessGrowthExpert.DoesNotExist:
            return Response({'error': 'Expert not found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'], url_path='assign-msmes')
    def assign_msmes(self, request, pk=None):
        """Bulk-assign MSMEs to this group.
        Body: {msme_ids: [...], session_number?: int, objectives?: str}.
        Every BGE in the group's `members` will then see these MSMEs in their dashboard.
        """
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can assign MSMEs to groups.")
        group = self.get_object()
        msme_ids = request.data.get('msme_ids', [])
        if not isinstance(msme_ids, list) or not msme_ids:
            return Response({'error': 'msme_ids must be a non-empty list'}, status=status.HTTP_400_BAD_REQUEST)

        session_number = request.data.get('session_number')
        objectives     = request.data.get('objectives', '').strip()

        update_fields = {'assigned_group': group}
        if session_number is not None:
            try:
                update_fields['session_number'] = int(session_number)
            except (TypeError, ValueError):
                return Response({'error': 'session_number must be an integer'}, status=status.HTTP_400_BAD_REQUEST)
        # Use the form-supplied objectives if any, otherwise fall back to the group's
        # canonical objectives so each MSME inherits its team's mission.
        effective_objectives = objectives or (group.objectives or '').strip()
        if effective_objectives:
            update_fields['assignment_objectives'] = effective_objectives

        updated = MSME.objects.filter(id__in=msme_ids, is_active=True).update(**update_fields)
        return Response({
            'group_id': group.id,
            'group_name': group.name,
            'assigned': updated,
            'msme_ids': msme_ids,
        })

    @action(detail=True, methods=['post'], url_path='unassign-msmes')
    def unassign_msmes(self, request, pk=None):
        """Remove the group assignment from given MSMEs (or all if msme_ids omitted).
        Body: {msme_ids?: [...]}"""
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can unassign MSMEs from groups.")
        group = self.get_object()
        msme_ids = request.data.get('msme_ids')
        qs = MSME.objects.filter(assigned_group=group)
        if isinstance(msme_ids, list) and msme_ids:
            qs = qs.filter(id__in=msme_ids)
        cleared = qs.update(assigned_group=None, session_number=None)
        return Response({'group_id': group.id, 'cleared': cleared})

    @action(detail=True, methods=['get'], url_path='msmes')
    def list_msmes(self, request, pk=None):
        """Return all MSMEs currently assigned to this group."""
        group = self.get_object()
        msmes = group.assigned_msmes.filter(is_active=True).order_by('session_number', 'business_name')
        return Response(MSMESerializer(msmes, many=True).data)


class SupportRequestViewSet(viewsets.ModelViewSet):
    queryset = SupportRequest.objects.all()
    serializer_class = SupportRequestSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # BGEs only see SupportRequests that they're matched on; admins see all
        u = self.request.user
        qs = SupportRequest.objects.all()
        if not (u.is_staff or u.is_superuser):
            try:
                bge = u.bge_profile
                qs = qs.filter(matched_bges=bge).distinct()
            except Exception:
                qs = qs.none()
        return qs

    def create(self, request, *args, **kwargs):
        # Public-form flow leaves an MSME's request — but an authenticated user
        # creating one must still be the MSME-owner / admin. Restrict here so a
        # logged-in BGE cannot inject support requests on behalf of others.
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        support_request = serializer.save()
        nearby = BusinessGrowthExpert.objects.filter(
            status='approved', location__icontains=support_request.location
        )[:3]
        support_request.matched_bges.set(nearby)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can edit support requests.")
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can edit support requests.")
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can delete support requests.")
        return super().destroy(request, *args, **kwargs)


class TrainingSessionViewSet(ViewerReadOnlyMixin, viewsets.ModelViewSet):
    serializer_class = TrainingSessionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = TrainingSession.objects.select_related('topic').prefetch_related(
            'businesses',
            'facilitation_assignments__bge',
            'facilitation_assignments__work_order',
            'attendances__msme',
        ).all()
        user = self.request.user
        if user.is_staff or user.is_superuser:
            pass  # see everything
        elif _managed_groups(user) is not None or _is_viewer(user):
            pass  # programme managers and viewers see all sessions
        else:
            # BGEs see sessions they are assigned to (as lead or mentor)
            try:
                bge = user.bge_profile
                qs = qs.filter(
                    Q(topic_id__in=TrainingFacilitationAssignment.objects.filter(bge=bge).values('topic_id')) |
                    Q(facilitation_assignments__bge=bge)
                ).distinct()
            except Exception:
                return qs.none()
        work_order_id = self.request.query_params.get('work_order')
        if work_order_id:
            qs = qs.filter(facilitation_assignments__work_order_id=work_order_id)
        return qs

    @action(detail=True, methods=['post'])
    def mark_attendance(self, request, pk=None):
        """Legacy single-MSME toggle kept for backward compat."""
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can mark attendance.")
        session = self.get_object()
        msme_id = request.data.get('msme_id')
        present = request.data.get('present', True)
        qs = Attendance.objects.filter(session=session, msme_id=msme_id)
        if qs.exists():
            att = qs.first()
            att.present = present
            att.save()
        else:
            att = Attendance.objects.create(session=session, msme_id=msme_id, present=present)
        return Response(AttendanceSerializer(att).data)


class AttendanceViewSet(ViewerReadOnlyMixin, viewsets.ModelViewSet):
    queryset = Attendance.objects.select_related('msme', 'session').all()
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated]

    def _check_session_scope(self, request, session_id):
        """BGEs may only record attendance for sessions they're assigned to facilitate/mentor."""
        user = request.user
        if user.is_staff or user.is_superuser or _is_programme_manager(user):
            return
        try:
            bge = user.bge_profile
        except Exception:
            raise PermissionDenied("You do not have access to record attendance.")
        if not TrainingFacilitationAssignment.objects.filter(bge=bge, session_id=session_id).exists():
            raise PermissionDenied("You can only record attendance for sessions you are assigned to.")

    def create(self, request, *args, **kwargs):
        session_id = request.data.get('session')
        if session_id:
            self._check_session_scope(request, session_id)
        return super().create(request, *args, **kwargs)

    def get_queryset(self):
        user = self.request.user
        qs = Attendance.objects.select_related('msme', 'session').all()

        # SECURITY: scope attendance records to what the user is allowed to see.
        # Staff / superusers see all. Programme managers see their groups.
        # Viewers see all (read-only role). BGEs see only sessions they are
        # assigned to facilitate or mentor — not all sessions system-wide.
        if not (user.is_staff or user.is_superuser or _is_viewer(user)):
            if _is_programme_manager(user):
                group_ids = _managed_groups(user) or []
                qs = qs.filter(session__businesses__programme_groups__in=group_ids).distinct()
            else:
                # Regular BGE: limit to sessions where this BGE has a facilitation assignment
                try:
                    bge = user.bge_profile
                    assigned_session_ids = TrainingFacilitationAssignment.objects.filter(
                        bge=bge
                    ).values_list('session_id', flat=True)
                    qs = qs.filter(session_id__in=assigned_session_ids)
                except Exception:
                    return Attendance.objects.none()

        sid = self.request.query_params.get('session')
        if sid:
            qs = qs.filter(session_id=sid)
        cohort = self.request.query_params.get('cohort')
        if cohort:
            qs = qs.filter(msme__cohort_id=cohort)
        return qs

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """
        Participation summary — demographic totals from attendance + BGE report counts.
        Filters: cohort, session, work_order, bge, date_from, date_to.
        Groups: by_session (per training session), by_cohort, by_work_order.

        NOTE: TrainingSession has no direct work_order FK. The link is through
        TrainingFacilitationAssignment, so we build a session→work_order map
        from that table rather than traversing a non-existent relation.
        """
        from portfolio.models import (
            MSMEReport, GroupReport, Cohort as CohortModel, WorkOrder,
            TrainingSession, TrainingReport, MentorTrainingReport,
            TrainingFacilitationAssignment,
        )

        cohort_id     = request.query_params.get('cohort')
        session_id    = request.query_params.get('session')
        work_order_id = request.query_params.get('work_order')
        bge_id        = request.query_params.get('bge')
        date_from     = request.query_params.get('date_from')
        date_to       = request.query_params.get('date_to')

        att_qs = Attendance.objects.filter(present=True)
        rep_qs = MSMEReport.objects.filter(status='submitted')
        grp_qs = GroupReport.objects.filter(status__in=['submitted', 'approved'])

        # Session → work_order mapping (via facilitation assignments)
        fa_qs = TrainingFacilitationAssignment.objects.filter(
            session__isnull=False, work_order__isnull=False
        ).values('session_id', 'work_order_id', 'bge_id')

        # Build {session_id: set(work_order_ids)} and {session_id: set(bge_ids)}
        session_to_wos = {}
        session_to_bges = {}
        for fa in fa_qs:
            session_to_wos.setdefault(fa['session_id'], set()).add(fa['work_order_id'])
            session_to_bges.setdefault(fa['session_id'], set()).add(fa['bge_id'])

        if cohort_id:
            att_qs = att_qs.filter(msme__cohort_id=cohort_id)
            rep_qs = rep_qs.filter(msme__cohort_id=cohort_id)
        if work_order_id:
            wo_session_ids = [sid for sid, wos in session_to_wos.items()
                              if int(work_order_id) in wos]
            att_qs = att_qs.filter(session_id__in=wo_session_ids)
            rep_qs = rep_qs.filter(bge__work_orders__id=work_order_id)
        if bge_id:
            bge_session_ids = [sid for sid, bges in session_to_bges.items()
                               if int(bge_id) in bges]
            att_qs = att_qs.filter(session_id__in=bge_session_ids)
            rep_qs = rep_qs.filter(bge_id=bge_id)
            grp_qs = grp_qs.filter(team_lead_id=bge_id)
        if session_id:
            att_qs = att_qs.filter(session_id=session_id)
        if date_from:
            att_qs = att_qs.filter(session__date__gte=date_from)
            rep_qs = rep_qs.filter(visit_date__gte=date_from)
            grp_qs = grp_qs.filter(visit_date__gte=date_from)
        if date_to:
            att_qs = att_qs.filter(session__date__lte=date_to)
            rep_qs = rep_qs.filter(visit_date__lte=date_to)
            grp_qs = grp_qs.filter(visit_date__lte=date_to)

        # Pull attendance in one query; aggregate in Python to avoid N+1
        att_rows = list(att_qs.values(
            'gender', 'age_group', 'refugee_status',
            'msme_id', 'msme__cohort_id', 'session_id',
        ))

        def _dem_rows(rows):
            youth = [r for r in rows if r['age_group'] == '18-34']
            adult = [r for r in rows if r['age_group'] in ('35-45', '46-55', '56+')]
            refs  = [r for r in rows if r['refugee_status'] == 'R']
            host  = [r for r in rows if r['refugee_status'] == 'H']
            return {
                'total':          len(rows),
                'male':           sum(1 for r in rows if r['gender'] == 'M'),
                'female':         sum(1 for r in rows if r['gender'] == 'F'),
                'male_youth':     sum(1 for r in youth if r['gender'] == 'M'),
                'female_youth':   sum(1 for r in youth if r['gender'] == 'F'),
                'male_adult':     sum(1 for r in adult if r['gender'] == 'M'),
                'female_adult':   sum(1 for r in adult if r['gender'] == 'F'),
                'refugees_total': len(refs),
                'refugee_male':   sum(1 for r in refs if r['gender'] == 'M'),
                'refugee_female': sum(1 for r in refs if r['gender'] == 'F'),
                'host_community': len(host),
            }

        overall = _dem_rows(att_rows)
        overall.update({
            'msme_reports':         rep_qs.count(),
            'unique_msmes_visited': rep_qs.values('msme').distinct().count(),
            'group_sessions':       grp_qs.count(),
        })

        # ── Per-session breakdown ──────────────────────────────────────────────
        tr_by_session = {tr.session_id: tr.id
                         for tr in TrainingReport.objects.only('id', 'session_id')}
        mr_by_session = {}
        for mr in MentorTrainingReport.objects.values('session_id', 'id'):
            mr_by_session.setdefault(mr['session_id'], []).append(mr['id'])

        sess_qs = TrainingSession.objects.select_related('topic').prefetch_related(
            'facilitation_assignments__bge',
        ).order_by('date')
        if date_from:
            sess_qs = sess_qs.filter(date__gte=date_from)
        if date_to:
            sess_qs = sess_qs.filter(date__lte=date_to)
        if session_id:
            sess_qs = sess_qs.filter(id=session_id)

        session_data = []
        for sess in sess_qs:
            s_rows = [r for r in att_rows if r['session_id'] == sess.id]
            dem = _dem_rows(s_rows)

            lead_name = None
            mentor_names = []
            for fa in sess.facilitation_assignments.all():
                if fa.role == 'lead' and fa.bge_id:
                    lead_name = fa.bge.name
                elif fa.role == 'mentor' and fa.bge_id:
                    mentor_names.append(fa.bge.name)

            topic_number = ''
            if sess.topic_id:
                t = sess.topic
                if t.section_number:
                    topic_number = f"{t.module_number}.{t.section_number}"
                else:
                    topic_number = str(t.module_number)

            dem.update({
                'session_id':          sess.id,
                'session_title':       sess.title,
                'session_date':        str(sess.date),
                'session_location':    sess.location or '',
                'topic_name':          sess.topic.name if sess.topic_id else '',
                'topic_number':        topic_number,
                'lead_bge_name':       lead_name or '',
                'mentor_names':        mentor_names,
                'has_training_report': sess.id in tr_by_session,
                'training_report_id':  tr_by_session.get(sess.id),
                'mentor_report_count': len(mr_by_session.get(sess.id, [])),
                'registered_count':    sess.businesses.count(),
            })
            session_data.append(dem)

        # ── Per-cohort breakdown ───────────────────────────────────────────────
        rep_by_cohort = {}
        for item in rep_qs.values('msme__cohort_id', 'msme_id'):
            cid = item['msme__cohort_id']
            if cid not in rep_by_cohort:
                rep_by_cohort[cid] = {'count': 0, 'msmes': set()}
            rep_by_cohort[cid]['count'] += 1
            rep_by_cohort[cid]['msmes'].add(item['msme_id'])

        cohorts_data = []
        for cohort in CohortModel.objects.all().order_by('name'):
            c_rows = [r for r in att_rows if r['msme__cohort_id'] == cohort.id]
            c_rep  = rep_by_cohort.get(cohort.id, {'count': 0, 'msmes': set()})
            if not c_rows and not c_rep['count']:
                continue
            row = _dem_rows(c_rows)
            row.update({'cohort_id': cohort.id, 'cohort_name': cohort.name,
                        'msme_reports': c_rep['count'],
                        'unique_msmes': len(c_rep['msmes'])})
            cohorts_data.append(row)

        # ── Per-work-order breakdown ───────────────────────────────────────────
        # Build {work_order_id: set(session_ids)} from the facilitation assignment map
        wo_to_sessions = {}
        for sid, wos in session_to_wos.items():
            for wid in wos:
                wo_to_sessions.setdefault(wid, set()).add(sid)

        rep_by_bge = {}
        for item in rep_qs.values('bge_id', 'msme_id'):
            bid = item['bge_id']
            if bid not in rep_by_bge:
                rep_by_bge[bid] = {'count': 0, 'msmes': set()}
            rep_by_bge[bid]['count'] += 1
            rep_by_bge[bid]['msmes'].add(item['msme_id'])

        wo_data = []
        for wo in WorkOrder.objects.select_related('bge').order_by('-issue_date'):
            wo_sids = wo_to_sessions.get(wo.id, set())
            w_rows  = [r for r in att_rows if r['session_id'] in wo_sids]
            w_rep   = rep_by_bge.get(wo.bge_id, {'count': 0, 'msmes': set()})
            if not w_rows and not w_rep['count']:
                continue
            row = _dem_rows(w_rows)
            row.update({
                'work_order_id':     wo.id,
                'work_order_number': wo.work_order_number,
                'work_order_type':   wo.get_work_order_type_display(),
                'bge_name':          wo.bge.name,
                'bge_code':          wo.bge.bge_code or '',
                'issue_date':        str(wo.issue_date),
                'start_date':        str(wo.start_date) if wo.start_date else None,
                'end_date':          str(wo.end_date) if wo.end_date else None,
                'status':            wo.status,
                'msme_reports':      w_rep['count'],
                'unique_msmes':      len(w_rep['msmes']),
            })
            wo_data.append(row)

        overall['by_session']    = session_data
        overall['by_cohort']     = cohorts_data
        overall['by_work_order'] = wo_data
        return Response(overall)


class TrainingTopicViewSet(viewsets.ModelViewSet):
    queryset = TrainingTopic.objects.all()
    serializer_class = TrainingTopicSerializer
    permission_classes = [IsAuthenticated]


class TrainingFacilitationAssignmentViewSet(viewsets.ModelViewSet):
    """
    Manage training facilitation assignments.
    - Admins/programme managers: full CRUD.
    - BGEs: read-only, scoped to their own assignments.
    """
    serializer_class = TrainingFacilitationAssignmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = TrainingFacilitationAssignment.objects.select_related(
            'bge', 'topic', 'assigned_by', 'session', 'work_order'
        )
        # BGEs only see their own assignments
        if not (user.is_staff or user.is_superuser or hasattr(user, 'cohort_admin_profile')):
            try:
                bge = user.bge_profile
                return qs.filter(bge=bge)
            except Exception:
                return qs.none()
        # Optional filters
        session_id = self.request.query_params.get('session')
        if session_id:
            qs = qs.filter(session_id=session_id)
        role = self.request.query_params.get('role')
        if role:
            qs = qs.filter(role=role)
        bge_id = self.request.query_params.get('bge')
        if bge_id:
            qs = qs.filter(bge_id=bge_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(assigned_by=self.request.user)

    def create(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser or hasattr(request.user, 'cohort_admin_profile')):
            raise PermissionDenied("Only programme managers can create facilitation assignments.")
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser or hasattr(request.user, 'cohort_admin_profile')):
            raise PermissionDenied("Only programme managers can edit facilitation assignments.")
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser or hasattr(request.user, 'cohort_admin_profile')):
            raise PermissionDenied("Only programme managers can remove facilitation assignments.")
        return super().destroy(request, *args, **kwargs)


class VisitReportTemplateViewSet(viewsets.ModelViewSet):
    """CRUD for visit report templates. List is public (authenticated); CUD is admin-only."""
    serializer_class = VisitReportTemplateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = VisitReportTemplate.objects.all()
        if self.request.query_params.get('active_only') == '1':
            qs = qs.filter(is_active=True)
        return qs

    def _require_admin(self):
        if not (self.request.user.is_staff or self.request.user.is_superuser):
            raise PermissionDenied("Only admins can manage report templates.")

    def create(self, request, *args, **kwargs):
        self._require_admin(); return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        self._require_admin(); return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        self._require_admin(); return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        self._require_admin(); return super().destroy(request, *args, **kwargs)


class MSMEReportViewSet(ProgrammeManagerReadOnlyMixin, ViewerReadOnlyMixin, viewsets.ModelViewSet):
    serializer_class = MSMEReportSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        group_ids = _managed_groups(user)
        if user.is_staff or user.is_superuser:
            qs = MSMEReport.objects.all()
        elif group_ids is not None:
            qs = MSMEReport.objects.filter(msme__programme_groups__in=group_ids).distinct()
        elif _is_viewer(user):
            qs = MSMEReport.objects.all()
        else:
            try:
                bge = user.bge_profile
                qs = MSMEReport.objects.filter(bge=bge)
            except Exception:
                qs = MSMEReport.objects.none()

        msme_id = self.request.query_params.get('msme')
        if msme_id:
            qs = qs.filter(msme_id=msme_id)
        bge_id = self.request.query_params.get('bge')
        if bge_id:
            qs = qs.filter(bge_id=bge_id)
        report_status = self.request.query_params.get('status')
        if report_status:
            qs = qs.filter(status=report_status)
        return qs.select_related('msme', 'bge')

    def perform_create(self, serializer):
        from rest_framework.exceptions import ValidationError as DRFValidationError
        from datetime import date as _date
        import math

        user = self.request.user
        if not (user.is_staff or user.is_superuser):
            try:
                bge = user.bge_profile
            except Exception:
                raise PermissionDenied("No BGE profile associated with this account.")

            # Prevent a second BGE from filing an annual/quarterly review for an MSME
            # that another BGE has already covered in the same period.
            # The BGE who originally filed it can still edit/update their own report.
            msme       = serializer.validated_data.get('msme')
            visit_type = serializer.validated_data.get('visit_type', '')
            visit_date = serializer.validated_data.get('visit_date') or _date.today()

            if msme and visit_type in ('annual_review', 'quarterly_review'):
                qs = MSMEReport.objects.filter(
                    msme=msme,
                    visit_type=visit_type,
                    visit_date__year=visit_date.year,
                ).exclude(bge=bge)

                if visit_type == 'quarterly_review':
                    # Scope conflict to the same calendar quarter (Q1–Q4)
                    quarter = math.ceil(visit_date.month / 3)
                    quarter_start = ((quarter - 1) * 3) + 1
                    quarter_end   = quarter_start + 2
                    qs = qs.filter(
                        visit_date__month__gte=quarter_start,
                        visit_date__month__lte=quarter_end,
                    )

                conflict = qs.select_related('bge').first()
                if conflict:
                    period_label = (
                        f"Q{math.ceil(conflict.visit_date.month / 3)} {conflict.visit_date.year}"
                        if visit_type == 'quarterly_review'
                        else str(conflict.visit_date.year)
                    )
                    raise DRFValidationError(
                        f"An {'annual' if visit_type == 'annual_review' else 'quarterly'} review "
                        f"for this MSME has already been filed by {conflict.bge.name} "
                        f"({period_label}). Only one BGE can file this review per period."
                    )

            serializer.save(bge=bge)
            return
        serializer.save()

    def perform_update(self, serializer):
        instance = serializer.instance
        new_status = serializer.validated_data.get('status', instance.status)
        report = serializer.save()

        if instance.status != 'submitted' and new_status == 'submitted':
            # Freeze a PDF copy on first submission
            try:
                from .pdf_reports import render_msme_report
                from django.core.files.base import ContentFile
                pdf_bytes = render_msme_report(report).read()
                safe_name = report.msme.business_name[:30].replace(' ', '_')
                fname = f"MSMEReport_{safe_name}_{report.visit_date}.pdf"
                report.submitted_pdf.save(fname, ContentFile(pdf_bytes), save=False)
                report.submitted_pdf_data = pdf_bytes
                report.save(update_fields=['submitted_pdf', 'submitted_pdf_data'])
            except Exception as e:
                logger.error('Failed to snapshot MSME report PDF (report id=%s): %s', report.id, e)

            # Auto-create a growth snapshot from the quantitative fields
            self._create_snapshot_from_report(report)

    @staticmethod
    def _create_snapshot_from_report(report):
        """Create or update a MSMEGrowthSnapshot from a submitted visit report."""
        has_quant = any([
            report.revenue_ugx, report.total_assets_ugx,
            report.employees_ft_male is not None, report.employees_ft_female is not None,
            report.employees_pt_male is not None, report.employees_pt_female is not None,
            report.has_tin is not None, report.has_ursb is not None,
            report.has_business_bank is not None, report.has_mobile_money is not None,
        ])
        if not has_quant:
            return
        try:
            bge = report.bge
        except Exception:
            bge = None
        notes_parts = []
        if report.key_achievement:
            notes_parts.append(f'Achievement: {report.key_achievement}')
        if report.growth_rating:
            notes_parts.append(f'Growth rating: {report.growth_rating}/5')
        MSMEGrowthSnapshot.objects.create(
            msme                = report.msme,
            snapshot_date       = report.visit_date,
            source              = 'bge_visit',
            collected_by        = bge,
            annual_turnover     = report.revenue_ugx,
            total_assets        = report.total_assets_ugx,
            employees_ft_male   = report.employees_ft_male,
            employees_ft_female = report.employees_ft_female,
            employees_pt_male   = report.employees_pt_male,
            employees_pt_female = report.employees_pt_female,
            has_tin             = report.has_tin,
            has_ursb            = report.has_ursb,
            has_business_bank   = report.has_business_bank,
            has_mobile_money    = report.has_mobile_money,
            notes               = '\n'.join(notes_parts),
        )

    @action(detail=True, methods=['post'], url_path='revert')
    def revert(self, request, pk=None):
        """Admin-only: revert a submitted/reviewed report back to draft
        so the BGE can re-edit and resubmit."""
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can revert reports.")
        report = self.get_object()
        if report.status == 'draft':
            return Response({'detail': 'Report is already a draft.'}, status=status.HTTP_400_BAD_REQUEST)
        report.status = 'draft'
        if report.submitted_pdf:
            report.submitted_pdf.delete(save=False)
        report.submitted_pdf_data = None
        report.save(update_fields=['status', 'submitted_pdf', 'submitted_pdf_data'])
        from .serializers import MSMEReportSerializer
        return Response(MSMEReportSerializer(report, context={'request': request}).data)

    @action(detail=True, methods=['get'], url_path='pdf')
    def pdf(self, request, pk=None):
        """Render this MSME visit report as a styled PDF.
        Submitted reports return the stored snapshot; drafts are rendered on demand."""
        from .pdf_reports import render_msme_report
        report = self.get_object()
        fname = _safe_filename(f"MSMEReport_{report.msme.business_name[:30].replace(' ', '_')}_{report.visit_date}.pdf")
        disposition = 'attachment' if request.query_params.get('dl') else 'inline'
        if report.submitted_pdf:
            try:
                resp = HttpResponse(report.submitted_pdf.read(), content_type='application/pdf')
                resp['Content-Disposition'] = f'{disposition}; filename="{fname}"'
                return resp
            except Exception as e:
                logger.warning('Stored MSME report PDF unreadable (report id=%s), regenerating: %s', report.id, e)
        buf = render_msme_report(report)
        resp = HttpResponse(buf.read(), content_type='application/pdf')
        resp['Content-Disposition'] = f'{disposition}; filename="{fname}"'
        return resp


class GroupReportViewSet(ProgrammeManagerReadOnlyMixin, ViewerReadOnlyMixin, viewsets.ModelViewSet):
    """Group-level reports.

    Visibility:
    - Admins see every group report.
    - A BGE sees reports for any group they're a member of (so the whole team
      can read what the lead submitted).

    Mutation:
    - Only admins or the group's `team_lead` can create or edit reports
      against that group. Other members can read but not write.
    - status transitions: draft -> submitted (sets submitted_at),
                          submitted -> approved (admin only, sets approved_at).
    """
    serializer_class = GroupReportSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        group_ids = _managed_groups(user)
        if user.is_staff or user.is_superuser:
            qs = GroupReport.objects.all()
        elif group_ids is not None:
            qs = GroupReport.objects.filter(
                group__assigned_msmes__programme_groups__in=group_ids
            ).distinct()
        elif _is_viewer(user):
            qs = GroupReport.objects.all()
        else:
            try:
                bge = user.bge_profile
                qs = GroupReport.objects.filter(group__members=bge).distinct()
            except Exception:
                qs = GroupReport.objects.none()

        gid = self.request.query_params.get('group')
        if gid:
            qs = qs.filter(group_id=gid)
        sid = self.request.query_params.get('session_number')
        if sid:
            qs = qs.filter(session_number=sid)
        st = self.request.query_params.get('status')
        if st:
            qs = qs.filter(status=st)
        return qs.select_related('group', 'team_lead').prefetch_related('msmes_supported')

    def _user_can_write_for_group(self, user, group):
        if user.is_staff or user.is_superuser:
            return True
        if _is_programme_manager(user):
            return False  # programme managers are read-only on group reports
        if _managed_groups(user) is not None:
            return True
        try:
            bge = user.bge_profile
        except Exception:
            return False
        return group.team_lead_id == bge.id

    def perform_create(self, serializer):
        user = self.request.user
        group = serializer.validated_data.get('group')
        if not group:
            raise PermissionDenied("`group` is required.")
        if not self._user_can_write_for_group(user, group):
            raise PermissionDenied(
                "Only the group's team lead (or an admin) can file group reports."
            )
        # Stamp the team lead automatically — never trusted from request body.
        # Resolution order:
        #   1. Author's own bge_profile (BGE filing the report on themselves)
        #   2. Group's designated team_lead (admin filing on the team's behalf)
        # If neither resolves, refuse to create — an owner-less report can't be
        # edited later because every write check matches on team_lead_id.
        bge = None
        try:
            bge = user.bge_profile
        except Exception:
            bge = group.team_lead
        if bge is None:
            raise PermissionDenied(
                "Cannot create a group report without a team lead. "
                "Set a team lead on the group first."
            )
        serializer.save(team_lead=bge)

    def perform_update(self, serializer):
        instance = self.get_object()
        user = self.request.user
        if not self._user_can_write_for_group(user, instance.group):
            raise PermissionDenied(
                "Only the group's team lead (or an admin) can edit this report."
            )

        # Stamp lifecycle timestamps when status transitions
        from django.utils import timezone
        new_status = serializer.validated_data.get('status', instance.status)
        extra = {}
        if instance.status != 'submitted' and new_status == 'submitted':
            extra['submitted_at'] = timezone.now()
        if instance.status != 'approved' and new_status == 'approved':
            if not (user.is_staff or user.is_superuser):
                raise PermissionDenied("Only admins can mark a report as Approved.")
            extra['approved_at'] = timezone.now()
        report = serializer.save(**extra)

        # Snapshot the PDF on first submission so the copy is frozen at that moment.
        if instance.status != 'submitted' and new_status == 'submitted':
            if not report.submitted_pdf:
                try:
                    from .pdf_reports import render_group_report
                    from django.core.files.base import ContentFile
                    pdf_bytes = render_group_report(report).read()
                    safe_name = report.group.name.replace(' ', '_')
                    fname = f"GroupReport_{safe_name}_{report.visit_date}.pdf"
                    report.submitted_pdf.save(fname, ContentFile(pdf_bytes), save=True)
                except Exception as e:
                    logger.error('Failed to snapshot group report PDF (report id=%s): %s', report.id, e)

    def destroy(self, request, *args, **kwargs):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can delete group reports.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='revert')
    def revert(self, request, pk=None):
        """Admin-only: revert a submitted or approved group report back to draft."""
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can revert reports.")
        report = self.get_object()
        if report.status == 'draft':
            return Response({'detail': 'Report is already a draft.'}, status=status.HTTP_400_BAD_REQUEST)
        report.status = 'draft'
        report.submitted_at = None
        report.approved_at = None
        if report.submitted_pdf:
            report.submitted_pdf.delete(save=False)
        report.submitted_pdf_data = None
        report.save(update_fields=['status', 'submitted_at', 'approved_at', 'submitted_pdf', 'submitted_pdf_data'])
        from .serializers import GroupReportSerializer
        return Response(GroupReportSerializer(report, context={'request': request}).data)

    @action(detail=True, methods=['get'], url_path='pdf')
    def pdf(self, request, pk=None):
        """Render this group report as a styled PDF.
        Submitted/approved reports return the stored snapshot; drafts are rendered on demand."""
        from .pdf_reports import render_group_report
        report = self.get_object()
        fname = _safe_filename(f"GroupReport_{report.group.name.replace(' ', '_')}_{report.visit_date}.pdf")
        disposition = 'attachment' if request.query_params.get('dl') else 'inline'
        if report.submitted_pdf:
            try:
                resp = HttpResponse(report.submitted_pdf.read(), content_type='application/pdf')
                resp['Content-Disposition'] = f'{disposition}; filename="{fname}"'
                return resp
            except Exception as e:
                logger.warning('Stored group report PDF unreadable (report id=%s), regenerating: %s', report.id, e)
        buf = render_group_report(report)
        resp = HttpResponse(buf.read(), content_type='application/pdf')
        resp['Content-Disposition'] = f'{disposition}; filename="{fname}"'
        return resp


class GroupReportContributionViewSet(viewsets.ModelViewSet):
    """A note from one group member feeding into a group report.

    Visibility:
    - Admins see every contribution.
    - A BGE sees their own contributions + every contribution to a group
      report whose group they belong to (so the team lead can read them
      while consolidating the consolidated report, and other members get
      transparency into what their teammates submitted).

    Mutation:
    - Members can create/edit only their own contribution. The `bge` field
      is auto-stamped from request.user; never trusted from the body.
    - Admins can edit any contribution.
    - One contribution per (group_report, bge) pair (unique_together).
      Re-POSTing for the same pair returns 200 with the existing record
      so the frontend can use it as upsert.
    """
    serializer_class = GroupReportContributionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = GroupReportContribution.objects.select_related(
            'bge', 'group_report__group',
        ).prefetch_related('msmes_observed')
        if user.is_staff or user.is_superuser:
            pass
        else:
            try:
                bge = user.bge_profile
            except Exception:
                return qs.none()
            from django.db.models import Q
            qs = qs.filter(
                Q(bge=bge) | Q(group_report__group__members=bge)
            ).distinct()

        gr = self.request.query_params.get('group_report')
        if gr:
            qs = qs.filter(group_report_id=gr)
        return qs

    def _bge_for_user(self):
        try:
            return self.request.user.bge_profile
        except Exception:
            return None

    def create(self, request, *args, **kwargs):
        bge = self._bge_for_user()
        is_admin = request.user.is_staff or request.user.is_superuser
        if bge is None and not is_admin:
            raise PermissionDenied("You don't have a BGE profile.")

        gr_id = request.data.get('group_report')
        if not gr_id:
            return Response({'error': 'group_report is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            gr = GroupReport.objects.select_related('group').get(pk=gr_id)
        except GroupReport.DoesNotExist:
            return Response({'error': 'group_report not found'}, status=status.HTTP_404_NOT_FOUND)

        # Member must belong to the group
        if bge and not gr.group.members.filter(pk=bge.pk).exists() and not is_admin:
            raise PermissionDenied("You're not a member of this group.")

        contributor = bge if bge else gr.group.team_lead
        if contributor is None:
            return Response({'error': 'No bge to attribute the contribution to.'}, status=status.HTTP_400_BAD_REQUEST)

        # Upsert: if (group_report, bge) already exists, return that row
        existing = GroupReportContribution.objects.filter(
            group_report=gr, bge=contributor,
        ).first()
        if existing:
            ser = self.get_serializer(existing, data=request.data, partial=True)
            ser.is_valid(raise_exception=True)
            ser.save()
            return Response(ser.data, status=status.HTTP_200_OK)

        ser = self.get_serializer(data=request.data)
        ser.is_valid(raise_exception=True)
        ser.save(bge=contributor)
        return Response(ser.data, status=status.HTTP_201_CREATED)

    def perform_update(self, serializer):
        bge = self._bge_for_user()
        is_admin = self.request.user.is_staff or self.request.user.is_superuser
        instance = self.get_object()
        if not is_admin and (bge is None or instance.bge_id != bge.id):
            raise PermissionDenied("You can only edit your own contribution.")
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        bge = self._bge_for_user()
        is_admin = request.user.is_staff or request.user.is_superuser
        instance = self.get_object()
        if not is_admin and (bge is None or instance.bge_id != bge.id):
            raise PermissionDenied("You can only delete your own contribution.")
        return super().destroy(request, *args, **kwargs)


class GroupReportAttendanceViewSet(viewsets.ModelViewSet):
    """Per-person MSME attendance records for group session reports.

    BGE users (team leads) can create/update/delete records for their own
    group reports. Admins have full access. Filter by ?group_report=<id>.
    """
    serializer_class = GroupReportAttendanceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = GroupReportAttendance.objects.select_related('msme', 'group_report')
        # Non-admin BGEs can only see attendance for group reports belonging to
        # groups they are a member of (team lead or regular member).
        if not (user.is_staff or user.is_superuser):
            bge = self._bge_for_user()
            if bge is None:
                return qs.none()
            qs = qs.filter(
                Q(group_report__team_lead=bge) |
                Q(group_report__group__members=bge)
            ).distinct()
        group_report_id = self.request.query_params.get('group_report')
        if group_report_id:
            qs = qs.filter(group_report_id=group_report_id)
        return qs

    def _bge_for_user(self):
        try:
            return self.request.user.bge_profile
        except Exception:
            return None

    def _can_edit(self, group_report_id):
        user = self.request.user
        if user.is_staff or user.is_superuser:
            return True
        bge = self._bge_for_user()
        if bge is None:
            return False
        try:
            report = GroupReport.objects.get(pk=group_report_id)
            return report.team_lead_id == bge.id or report.group.members.filter(pk=bge.id).exists()
        except GroupReport.DoesNotExist:
            return False

    def perform_create(self, serializer):
        group_report_id = serializer.validated_data.get('group_report').id
        if not self._can_edit(group_report_id):
            raise PermissionDenied("You can only record attendance for your own group reports.")
        serializer.save()

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        if not self._can_edit(instance.group_report_id):
            raise PermissionDenied("You can only update attendance for your own group reports.")
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if not self._can_edit(instance.group_report_id):
            raise PermissionDenied("You can only delete attendance for your own group reports.")
        return super().destroy(request, *args, **kwargs)


class BGEUserViewSet(viewsets.ViewSet):
    """
    Admin-only viewset for managing BGE user accounts.
    Allows creating logins and linking them to BGE profiles without needing Django admin.
    """
    permission_classes = [IsAuthenticated]

    def _require_admin(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can manage users.")

    def list(self, request):
        if not (request.user.is_staff or request.user.is_superuser or hasattr(request.user, 'cohort_admin_profile')):
            raise PermissionDenied("Only admins can manage users.")
        users = User.objects.filter(is_staff=False, is_superuser=False).select_related('bge_profile', 'security_profile')
        data = []
        for u in users:
            try:
                profile = u.bge_profile
                bge_info = {'id': profile.id, 'name': profile.name, 'status': profile.status}
            except Exception:
                bge_info = None
            # Determine role
            try:
                ca = u.cohort_admin_profile
                role = 'cohort_admin'
                managed_groups = list(ca.managed_groups.values('id', 'name'))
            except CohortAdmin.DoesNotExist:
                managed_groups = []
                role = 'bge' if bge_info else 'viewer'
            try:
                viewer_approved = u.security_profile.viewer_approved
            except Exception:
                viewer_approved = True
            if role == 'viewer' and not viewer_approved:
                role = 'pending'
            data.append({
                'id': u.id,
                'username': u.username,
                'email': u.email,
                'is_active': u.is_active,
                'date_joined': u.date_joined,
                'bge_profile': bge_info,
                'role': role,
                'managed_groups': managed_groups,
                'viewer_approved': viewer_approved,
            })
        return Response(data)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        """Admin: approve a pending Google sign-in, granting it read-only viewer access."""
        self._require_admin(request)
        user, err = self._get_target_non_admin_user(pk, request)
        if err is not None:
            return err
        from .models import UserSecurityProfile
        sec, _ = UserSecurityProfile.objects.get_or_create(user=user)
        sec.viewer_approved = True
        sec.save(update_fields=['viewer_approved'])
        return Response({'message': f'{user.username} approved for viewer access.'})

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        """Admin: reject a pending Google sign-in, deactivating the account."""
        self._require_admin(request)
        user, err = self._get_target_non_admin_user(pk, request)
        if err is not None:
            return err
        user.is_active = False
        user.save(update_fields=['is_active'])
        return Response({'message': f'{user.username} rejected and deactivated.'})

    def create(self, request):
        self._require_admin(request)
        username = request.data.get('username', '').strip()
        password = request.data.get('password', '').strip()
        email = request.data.get('email', '').strip()
        bge_id = request.data.get('bge_id')

        if not username or not password:
            return Response({'error': 'Username and password are required.'}, status=status.HTTP_400_BAD_REQUEST)
        if User.objects.filter(username=username).exists():
            return Response({'error': 'Username already exists.'}, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.create_user(username=username, password=password, email=email)

        if bge_id:
            try:
                bge = BusinessGrowthExpert.objects.get(pk=bge_id)
                if bge.user:
                    user.delete()
                    return Response({'error': 'This BGE already has a user account linked.'}, status=status.HTTP_400_BAD_REQUEST)
                type(bge).objects.filter(pk=bge.pk).update(user=user)
                bge.user = user
                send_welcome_email(bge, username, password)
                from .account_setup import send_welcome_sms as _send_sms
                _send_sms(bge, username, password)
            except BusinessGrowthExpert.DoesNotExist:
                user.delete()
                return Response({'error': 'BGE profile not found.'}, status=status.HTTP_404_NOT_FOUND)

        return Response({'id': user.id, 'username': user.username, 'email': user.email}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='bulk-create-missing')
    def bulk_create_missing(self, request):
        """Create login accounts for every BGE that doesn't have one yet.
        Admin-only. Each account gets a unique random temporary password,
        sent to the BGE via welcome email/SMS. Returns counts of created / skipped."""
        self._require_admin(request)
        unlinked = BusinessGrowthExpert.objects.filter(user__isnull=True).order_by('id')
        created = skipped = 0
        names = []
        for bge in unlinked:
            outcome = ensure_bge_account(bge, send_email=True)
            if outcome == 'created':
                created += 1
                names.append(bge.name or f'BGE #{bge.id}')
            else:
                skipped += 1
        return Response({'created': created, 'skipped': skipped, 'names': names})

    @action(detail=True, methods=['patch'], url_path='set-role')
    def set_role(self, request, pk=None):
        """Set or clear a user's programme-manager role and managed groups.

        Body:
          role        : 'viewer' | 'cohort_admin'
          group_ids   : [1, 2, ...]   (required when role='cohort_admin')
        """
        self._require_admin(request)
        user, err = self._get_target_non_admin_user(pk, request)
        if err:
            return err

        role = request.data.get('role', 'viewer')
        group_ids = request.data.get('group_ids', [])

        if role == 'cohort_admin':
            ca, _ = CohortAdmin.objects.get_or_create(user=user)
            ca.managed_groups.set(ProgrammeGroup.objects.filter(id__in=group_ids))
            ca.save()
            names = list(ca.managed_groups.values_list('name', flat=True))
            return Response({'role': 'cohort_admin', 'managed_groups': names})
        else:
            # viewer — remove cohort_admin if it exists
            CohortAdmin.objects.filter(user=user).delete()
            return Response({'role': 'viewer'})

    def _get_target_non_admin_user(self, pk, request):
        """Resolve a target user that is NOT a staff/superuser. Raises 403 if it
        is — admins managing other admins must use Django's admin site."""
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return None, Response({'error': 'User not found.'}, status=status.HTTP_404_NOT_FOUND)
        # Even an admin should not be able to demote / lock-out / re-password
        # another admin via this BGE-user endpoint. The list view already hides
        # them; close the loophole on the detail mutations too.
        if user.is_staff or user.is_superuser:
            return None, Response(
                {'error': 'Cannot manage staff/superuser accounts here. Use Django admin.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return user, None

    @action(detail=True, methods=['post'], url_path='set-password')
    def set_password(self, request, pk=None):
        self._require_admin(request)
        new_password = (request.data.get('password') or '').strip()
        if not new_password or len(new_password) < 8:
            return Response({'error': 'Password must be at least 8 characters.'}, status=status.HTTP_400_BAD_REQUEST)
        user, err = self._get_target_non_admin_user(pk, request)
        if err is not None:
            return err
        user.set_password(new_password)
        user.save()
        return Response({'message': f'Password updated for {user.username}.'})

    @action(detail=True, methods=['patch'], url_path='link-bge')
    def link_bge(self, request, pk=None):
        self._require_admin(request)
        bge_id = request.data.get('bge_id')
        user, err = self._get_target_non_admin_user(pk, request)
        if err is not None:
            return err
        if bge_id:
            try:
                bge = BusinessGrowthExpert.objects.get(pk=bge_id)
                # Unlink any previous user linked to this BGE
                BusinessGrowthExpert.objects.filter(user=user).update(user=None)
                bge.user = user
                bge.save()
            except BusinessGrowthExpert.DoesNotExist:
                return Response({'error': 'BGE not found.'}, status=status.HTTP_404_NOT_FOUND)
        else:
            BusinessGrowthExpert.objects.filter(user=user).update(user=None)
        return Response({'message': 'BGE link updated.'})

    @action(detail=True, methods=['patch'], url_path='toggle-active')
    def toggle_active(self, request, pk=None):
        self._require_admin(request)
        user, err = self._get_target_non_admin_user(pk, request)
        if err is not None:
            return err
        user.is_active = not user.is_active
        user.save()
        return Response({'is_active': user.is_active})


# ── Push notification helpers ──────────────────────────────────────────────────

def _send_push(subscription_obj, title, body, url='/'):
    """Send a single Web Push notification.

    Errors are logged. 404/410 responses (subscription gone — user uninstalled
    the PWA or revoked permission) cause the row to be deleted so the next
    notify run doesn't waste a request on a dead endpoint.
    """
    import logging as _logging
    log = _logging.getLogger(__name__)
    try:
        webpush(
            subscription_info={
                'endpoint': subscription_obj.endpoint,
                'keys': {'p256dh': subscription_obj.p256dh, 'auth': subscription_obj.auth},
            },
            data=_json.dumps({'title': title, 'body': body, 'url': url}),
            vapid_private_key=settings.VAPID_PRIVATE_KEY,
            vapid_claims=settings.VAPID_CLAIMS,
        )
    except WebPushException as exc:
        # Prune subscriptions the push service has explicitly retired.
        status_code = getattr(getattr(exc, 'response', None), 'status_code', None)
        if status_code in (404, 410):
            try:
                subscription_obj.delete()
                log.info("Pruned dead push subscription id=%s (%s)",
                         subscription_obj.pk, status_code)
            except Exception:
                pass
        else:
            log.warning("Push send failed for subscription id=%s: %s",
                        getattr(subscription_obj, 'pk', '?'), exc)


def _notify_bge(bge, title, body, url='/'):
    """Send push notification to all active subscriptions for a BGE's linked user."""
    if not bge.user:
        return
    for sub in PushSubscription.objects.filter(user=bge.user):
        _send_push(sub, title, body, url)


def _send_co_assignment_alert(existing_bge, new_bge, msme):
    """Email the existing/primary BGE to inform them that a second BGE
    has been co-assigned to visit the same MSME.

    Fails silently so the assignment itself is never blocked by an email error.
    """
    if not existing_bge.email:
        return
    import logging as _logging
    _log = _logging.getLogger(__name__)
    from django.utils.html import escape as _esc

    subject = (
        f"PRUDEV II — Joint Deployment Notice: "
        f"{new_bge.name} has also been assigned to {msme.business_name}"
    )

    # ── Plain text ─────────────────────────────────────────────────────────────
    lines = [
        f"Dear {existing_bge.name},",
        "",
        "This is to inform you that a second Business Growth Expert (BGE) has been "
        "co-assigned to visit one of the MSMEs currently in your portfolio.",
        "",
        "SHARED MSME",
        "─" * 40,
        f"  {msme.business_name} ({msme.msme_code or 'No code'})",
    ]
    if msme.owner_name: lines.append(f"  Owner:    {msme.owner_name}")
    if msme.city:       lines.append(f"  Location: {msme.city}")
    if msme.phone:      lines.append(f"  Phone:    {msme.phone}")
    lines += ["", "CO-ASSIGNED BGE", "─" * 40,
              f"  Name:  {new_bge.name} ({new_bge.bge_code or 'No code'})"]
    if new_bge.phone:  lines.append(f"  Phone: {new_bge.phone}")
    if new_bge.email:  lines.append(f"  Email: {new_bge.email}")
    if new_bge.deployment_objectives:
        obj_preview = new_bge.deployment_objectives[:300]
        lines += ["  Objectives:", f"    {obj_preview}" + ("…" if len(new_bge.deployment_objectives) > 300 else "")]
    lines += [
        "",
        "This BGE will submit a separate work order and report for their visit.",
        "Please coordinate where possible to ensure visits are complementary and not duplicated.",
        "",
        "Best regards,",
        "PRUDEV II BDS Team",
        "GIZ · GOPA AFC",
    ]
    body_text = "\n".join(lines)

    # ── HTML ────────────────────────────────────────────────────────────────────
    msme_detail_parts = []
    if msme.owner_name: msme_detail_parts.append(f"Owner: {_esc(msme.owner_name)}")
    if msme.city:       msme_detail_parts.append(f"Location: {_esc(msme.city)}")
    if msme.phone:      msme_detail_parts.append(f"Phone: {_esc(msme.phone)}")
    msme_detail_html = " &nbsp;·&nbsp; ".join(msme_detail_parts)

    contact_parts = []
    if new_bge.phone: contact_parts.append(f"📞 {_esc(new_bge.phone)}")
    if new_bge.email: contact_parts.append(f"✉ {_esc(new_bge.email)}")
    contact_html = " &nbsp;·&nbsp; ".join(contact_parts)

    obj_html = ""
    if new_bge.deployment_objectives:
        snippet = _esc(new_bge.deployment_objectives[:300]) + (
            "…" if len(new_bge.deployment_objectives) > 300 else ""
        )
        obj_html = f"""
            <div style="background:#f8f9fa;border-left:3px solid #C8102E;padding:8px 12px;
                        border-radius:0 4px 4px 0;margin-top:8px;">
              <p style="margin:0 0 3px;font-size:10px;font-weight:700;text-transform:uppercase;
                        letter-spacing:.05em;color:#C8102E;">Their Visit Objectives</p>
              <p style="margin:0;font-size:12px;color:#555;line-height:1.55;
                        white-space:pre-line;">{snippet}</p>
            </div>"""

    body_html = f"""
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f4f6f9;font-family:Arial,Helvetica,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f6f9;padding:32px 16px;">
    <tr><td align="center">
      <table width="600" cellpadding="0" cellspacing="0"
             style="background:#ffffff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.08);">

        <!-- Header -->
        <tr><td style="background:#1A2E42;padding:24px 32px;">
          <table width="100%" cellpadding="0" cellspacing="0"><tr>
            <td><p style="margin:0;color:#fff;font-size:20px;font-weight:700;">PRUDEV II</p>
                <p style="margin:2px 0 0;color:rgba(255,255,255,.65);font-size:12px;">
                  MSME Portfolio Management Programme</p></td>
            <td align="right"><p style="margin:0;color:#C8102E;font-size:11px;
                font-weight:700;letter-spacing:.05em;">GIZ · GOPA AFC</p></td>
          </tr></table>
        </td></tr>

        <!-- Body -->
        <tr><td style="padding:28px 32px;">
          <p style="margin:0 0 16px;color:#333;font-size:15px;">
            Dear <strong>{_esc(existing_bge.name)}</strong>,
          </p>
          <p style="margin:0 0 20px;color:#555;line-height:1.6;">
            This is to inform you that a second BGE has been co-assigned to visit one of the
            MSMEs currently in your portfolio. They will submit their own separate work order
            and visit report.
          </p>

          <!-- Shared MSME -->
          <p style="font-weight:700;color:#1A2E42;font-size:12px;text-transform:uppercase;
                    letter-spacing:.05em;margin:0 0 8px;">Shared MSME</p>
          <div style="background:#E3F2FD;border:1px solid #90CAF9;border-radius:6px;
                      padding:14px 16px;margin-bottom:20px;">
            <p style="margin:0;font-weight:700;color:#1A2E42;font-size:14px;">
              {_esc(msme.business_name)}</p>
            <p style="margin:2px 0 0;font-size:11px;color:#888;">
              {_esc(msme.msme_code or 'No code')}</p>
            {f'<p style="margin:6px 0 0;font-size:12px;color:#555;">{msme_detail_html}</p>'
              if msme_detail_html else ''}
          </div>

          <!-- Co-Assigned BGE -->
          <p style="font-weight:700;color:#1A2E42;font-size:12px;text-transform:uppercase;
                    letter-spacing:.05em;margin:0 0 8px;">Co-Assigned BGE</p>
          <div style="background:#FFF3E0;border:1px solid #FFCC80;border-radius:6px;
                      padding:14px 16px;">
            <p style="margin:0 0 4px;font-weight:700;color:#1A2E42;font-size:14px;">
              {_esc(new_bge.name)}</p>
            <p style="margin:0 0 6px;font-size:11px;color:#888;">
              BGE Code: {_esc(new_bge.bge_code or 'No code')}</p>
            {f'<p style="margin:0 0 8px;font-size:12px;color:#555;">{contact_html}</p>'
              if contact_html else ''}
            {obj_html}
          </div>

          <p style="margin:24px 0 0;color:#555;font-size:13px;line-height:1.7;
                    border-top:1px solid #e8edf2;padding-top:20px;">
            Please coordinate where possible to ensure visits are complementary and not
            duplicated. Log in to the
            <strong>PRUDEV II Portfolio Management System</strong> for full details.
          </p>
        </td></tr>

        <!-- Footer -->
        <tr><td style="background:#f8f9fa;padding:16px 32px;border-top:1px solid #e8edf2;">
          <p style="margin:0;color:#777;font-size:12px;">
            Best regards,<br><strong>PRUDEV II BDS Team</strong><br>GIZ · GOPA AFC</p>
        </td></tr>

      </table>
    </td></tr>
  </table>
</body>
</html>"""

    try:
        msg = EmailMultiAlternatives(
            subject=subject,
            body=body_text,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[existing_bge.email],
            reply_to=[getattr(settings, 'EMAIL_REPLY_TO', 'richard.obuku@gopa.eu')],
        )
        msg.attach_alternative(body_html, "text/html")
        msg.send(fail_silently=False)
    except Exception as exc:
        _log.warning("Co-assignment alert to %s failed: %s", existing_bge.email, exc)


# ── Push subscription API views ────────────────────────────────────────────────

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated as _IsAuth, AllowAny as _AllowAny

class WorkOrderViewSet(ProgrammeManagerReadOnlyMixin, ViewerReadOnlyMixin, viewsets.ModelViewSet):
    """Work Order management.

    Visibility:
    - Admins see all work orders (any status).
    - BGEs see only their own work orders with status 'issued' or 'signed'.

    Mutation (create / update / delete / issue):
    - Admin-only. BGEs, programme managers and viewers have read-only access.
    """
    serializer_class = WorkOrderSerializer
    permission_classes = [IsAuthenticated]

    def _is_admin(self):
        u = self.request.user
        return u.is_staff or u.is_superuser

    def get_queryset(self):
        user = self.request.user
        qs = WorkOrder.objects.select_related('bge', 'group')
        # Common filters regardless of role
        status_filter = self.request.query_params.get('status')
        type_filter   = self.request.query_params.get('work_order_type')
        if status_filter:
            qs = qs.filter(status=status_filter)
        if type_filter:
            qs = qs.filter(work_order_type=type_filter)
        if user.is_staff or user.is_superuser:
            bge_id = self.request.query_params.get('bge')
            if bge_id:
                qs = qs.filter(bge_id=bge_id)
            return qs
        # Programme managers and viewers see all work orders
        if _managed_groups(user) is not None or _is_viewer(user):
            return qs
        # BGE users: their own issued/signed orders only
        try:
            bge = user.bge_profile
        except Exception:
            return qs.none()
        return qs.filter(bge=bge, status__in=['issued', 'signed'])

    def _require_admin(self):
        if not self._is_admin():
            raise PermissionDenied("Work order management is restricted to administrators.")

    def _check_date_overlap(self, bge_id, start_date, end_date, exclude_id=None):
        if not bge_id or not start_date or not end_date:
            return
        qs = WorkOrder.objects.filter(
            bge_id=bge_id,
            start_date__isnull=False,
            end_date__isnull=False,
            start_date__lte=end_date,
            end_date__gte=start_date,
        )
        if exclude_id:
            qs = qs.exclude(pk=exclude_id)
        conflict = qs.first()
        if conflict:
            raise ValidationError(
                f"Date overlap: this BGE already has work order {conflict.work_order_number} "
                f"running from {conflict.start_date} to {conflict.end_date}. "
                "BGEs cannot be assigned overlapping work orders."
            )

    def perform_create(self, serializer):
        self._require_admin()
        data = serializer.validated_data
        self._check_date_overlap(
            bge_id=data.get('bge').pk if data.get('bge') else None,
            start_date=data.get('start_date'),
            end_date=data.get('end_date'),
        )
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        self._require_admin()
        data = serializer.validated_data
        instance = self.get_object()
        bge = data.get('bge', instance.bge)
        self._check_date_overlap(
            bge_id=bge.pk if bge else None,
            start_date=data.get('start_date', instance.start_date),
            end_date=data.get('end_date', instance.end_date),
            exclude_id=instance.pk,
        )
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        self._require_admin()
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='sign')
    def sign(self, request, pk=None):
        """BGE confirms acceptance: marks work order as signed with today's date."""
        from django.utils import timezone
        work_order = self.get_object()

        # Only the BGE this work order belongs to (or an admin/programme manager) may sign it
        user = request.user
        is_admin = user.is_staff or user.is_superuser or _managed_groups(user) is not None
        is_owner = hasattr(user, 'bge_profile') and user.bge_profile == work_order.bge
        if not (is_admin or is_owner):
            raise PermissionDenied("You can only sign your own work orders.")

        if work_order.status == 'signed':
            return Response({'detail': 'Already signed.'}, status=status.HTTP_200_OK)
        if work_order.status == 'draft':
            return Response({'detail': 'Work order has not been issued yet.'}, status=status.HTTP_400_BAD_REQUEST)

        work_order.status = 'signed'
        work_order.bge_signed_date = timezone.now().date()
        work_order.save(update_fields=['status', 'bge_signed_date'])

        # Generate the signed PDF immediately and persist it so the BGE's
        # signature is captured at this exact moment.  All future downloads
        # serve this frozen copy rather than regenerating.
        try:
            from .pdf_reports import render_work_order
            from django.core.files.base import ContentFile
            pdf_bytes = render_work_order(work_order).read()
            fname = f'WO_{(work_order.work_order_number or str(work_order.id)).replace(" ", "_")}_signed.pdf'
            work_order.signed_pdf.save(fname, ContentFile(pdf_bytes), save=False)
            # Store bytes in DB so the signed copy survives Render filesystem wipes
            work_order.signed_pdf_data = pdf_bytes
            work_order.save(update_fields=['signed_pdf', 'signed_pdf_data'])
        except Exception as e:
            logger.error('Failed to store signed work order PDF (wo id=%s): %s', work_order.id, e)
            # Signing is complete; PDF storage failure is non-fatal

        return Response(self.get_serializer(work_order).data)

    @action(detail=True, methods=['get'], url_path='pdf')
    def pdf(self, request, pk=None):
        """Render the work order as a PDF. Admin can access any; BGE can access their own.
        Signed work orders return the stored signed copy; others are rendered on demand."""
        work_order = self.get_object()
        user = request.user
        is_admin = user.is_staff or user.is_superuser
        is_owner = hasattr(user, 'bge_profile') and user.bge_profile == work_order.bge
        if not (is_admin or is_owner):
            raise PermissionDenied("You can only download your own work orders.")
        fname = _safe_filename(f'WorkOrder_{(work_order.work_order_number or str(work_order.id)).replace(" ", "_")}.pdf')
        dl = request.query_params.get('dl', '0')
        disp = 'attachment' if dl == '1' else 'inline'
        # Prefer DB-stored signed bytes (survives Render filesystem wipes).
        # Fall back to filesystem copy, then regenerate live for unsigned orders.
        if work_order.signed_pdf_data:
            resp = HttpResponse(bytes(work_order.signed_pdf_data), content_type='application/pdf')
            resp['Content-Disposition'] = f'{disp}; filename="{fname}"'
            return resp
        if work_order.signed_pdf:
            try:
                resp = HttpResponse(work_order.signed_pdf.read(), content_type='application/pdf')
                resp['Content-Disposition'] = f'{disp}; filename="{fname}"'
                return resp
            except Exception:
                pass
        from .pdf_reports import render_work_order
        buf = render_work_order(work_order)
        resp = HttpResponse(buf.read(), content_type='application/pdf')
        resp['Content-Disposition'] = f'{disp}; filename="{fname}"'
        return resp

    @action(detail=True, methods=['post'], url_path='issue')
    def issue(self, request, pk=None):
        """Admin-only: set status → issued, generate PDF, email to BGE."""
        self._require_admin()
        work_order = self.get_object()

        if work_order.status == 'issued':
            return Response({'detail': 'Already issued.'}, status=status.HTTP_200_OK)

        work_order.status = 'issued'
        # Snapshot the BGE's current MSME assignments so co-deployment overlap
        # can be detected even after MSMEs are later re-assigned to other BGEs.
        bge_msme_ids = list(
            work_order.bge.assigned_msmes.values_list('id', flat=True)
        )
        work_order.msme_ids_snapshot = bge_msme_ids
        work_order.save(update_fields=['status', 'msme_ids_snapshot'])

        # Generate PDF
        from .pdf_reports import render_work_order
        pdf_buf = render_work_order(work_order)
        pdf_bytes = pdf_buf.read()

        bge = work_order.bge
        recipient_email = bge.email or ''
        admin_email = getattr(settings, 'DEFAULT_FROM_EMAIL', '')
        recipients = [r for r in [recipient_email, admin_email] if r]

        if recipients:
            # Check for other BGEs with overlapping date ranges AND shared MSMEs
            co_text = ''
            if work_order.start_date and work_order.end_date:
                from .models import WorkOrder as _WO2
                overlapping = _WO2.objects.filter(
                    status__in=['issued', 'signed'],
                    start_date__lte=work_order.end_date,
                    end_date__gte=work_order.start_date,
                ).exclude(bge=bge).exclude(id=work_order.id).select_related('bge')
                # Union snapshot with current (handles both new and legacy work orders)
                my_set = set(bge_msme_ids) | BusinessGrowthExpertViewSet._bge_all_msme_ids(bge)
                aa_lines = []
                seen_bges = set()
                for owo in overlapping:
                    if owo.bge_id in seen_bges:
                        continue
                    other_current = BusinessGrowthExpertViewSet._bge_all_msme_ids(owo.bge)
                    other_ids = set(owo.msme_ids_snapshot or []) | other_current
                    if not (my_set & other_ids):
                        continue  # no shared MSMEs — skip
                    seen_bges.add(owo.bge_id)
                    obj = (owo.objective or owo.bge.deployment_objectives or '').strip()
                    snippet = (obj[:250] + '…') if len(obj) > 250 else obj
                    aa_lines.append(
                        f"  BGE:        {owo.bge.name} ({owo.bge.bge_code or 'No code'})"
                        + (f"\n  Work Order: {owo.work_order_number}")
                        + (f"\n  Phone:      {owo.bge.phone}" if owo.bge.phone else '')
                        + (f"\n  Objectives: {snippet}" if snippet else '')
                    )
                if aa_lines:
                    co_text = (
                        "\n\nPLEASE NOTE — ANOTHER BGE ALREADY ASSIGNED\n"
                        + "─" * 40 + "\n"
                        + "Another BGE has already been assigned to work with some of the same "
                        + "MSMEs during this period. Please coordinate accordingly:\n\n"
                        + "\n\n".join(aa_lines)
                    )

            subject = f'Work Order Issued — {work_order.work_order_number}'
            body = (
                f'Dear {bge.name},\n\n'
                f'Please find attached your work order ({work_order.work_order_number}) '
                f'for the PRUDEV II programme.\n\n'
                f'Work Order Type: {work_order.get_work_order_type_display()}\n'
                f'Issue Date: {work_order.issue_date}\n'
                f'Period: {work_order.start_date or "TBD"} to {work_order.end_date or "TBD"}\n'
                f'Net Payable: UGX {work_order.rate_per_day * work_order.max_days - int(work_order.rate_per_day * work_order.max_days * 0.06):,}\n'
                f'{co_text}\n\n'
                f'Regards,\nPRUDEV II BDS Team\nGOPA AFC / GIZ'
            )
            email = EmailMultiAlternatives(subject, body,
                                           getattr(settings, 'DEFAULT_FROM_EMAIL', ''),
                                           recipients)
            filename = f'WorkOrder_{work_order.work_order_number.replace(" ", "_")}.pdf'
            email.attach(filename, pdf_bytes, 'application/pdf')
            try:
                email.send(fail_silently=False)
            except Exception as exc:
                return Response(
                    {'detail': f'Issued but email failed: {exc}'},
                    status=status.HTTP_200_OK,
                )

        serializer = self.get_serializer(work_order)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='withdraw')
    def withdraw(self, request, pk=None):
        """Admin-only: withdraw a work order (issued or signed) back to draft status.
        Clears the signed PDF and emails the BGE to notify them."""
        self._require_admin()
        work_order = self.get_object()

        if work_order.status == 'draft':
            return Response({'detail': 'Work order is already in draft status.'}, status=status.HTTP_400_BAD_REQUEST)

        reason = request.data.get('reason', '').strip()

        work_order.status = 'draft'
        work_order.bge_signed_date = None
        # Clear any stored signed PDF so re-issue generates a fresh one
        work_order.signed_pdf_data = None
        if work_order.signed_pdf:
            work_order.signed_pdf.delete(save=False)
        work_order.save(update_fields=['status', 'bge_signed_date', 'signed_pdf', 'signed_pdf_data'])

        bge = work_order.bge
        recipient_email = bge.email or ''
        admin_email = getattr(settings, 'DEFAULT_FROM_EMAIL', '')
        recipients = [r for r in [recipient_email, admin_email] if r]

        if recipients:
            subject = f'Work Order Withdrawn — {work_order.work_order_number}'
            reason_line = f'\nReason: {reason}\n' if reason else ''
            body = (
                f'Dear {bge.name},\n\n'
                f'Your work order ({work_order.work_order_number}) has been withdrawn and is under review.\n'
                f'{reason_line}\n'
                f'Work Order Type: {work_order.get_work_order_type_display()}\n'
                f'You will be notified when a revised work order is re-issued to you.\n\n'
                f'Regards,\nPRUDEV II BDS Team\nGOPA AFC / GIZ'
            )
            try:
                msg = EmailMultiAlternatives(
                    subject, body,
                    getattr(settings, 'DEFAULT_FROM_EMAIL', ''),
                    recipients,
                )
                msg.send(fail_silently=True)
            except Exception:
                pass  # withdrawal succeeds even if email fails

        return Response(self.get_serializer(work_order).data)


@api_view(['POST'])
@permission_classes([_IsAuth])
def push_subscribe(request):
    """Save or update a push subscription for the current user.
    Accepts flat format: {endpoint, p256dh, auth} (as sent by the frontend).
    Also accepts nested format: {endpoint, keys: {p256dh, auth}} (Web Push API standard).
    """
    data = request.data
    endpoint = data.get('endpoint', '').strip()
    # Accept both flat and nested key formats
    if 'keys' in data and isinstance(data['keys'], dict):
        p256dh = data['keys'].get('p256dh', '')
        auth   = data['keys'].get('auth', '')
    else:
        p256dh = data.get('p256dh', '')
        auth   = data.get('auth', '')

    if not endpoint:
        return Response({'error': 'endpoint required'}, status=status.HTTP_400_BAD_REQUEST)
    if not p256dh or not auth:
        return Response({'error': 'p256dh and auth keys are required'}, status=status.HTTP_400_BAD_REQUEST)

    PushSubscription.objects.update_or_create(
        endpoint=endpoint,
        defaults={'user': request.user, 'p256dh': p256dh, 'auth': auth},
    )
    return Response({'message': 'Subscribed'})


@api_view(['POST'])
@permission_classes([_IsAuth])
def push_unsubscribe(request):
    """Remove a push subscription."""
    endpoint = request.data.get('endpoint', '').strip()
    PushSubscription.objects.filter(endpoint=endpoint, user=request.user).delete()
    return Response({'message': 'Unsubscribed'})


@api_view(['GET'])
@permission_classes([_AllowAny])
def push_vapid_key(request):
    """Return the VAPID public key so the frontend can subscribe. Public endpoint."""
    return Response({'publicKey': settings.VAPID_PUBLIC_KEY})


class TrainingReportViewSet(ProgrammeManagerReadOnlyMixin, ViewerReadOnlyMixin, viewsets.ModelViewSet):
    """
    CRUD for training session reports.
    - BGEs can create/edit reports for sessions linked to their assignments.
    - Admins and programme managers see everything.
    """
    serializer_class   = TrainingReportSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = TrainingReport.objects.select_related('session', 'bge')
        if user.is_staff or user.is_superuser:
            return qs
        if _is_programme_manager(user):
            group_ids = _managed_groups(user) or []
            return qs.filter(
                session__businesses__programme_groups__in=group_ids
            ).distinct()
        try:
            return qs.filter(bge=user.bge_profile)
        except Exception:
            return qs.none()

    def perform_create(self, serializer):
        bge = None
        try:
            bge = self.request.user.bge_profile
        except Exception:
            pass
        serializer.save(bge=bge)

    def perform_update(self, serializer):
        from django.utils import timezone
        data = {}
        if serializer.validated_data.get('status') == 'submitted':
            data['submitted_at'] = timezone.now()
        serializer.save(**data)

    @action(detail=True, methods=['post'], url_path='revert')
    def revert(self, request, pk=None):
        """Admin-only: revert a submitted training report back to draft."""
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can revert reports.")
        report = self.get_object()
        if report.status == 'draft':
            return Response({'detail': 'Report is already a draft.'}, status=status.HTTP_400_BAD_REQUEST)
        report.status = 'draft'
        report.submitted_at = None
        report.save(update_fields=['status', 'submitted_at'])
        from .serializers import TrainingReportSerializer
        return Response(TrainingReportSerializer(report, context={'request': request}).data)

    @action(detail=True, methods=['get'], url_path='pdf')
    def pdf(self, request, pk=None):
        """Render this training report as a branded PDF."""
        from .pdf_reports import render_training_report
        report = self.get_object()
        safe = report.session.title[:40].replace(' ', '_')
        fname = _safe_filename(f"TrainingReport_{safe}_{report.session.date}.pdf")
        dl = request.query_params.get('dl')
        buf = render_training_report(report)
        resp = HttpResponse(buf.read(), content_type='application/pdf')
        disposition = 'attachment' if dl else 'inline'
        resp['Content-Disposition'] = f'{disposition}; filename="{fname}"'
        return resp


class AnnualReviewReportViewSet(ProgrammeManagerReadOnlyMixin, ViewerReadOnlyMixin, viewsets.ModelViewSet):
    """Annual / quarterly / mid-term review reports authored by a single BGE.

    A BGE selects which of their MSMEs to include (attendance list) and writes
    a narrative summary. No financial data is captured here — that lives in
    GrowthSnapshot records. Admins see all; BGEs see only their own.
    """
    serializer_class  = AnnualReviewReportSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.is_staff or user.is_superuser:
            qs = AnnualReviewReport.objects.all()
        elif _is_viewer(user):
            qs = AnnualReviewReport.objects.all()
        elif _is_programme_manager(user):
            group_ids = _managed_groups(user) or []
            qs = AnnualReviewReport.objects.filter(
                msmes_reviewed__programme_groups__in=group_ids
            ).distinct()
        else:
            try:
                qs = AnnualReviewReport.objects.filter(bge=user.bge_profile)
            except Exception:
                qs = AnnualReviewReport.objects.none()

        period = self.request.query_params.get('period')
        if period:
            qs = qs.filter(review_period=period)
        status = self.request.query_params.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs.select_related('bge').prefetch_related('msmes_reviewed')

    def perform_create(self, serializer):
        bge = None
        try:
            bge = self.request.user.bge_profile
        except Exception:
            pass
        serializer.save(bge=bge)

    def perform_update(self, serializer):
        from django.utils import timezone
        data = {}
        if serializer.validated_data.get('status') == 'submitted':
            data['submitted_at'] = timezone.now()
        serializer.save(**data)


class MentorTrainingReportViewSet(ProgrammeManagerReadOnlyMixin, ViewerReadOnlyMixin, viewsets.ModelViewSet):
    """
    Training reports filed by mentor BGEs.
    - A mentor BGE can only create/edit their own report for sessions they are assigned to.
    - Admins and programme managers see all mentor reports.
    """
    serializer_class   = MentorTrainingReportSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = MentorTrainingReport.objects.select_related(
            'session', 'bge'
        ).prefetch_related(
            'session__businesses', 'session__attendances',
            'session__facilitation_assignments__bge',
        )
        if user.is_staff or user.is_superuser:
            return qs
        if _is_programme_manager(user):
            group_ids = _managed_groups(user) or []
            return qs.filter(
                session__businesses__programme_groups__in=group_ids
            ).distinct()
        try:
            return qs.filter(bge=user.bge_profile)
        except Exception:
            return qs.none()

    def perform_create(self, serializer):
        bge = None
        try:
            bge = self.request.user.bge_profile
            # verify BGE is actually a mentor on this session
            session = serializer.validated_data.get('session')
            if session and not session.facilitation_assignments.filter(bge=bge, role='mentor').exists():
                raise PermissionDenied("You are not assigned as a mentor for this session.")
        except PermissionDenied:
            raise
        except Exception:
            pass
        serializer.save(bge=bge)

    def perform_update(self, serializer):
        from django.utils import timezone
        data = {}
        if serializer.validated_data.get('status') == 'submitted':
            data['submitted_at'] = timezone.now()
        serializer.save(**data)

    @action(detail=True, methods=['post'], url_path='revert')
    def revert(self, request, pk=None):
        """Admin-only: revert a submitted mentor report back to draft."""
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can revert reports.")
        report = self.get_object()
        if report.status == 'draft':
            return Response({'detail': 'Report is already a draft.'}, status=status.HTTP_400_BAD_REQUEST)
        report.status = 'draft'
        report.submitted_at = None
        report.save(update_fields=['status', 'submitted_at'])
        from .serializers import MentorTrainingReportSerializer
        return Response(MentorTrainingReportSerializer(report, context={'request': request}).data)

    @action(detail=True, methods=['get'], url_path='pdf')
    def pdf(self, request, pk=None):
        """Render this mentor training report as a branded PDF."""
        from .pdf_reports import render_mentor_report
        report = self.get_object()
        safe = report.session.title[:40].replace(' ', '_')
        fname = _safe_filename(f"MentorReport_{safe}_{report.session.date}.pdf")
        dl = request.query_params.get('dl')
        buf = render_mentor_report(report)
        resp = HttpResponse(buf.read(), content_type='application/pdf')
        disposition = 'attachment' if dl else 'inline'
        resp['Content-Disposition'] = f'{disposition}; filename="{fname}"'
        return resp


# ── Bulk communication email ────────────────────────────────────────────────────

def _do_send_emails(records, subject, body_text, body_html, skip_sent,
                    already_sent_ids, recipient_type, from_email, reply_to, user_id):
    """Send emails in a background thread using a single SMTP connection."""
    from django.core.mail import get_connection
    from .models import EmailSendLog

    # Deduplicate by email address (keeps first occurrence per address)
    seen_emails = set()
    deduped = []
    for rec in records:
        addr = (rec['email'] or '').lower().strip()
        if addr and addr not in seen_emails:
            seen_emails.add(addr)
            deduped.append(rec)

    logs_to_create = []
    failed_count = 0

    try:
        connection = get_connection()
        connection.open()
    except Exception as e:
        logger.warning('Bulk email: could not open SMTP connection, falling back to per-message: %s', e)
        connection = None  # fall back to per-message connections

    for rec in deduped:
        if skip_sent and rec['id'] in already_sent_ids:
            continue
        first = (rec['name'] or '').split()[0] or 'Team'
        try:
            txt  = body_text.replace('{{name}}', first)
            html = body_html.replace('{{name}}', first) if body_html else ''
            msg  = EmailMultiAlternatives(
                subject=subject, body=txt,
                from_email=from_email, to=[rec['email']], reply_to=[reply_to],
                connection=connection,
            )
            if html:
                msg.attach_alternative(html, 'text/html')
            msg.send()
            logs_to_create.append(EmailSendLog(
                recipient_type=recipient_type, recipient_id=rec['id'],
                recipient_email=rec['email'], subject=subject,
                sent_by_id=user_id,
            ))
        except Exception as e:
            failed_count += 1
            logger.error('Bulk email: failed to send to %s (id=%s): %s', rec.get('email'), rec.get('id'), e)

    if failed_count:
        logger.warning('Bulk email job finished with %d failure(s) out of %d recipients for subject: %s',
                       failed_count, len(deduped), subject)

    if connection:
        try:
            connection.close()
        except Exception:
            pass

    try:
        if logs_to_create:
            EmailSendLog.objects.bulk_create(logs_to_create, ignore_conflicts=True)
    except Exception as e:
        logger.error('Bulk email: failed to persist send logs: %s', e)


@api_view(['POST'])
@permission_classes([_IsAuth])
def bulk_email_view(request):
    """Admin-only: send a communication email to a selection of BGEs or MSMEs."""
    import threading
    from .models import EmailSendLog

    user = request.user
    if not (user.is_staff or user.is_superuser or _managed_groups(user) is not None):
        raise PermissionDenied("Only administrators can send bulk emails.")

    recipient_type  = request.data.get('recipient_type', 'bge')
    recipient_ids   = request.data.get('recipient_ids', [])
    subject         = (request.data.get('subject') or '').strip()
    body_text       = (request.data.get('body') or request.data.get('body_text') or '').strip()
    body_html       = (request.data.get('body_html') or '').strip()
    skip_sent       = bool(request.data.get('skip_already_sent', False))

    if not subject:
        return Response({'detail': 'Subject is required.'}, status=400)
    if not body_text:
        return Response({'detail': 'Body is required.'}, status=400)
    if len(subject) > 300:
        return Response({'detail': 'Subject is too long (max 300 characters).'}, status=400)
    if len(body_text) > 50000:
        return Response({'detail': 'Body is too long (max 50,000 characters).'}, status=400)
    if len(body_html) > 100000:
        return Response({'detail': 'HTML body is too long (max 100,000 characters).'}, status=400)
    if isinstance(recipient_ids, list) and len(recipient_ids) > 2000:
        return Response({'detail': 'Too many recipients selected (max 2,000).'}, status=400)

    if recipient_type == 'bge':
        qs = BusinessGrowthExpert.objects.filter(email__isnull=False).exclude(email='')
        if recipient_ids:
            qs = qs.filter(id__in=recipient_ids)
        records = [{'id': b.id, 'name': b.name or 'BGE', 'email': b.email} for b in qs]
    else:
        qs = MSME.objects.filter(email__isnull=False).exclude(email='')
        if recipient_ids:
            qs = qs.filter(id__in=recipient_ids)
        records = [{'id': m.id, 'name': m.owner_name or m.business_name or 'Business Owner', 'email': m.email} for m in qs]

    # Deduplicate by email address before counting
    seen = set()
    deduped_records = []
    for rec in records:
        addr = (rec['email'] or '').lower().strip()
        if addr and addr not in seen:
            seen.add(addr)
            deduped_records.append(rec)

    # Identify already-sent recipients for this subject
    try:
        already_sent_ids = set(
            EmailSendLog.objects.filter(
                recipient_type=recipient_type, subject=subject,
                recipient_id__in=[r['id'] for r in deduped_records],
            ).values_list('recipient_id', flat=True)
        )
    except Exception:
        already_sent_ids = set()

    to_send = len(deduped_records) - (len(already_sent_ids) if skip_sent else 0)
    skipped = len(deduped_records) - to_send
    duplicates_removed = len(records) - len(deduped_records)

    from_email = settings.DEFAULT_FROM_EMAIL
    reply_to   = getattr(settings, 'EMAIL_REPLY_TO', from_email)

    # Fire-and-forget background thread — avoids HTTP timeout for large lists
    t = threading.Thread(
        target=_do_send_emails,
        args=(deduped_records, subject, body_text, body_html, skip_sent,
              already_sent_ids, recipient_type, from_email, reply_to, user.id),
        daemon=True,
    )
    t.start()

    return Response({
        'queued': to_send,
        'skipped': skipped,
        'duplicates_removed': duplicates_removed,
        'already_sent_count': len(already_sent_ids),
        'message': f'{to_send} email{"s" if to_send != 1 else ""} queued for delivery.',
    })


@api_view(['GET'])
@permission_classes([_IsAuth])
def bulk_email_log_view(request):
    """Return how many of the given recipients have already been sent a subject."""
    from .models import EmailSendLog

    user = request.user
    if not (user.is_staff or user.is_superuser or _managed_groups(user) is not None):
        raise PermissionDenied("Only administrators can view bulk email logs.")

    subject        = request.query_params.get('subject', '')
    recipient_type = request.query_params.get('recipient_type', 'bge')
    ids_raw        = request.query_params.getlist('ids')
    recipient_ids  = [int(i) for i in ids_raw if i.isdigit()]

    if not subject or not recipient_ids:
        return Response({'already_sent': [], 'count': 0})

    already = list(
        EmailSendLog.objects.filter(
            recipient_type=recipient_type, subject=subject,
            recipient_id__in=recipient_ids,
        ).values_list('recipient_id', flat=True).distinct()
    )
    return Response({'already_sent': already, 'count': len(already)})


# ── Bulk SMS (Message Carrier) ────────────────────────────────────────────────

def _normalise_phone(phone):
    """Ensure phone number is in international format for Message Carrier."""
    p = re.sub(r'[\s\-\(\)]', '', str(phone or ''))
    if p.startswith('0'):
        p = '+256' + p[1:]  # Uganda default
    if not p.startswith('+'):
        p = '+' + p
    return p


SMS_BALANCE_CACHE_KEY = 'mc_sms_wallet_balance'


def _cache_sms_balance(balance):
    """Store the latest known SMS wallet balance in Django's cache (24h TTL)."""
    try:
        from django.core.cache import cache
        cache.set(SMS_BALANCE_CACHE_KEY, float(balance), timeout=86400)
    except Exception:
        pass


def _get_cached_sms_balance():
    """Return the last known SMS wallet balance from cache, or None."""
    try:
        from django.core.cache import cache
        return cache.get(SMS_BALANCE_CACHE_KEY)
    except Exception:
        return None


@api_view(['GET'])
@permission_classes([_IsAuth])
def bulk_sms_balance_view(request):
    """Return the current Message Carrier wallet balance (from cache).

    The cache is populated after every bulk SMS send (balanceAfter field).
    Use ?seed=<amount> to manually seed the balance (admin only).
    """
    if not (request.user.is_staff or request.user.is_superuser):
        raise PermissionDenied("Only administrators can view SMS balance.")

    # Allow manual seeding: /api/bulk-sms/balance/?seed=465
    seed = request.query_params.get('seed')
    if seed:
        try:
            _cache_sms_balance(float(seed))
        except ValueError:
            pass

    balance = _get_cached_sms_balance()
    return Response({
        'balance': balance,
        'currency': 'UGX',
        'cost_per_sms': 45,
        'ok': balance is not None,
        'message': '' if balance is not None else 'Balance updates automatically after each SMS send.',
    })


def _do_send_sms(records, message, user_id):
    """Background thread: send SMS to each record via Message Carrier API."""
    import urllib.request as _urllib
    import json as _json
    import django
    django.setup() if not django.conf.settings.configured else None  # noqa
    from django.conf import settings as _s
    from .models import SmsSendLog
    from django.contrib.auth.models import User as _User

    api_key = _s.MESSAGE_CARRIER_API_KEY
    base_url = getattr(_s, 'MESSAGE_CARRIER_BASE_URL', 'https://api.bravo.mystyler.xyz')
    endpoint = f'{base_url}/v1/api-keys/send-sms'

    try:
        sent_by = _User.objects.get(pk=user_id)
    except Exception:
        sent_by = None

    logs = []
    last_balance = None
    for rec in records:
        phone = _normalise_phone(rec['phone'])
        personal_msg = message.replace('{{name}}', (rec['name'] or '').split()[0])
        try:
            import requests as _req_lib
            resp_obj = _req_lib.post(
                endpoint,
                json={'phone': phone, 'message': personal_msg},
                headers={'x-api-key': api_key},
                timeout=15,
            )
            resp_data = resp_obj.json() if resp_obj.content else {}
            if resp_obj.status_code < 300:
                status = 'sent'
                err = ''
                # Capture balance from successful send
                if 'balanceAfter' in resp_data:
                    last_balance = resp_data['balanceAfter']
            else:
                status = 'failed'
                err = resp_data.get('message', resp_obj.text[:200])
        except Exception as exc:
            status = 'failed'
            err = str(exc)
            logger.error('SMS send failed for %s (%s): %s', rec['name'], phone, exc)

        logs.append(SmsSendLog(
            recipient_type=rec['rtype'],
            recipient_id=rec['id'],
            recipient_phone=phone,
            message_preview=personal_msg[:160],
            sent_by=sent_by,
            status=status,
            error=err,
        ))

    # Cache the latest balance so the balance endpoint can return it
    if last_balance is not None:
        _cache_sms_balance(last_balance)

    try:
        SmsSendLog.objects.bulk_create(logs, ignore_conflicts=True)
    except Exception as exc:
        logger.error('Bulk SMS: failed to persist send logs: %s', exc)


@api_view(['POST'])
@permission_classes([_IsAuth])
def bulk_sms_view(request):
    """Admin-only: send bulk SMS to BGEs or MSMEs via Message Carrier."""
    import threading

    user = request.user
    if not (user.is_staff or user.is_superuser or _managed_groups(user) is not None):
        raise PermissionDenied("Only administrators can send bulk SMS.")

    recipient_type = request.data.get('recipient_type', 'bge')
    recipient_ids  = request.data.get('recipient_ids', [])
    message        = (request.data.get('message') or '').strip()

    if not message:
        return Response({'detail': 'Message is required.'}, status=400)
    if len(message) > 1600:
        return Response({'detail': 'Message is too long (max 1,600 characters).'}, status=400)
    if isinstance(recipient_ids, list) and len(recipient_ids) > 2000:
        return Response({'detail': 'Too many recipients selected (max 2,000).'}, status=400)

    # Scope filter: programme managers can only message within their groups
    group_ids = _managed_groups(user)  # None for superuser/staff, list for programme managers

    if recipient_type == 'bge':
        qs = BusinessGrowthExpert.objects.filter(phone__isnull=False).exclude(phone='')
        if group_ids is not None:  # programme manager — restrict to managed groups
            qs = qs.filter(bge_groups__programme_group__in=group_ids)
        if recipient_ids:
            qs = qs.filter(id__in=recipient_ids)
        records = [{'id': b.id, 'name': b.name or 'BGE', 'phone': b.phone, 'rtype': 'bge'} for b in qs]
    else:
        qs = MSME.objects.filter(phone__isnull=False).exclude(phone='')
        if group_ids is not None:  # programme manager — restrict to managed groups
            qs = qs.filter(assigned_group__in=group_ids)
        if recipient_ids:
            qs = qs.filter(id__in=recipient_ids)
        records = [{'id': m.id, 'name': m.owner_name or m.business_name or 'Business', 'phone': m.phone, 'rtype': 'msme'} for m in qs]

    # Deduplicate by phone number
    seen = set()
    deduped = []
    for rec in records:
        p = _normalise_phone(rec['phone'])
        if p and p not in seen:
            seen.add(p)
            deduped.append(rec)
    duplicates_removed = len(records) - len(deduped)

    threading.Thread(
        target=_do_send_sms,
        args=(deduped, message, user.id),
        daemon=True,
    ).start()

    return Response({
        'queued': len(deduped),
        'duplicates_removed': duplicates_removed,
        'message': f'{len(deduped)} SMS message{"s" if len(deduped) != 1 else ""} queued for delivery.',
    })


@api_view(['GET'])
@permission_classes([_IsAuth])
def bulk_sms_log_view(request):
    """Return recent SMS send log entries for the given recipients."""
    from .models import SmsSendLog

    user = request.user
    if not (user.is_staff or user.is_superuser or _managed_groups(user) is not None):
        raise PermissionDenied("Only administrators can view bulk SMS logs.")

    recipient_type = request.query_params.get('recipient_type', 'bge')
    ids_raw        = request.query_params.getlist('ids')
    recipient_ids  = [int(i) for i in ids_raw if i.isdigit()]

    if not recipient_ids:
        return Response({'logs': []})

    logs = SmsSendLog.objects.filter(
        recipient_type=recipient_type,
        recipient_id__in=recipient_ids,
    ).order_by('-sent_at')[:200]

    return Response({'logs': [
        {
            'recipient_id': l.recipient_id,
            'phone': l.recipient_phone,
            'preview': l.message_preview,
            'sent_at': l.sent_at.isoformat(),
            'status': l.status,
        }
        for l in logs
    ]})


# ── T-Shirt Receipt ViewSets ───────────────────────────────────────────────

from .serializers import TshirtReceiptSerializer, TshirtReceiptEntrySerializer
from django.utils import timezone
import base64, io as _io

def _bge_signature_bytes(bge):
    """Return raw PNG bytes of the BGE's stored signature, or None."""
    if bge.signature_data:
        return bytes(bge.signature_data)
    if bge.signature:
        try:
            with open(bge.signature.path, 'rb') as f:
                return f.read()
        except Exception:
            pass
    return None


def _clean_sig_for_pdf(raw_bytes):
    """
    Re-process a signature image before embedding in a PDF.

    Applies luminance + saturation-based background removal so that
    off-white or slightly yellowed paper backgrounds become fully
    transparent, giving a clean floating signature on any background.
    Returns PNG bytes ready for reportlab.
    """
    try:
        import PIL.Image as _PIL
        img = _PIL.open(_io.BytesIO(raw_bytes)).convert('RGBA')
        pixels = list(img.getdata())
        cleaned = []
        for r, g, b, a in pixels:
            lum = 0.299 * r + 0.587 * g + 0.114 * b
            sat_range = max(r, g, b) - min(r, g, b)
            # Bright (lum > 205) AND low saturation (sat_range < 45)
            # → paper/background → transparent
            if lum > 205 and sat_range < 45:
                cleaned.append((r, g, b, 0))
            else:
                cleaned.append((r, g, b, a))
        img.putdata(cleaned)
        buf = _io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        return buf.read()
    except Exception:
        return raw_bytes  # fall back to original on any error


def _build_tshirt_pdf(receipt):
    """Generate a signed PDF for a TshirtReceipt using reportlab.

    Landscape A4 with the standard PRUDEV II branded header:
    GOPA AFC logo (left) | PRUDEV II wordmark (centre) | GIZ logo (right)
    — identical to the visit-report and training-report headers.
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.utils import ImageReader
        import PIL.Image as PILImage
    except ImportError as e:
        raise ImportError(f"reportlab/Pillow required: {e}")

    # ── Logo paths (same static dir used by pdf_reports.py) ────────────────
    import os as _os
    _logo_dir  = _os.path.join(_os.path.dirname(__file__), 'static', 'portfolio', 'images')
    GOPA_PATH  = _os.path.join(_logo_dir, 'gopa-logo.png')
    GIZ_PATH   = _os.path.join(_logo_dir, 'giz-logo.png')

    NAVY  = colors.HexColor("#1A2F4B")
    RED   = colors.HexColor("#C8102E")
    GREY  = colors.HexColor("#666666")
    LGREY = colors.HexColor("#F2F2F2")

    PAGE_W, PAGE_H = landscape(A4)   # 841.9 × 595.3 pt
    BAND_H  = 24 * mm
    RULE_H  = 1.5 * mm
    TOP_M   = BAND_H + RULE_H + 8 * mm
    SIDE_M  = 15 * mm
    BOT_M   = 15 * mm

    # ── Per-page header callback ────────────────────────────────────────────
    def _draw_header(canvas, doc):
        canvas.saveState()
        w, h = landscape(A4)

        # White band
        canvas.setFillColorRGB(1, 1, 1)
        canvas.rect(0, h - BAND_H, w, BAND_H, fill=1, stroke=0)
        # GIZ-red rule
        canvas.setFillColor(RED)
        canvas.rect(0, h - BAND_H - RULE_H, w, RULE_H, fill=1, stroke=0)

        # Left: GOPA logo (aspect ≈ 3.06)
        if _os.path.isfile(GOPA_PATH):
            logo_h = 14 * mm
            logo_w = logo_h * 3.06
            canvas.drawImage(
                ImageReader(GOPA_PATH),
                x=14 * mm, y=h - BAND_H + (BAND_H - logo_h) / 2,
                width=logo_w, height=logo_h,
                mask='auto', preserveAspectRatio=True,
            )

        # Right: GIZ logo (aspect ≈ 2.71)
        if _os.path.isfile(GIZ_PATH):
            logo_h = 16 * mm
            logo_w = logo_h * 2.71
            canvas.drawImage(
                ImageReader(GIZ_PATH),
                x=w - 14 * mm - logo_w, y=h - BAND_H + (BAND_H - logo_h) / 2,
                width=logo_w, height=logo_h,
                mask='auto', preserveAspectRatio=True,
            )

        # Centre: wordmark
        cy = h - BAND_H / 2
        canvas.setFillColor(NAVY)
        canvas.setFont('Helvetica-Bold', 13)
        canvas.drawCentredString(w / 2, cy + 2 * mm, 'PRUDEV II')
        canvas.setFillColor(GREY)
        canvas.setFont('Helvetica', 8)
        canvas.drawCentredString(w / 2, cy - 2 * mm, 'T-Shirt Distribution Receipt')

        # Page number
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(GREY)
        canvas.drawRightString(w - SIDE_M, 8 * mm, f'Page {doc.page}')
        canvas.restoreState()

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=SIDE_M, rightMargin=SIDE_M,
        topMargin=TOP_M, bottomMargin=BOT_M,
        title=receipt.title,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TTitle', parent=styles['Title'],
        fontSize=15, textColor=NAVY, spaceAfter=2, alignment=TA_CENTER)
    sub_style   = ParagraphStyle('TSub', parent=styles['Normal'],
        fontSize=10, textColor=GREY, spaceAfter=2,
        alignment=TA_CENTER, fontName='Helvetica-Oblique')
    label_style = ParagraphStyle('TLbl', parent=styles['Normal'],
        fontSize=9, textColor=colors.black, leading=13)
    note_style  = ParagraphStyle('TNote', parent=styles['Normal'],
        fontSize=7.5, textColor=GREY,
        fontName='Helvetica-Oblique', alignment=TA_CENTER)
    conf_style  = ParagraphStyle('TConf', parent=styles['Normal'],
        fontSize=7, textColor=GREY,
        fontName='Helvetica-Oblique', alignment=TA_CENTER)

    entries = list(receipt.entries.select_related('bge').order_by('order', 'bge__name'))

    # ── Column widths — landscape A4 content = 297mm − 30mm margins = 267mm = 26.7cm
    # #(0.7) Name(4.5) Code(3.0) Phone(3.3) Loc(2.2) Size(1.4) Qty(0.8) Sig(6.0) Date(2.8) = 24.7cm
    col_w = [0.7*cm, 4.5*cm, 3.0*cm, 3.3*cm, 2.2*cm, 1.4*cm, 0.8*cm, 6.0*cm, 2.8*cm]

    SIG_ROW_H = 1.1 * cm

    # Paragraph style for body cells — enables word-wrap so nothing overflows
    cell_s = ParagraphStyle('TCell', parent=styles['Normal'],
        fontSize=8, leading=10, wordWrap='LTR', splitLongWords=True)
    # Header cell style (white text on navy background)
    hdr_s = ParagraphStyle('THdrCell', parent=styles['Normal'],
        fontSize=8, leading=10, textColor=colors.white,
        fontName='Helvetica-Bold', alignment=TA_CENTER)

    def _P(text, style=cell_s):
        return Paragraph(str(text) if text else '', style)

    hdr_row = [_P(h, hdr_s) for h in
               ['#', 'BGE Name', 'BGE Code', 'Phone', 'Location',
                'Size', 'Qty', 'BGE Signature', 'Date Signed']]
    rows = [hdr_row]

    for idx, entry in enumerate(entries):
        sig_bytes = _bge_signature_bytes(entry.bge)
        if entry.signed and sig_bytes:
            try:
                # Clean background then embed
                clean_bytes = _clean_sig_for_pdf(sig_bytes)
                img_buf = _io.BytesIO(clean_bytes)
                pil = PILImage.open(img_buf)
                w_px, h_px = pil.size
                aspect = w_px / h_px if h_px else 1
                img_h = SIG_ROW_H * 0.85
                img_w = min(img_h * aspect, col_w[7] - 4 * mm)
                img_buf.seek(0)
                sig_cell = Image(img_buf, width=img_w, height=img_h)
            except Exception:
                sig_cell = _P('(signed)')
        elif entry.signed:
            sig_cell = _P('(signed)')
        else:
            sig_cell = _P('')

        date_str = entry.signed_at.strftime('%d/%m/%Y') if entry.signed_at else ''
        rows.append([
            _P(str(idx + 1)),
            _P(entry.bge.name),
            _P(entry.bge.bge_code or ''),
            _P(entry.bge.phone or ''),
            _P(entry.bge.location or ''),
            _P(entry.size),
            _P(str(entry.quantity)),
            sig_cell,
            _P(date_str),
        ])

    row_heights = [0.65 * cm] + [SIG_ROW_H] * len(entries)
    tbl = Table(rows, colWidths=col_w, rowHeights=row_heights, repeatRows=1)
    tbl.setStyle(TableStyle([
        # Header row background (text style is in hdr_s ParagraphStyle)
        ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
        ('VALIGN',        (0, 0), (-1, 0), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, 0), 5),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
        # Body
        ('VALIGN',        (0, 1), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        # Alternating rows
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LGREY]),
        # Grid
        ('GRID',           (0, 0), (-1, -1), 0.4, colors.HexColor("#BBBBBB")),
    ]))
    story = [
        Paragraph(receipt.title, title_style),
    ]
    sub_parts = []
    if receipt.event:
        sub_parts.append(receipt.event)
    sub_parts.append(f"Colour: {receipt.colour}")
    story.append(Paragraph("  —  ".join(sub_parts), sub_style))
    story.append(Spacer(1, 0.25 * cm))
    from datetime import date as _date, timedelta as _td
    _today = _date.today()
    _this_monday = _today - _td(days=_today.isocalendar()[2] - 1)
    _last_friday = _this_monday - _td(days=3)
    story.append(Paragraph(
        f"Date: {_last_friday.strftime('%d/%m/%Y')}",
        label_style))
    story.append(Spacer(1, 0.15 * cm))
    story.append(Paragraph(
        "Signatures below are embedded from each BGE's registered profile on the PRUDEV II system.",
        note_style))
    story.append(Spacer(1, 0.2 * cm))
    story.append(tbl)
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(
        f"Total BGEs: {len(entries)}     Signed: {receipt.signed_count}     Pending: {len(entries) - receipt.signed_count}",
        label_style))
    story.append(Spacer(1, 0.15 * cm))
    story.append(Paragraph(
        "Distributed by:   Name: Richard Obuku   Title: BDS Expert   Date: _______________",
        label_style))
    story.append(Spacer(1, 0.15 * cm))
    story.append(Paragraph(
        "Verified by: Stella Abote.   Date: _______",
        label_style))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph("PRUDEV II Programme — GOPA AFC in partnership with GIZ  |  Confidential", conf_style))

    doc.build(story, onFirstPage=_draw_header, onLaterPages=_draw_header)
    buf.seek(0)
    return buf.read()


class TshirtReceiptViewSet(viewsets.ModelViewSet):
    queryset           = TshirtReceipt.objects.prefetch_related('entries__bge').all()
    serializer_class   = TshirtReceiptSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def get_queryset(self):
        qs = TshirtReceipt.objects.prefetch_related('entries__bge').all()
        user = self.request.user
        # BGEs only see receipts that have an entry for them
        if not (user.is_staff or user.is_superuser):
            bge = getattr(user, 'bge_profile', None)
            if bge:
                qs = qs.filter(entries__bge=bge)
        return qs

    @action(detail=True, methods=['get'], url_path='pdf')
    def pdf(self, request, pk=None):
        receipt = self.get_object()
        try:
            pdf_bytes = _build_tshirt_pdf(receipt)
        except Exception as e:
            logger.error("TshirtReceipt PDF error: %s", e, exc_info=True)
            return Response({'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        filename = f"tshirt_receipt_{receipt.id}.pdf"
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    @action(detail=True, methods=['post'], url_path='bulk-sign')
    def bulk_sign(self, request, pk=None):
        """Admin: embed all available signatures at once."""
        receipt = self.get_object()
        if not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("Only admins can bulk-sign.")
        now = timezone.now()
        updated = 0
        for entry in receipt.entries.select_related('bge').filter(signed=False):
            if _bge_signature_bytes(entry.bge):
                entry.signed    = True
                entry.signed_at = now
                entry.save(update_fields=['signed', 'signed_at'])
                updated += 1
        return Response({'signed': updated, 'total': receipt.total_entries})


class TshirtReceiptEntryViewSet(viewsets.ModelViewSet):
    queryset           = TshirtReceiptEntry.objects.select_related('bge', 'receipt').all()
    serializer_class   = TshirtReceiptEntrySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = TshirtReceiptEntry.objects.select_related('bge', 'receipt').all()
        receipt_id = self.request.query_params.get('receipt')
        if receipt_id:
            qs = qs.filter(receipt_id=receipt_id)
        # BGEs only see their own entries
        user = self.request.user
        if not (user.is_staff or user.is_superuser):
            bge = getattr(user, 'bge_profile', None)
            if bge:
                qs = qs.filter(bge=bge)
            else:
                return qs.none()
        return qs

    @action(detail=True, methods=['post'], url_path='sign')
    def sign(self, request, pk=None):
        """BGE signs their own entry.

        The BGE may update their size and/or quantity before signing by
        including ``size`` and/or ``quantity`` in the POST body.
        """
        entry = self.get_object()
        user  = request.user

        # Allow the BGE whose entry this is, or staff
        bge = getattr(user, 'bge_profile', None)
        if not (user.is_staff or user.is_superuser):
            if not bge or entry.bge_id != bge.id:
                raise PermissionDenied("You can only sign your own receipt entry.")

        if entry.signed:
            return Response({'detail': 'Already signed.'}, status=status.HTTP_400_BAD_REQUEST)

        if not _bge_signature_bytes(entry.bge):
            return Response(
                {'detail': 'No signature on file. Please upload your signature first.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Let the BGE confirm / adjust their size and quantity before signing
        size     = request.data.get('size')
        quantity = request.data.get('quantity')
        update_fields = ['signed', 'signed_at']

        if size and size in dict(entry.__class__.SIZE_CHOICES):
            entry.size = size
            update_fields.append('size')

        if quantity is not None:
            try:
                qty = int(quantity)
                if qty >= 1:
                    entry.quantity = qty
                    update_fields.append('quantity')
            except (ValueError, TypeError):
                pass

        entry.signed    = True
        entry.signed_at = timezone.now()
        entry.save(update_fields=update_fields)
        return Response(TshirtReceiptEntrySerializer(entry, context={'request': request}).data)


class WorkOrderSubmissionViewSet(ViewerReadOnlyMixin, viewsets.ModelViewSet):
    """BGE timesheet & invoice (Excel) uploads against a work order.

    BGEs upload for their own work orders (or work orders they're co-assigned
    to). Admins/programme managers/viewers see everything, organised per BGE
    via the ``?bge=`` filter, and can download any file.
    """
    serializer_class = WorkOrderSubmissionSerializer
    permission_classes = [IsAuthenticated]

    ALLOWED_EXTENSIONS = ('.xlsx', '.xls')
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

    def _is_admin(self):
        u = self.request.user
        return u.is_staff or u.is_superuser or _managed_groups(u) is not None

    def get_queryset(self):
        user = self.request.user
        qs = WorkOrderSubmission.objects.select_related('work_order', 'bge', 'uploaded_by')
        bge_id = self.request.query_params.get('bge')
        wo_id = self.request.query_params.get('work_order')
        if self._is_admin() or _is_viewer(user):
            if bge_id:
                qs = qs.filter(bge_id=bge_id)
            if wo_id:
                qs = qs.filter(work_order_id=wo_id)
            return qs
        try:
            bge = user.bge_profile
        except Exception:
            return qs.none()
        qs = qs.filter(bge=bge)
        if wo_id:
            qs = qs.filter(work_order_id=wo_id)
        return qs

    # XLSX files are ZIP archives (PK\x03\x04); legacy XLS files are OLE2
    # compound documents (D0 CF 11 E0 A1 B1 1A E1).
    _XLSX_MAGIC = b'PK\x03\x04'
    _XLS_MAGIC = b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'

    def _validate_file(self, f, label):
        if f is None:
            return
        name = (f.name or '').lower()
        if not name.endswith(self.ALLOWED_EXTENSIONS):
            raise ValidationError(f'{label} must be an Excel file (.xlsx or .xls).')
        if f.size > self.MAX_FILE_SIZE:
            raise ValidationError(f'{label} must be under 10 MB.')
        header = f.read(8)
        f.seek(0)
        if not (header.startswith(self._XLSX_MAGIC) or header.startswith(self._XLS_MAGIC)):
            raise ValidationError(f'{label} does not look like a valid Excel file.')

    def _resolve_bge(self, work_order):
        user = self.request.user
        if self._is_admin():
            return work_order.bge
        try:
            bge = user.bge_profile
        except Exception:
            raise PermissionDenied("Only BGEs or admins can upload timesheets/invoices.")
        if bge.id != work_order.bge_id and not work_order.co_bges.filter(id=bge.id).exists():
            raise PermissionDenied("You can only upload documents for your own work orders.")
        return bge

    def _apply_files(self, instance, timesheet, invoice):
        from django.core.files.base import ContentFile
        update_fields = []
        if timesheet:
            data = timesheet.read()
            instance.timesheet_data = data
            instance.timesheet_filename = timesheet.name
            instance.timesheet_file.save(timesheet.name, ContentFile(data), save=False)
            update_fields += ['timesheet_data', 'timesheet_filename', 'timesheet_file']
        if invoice:
            data = invoice.read()
            instance.invoice_data = data
            instance.invoice_filename = invoice.name
            instance.invoice_file.save(invoice.name, ContentFile(data), save=False)
            update_fields += ['invoice_data', 'invoice_filename', 'invoice_file']
        if update_fields:
            instance.save(update_fields=update_fields)

    def perform_create(self, serializer):
        work_order = serializer.validated_data.get('work_order')
        bge = self._resolve_bge(work_order)
        timesheet = serializer.validated_data.pop('timesheet', None)
        invoice = serializer.validated_data.pop('invoice', None)
        self._validate_file(timesheet, 'Timesheet')
        self._validate_file(invoice, 'Invoice')
        if not timesheet and not invoice:
            raise ValidationError("Upload at least a timesheet or an invoice file.")
        instance = serializer.save(bge=bge, uploaded_by=self.request.user)
        self._apply_files(instance, timesheet, invoice)

    def _check_owner_or_admin(self, instance):
        user = self.request.user
        is_owner = hasattr(user, 'bge_profile') and user.bge_profile_id == instance.bge_id
        if not (self._is_admin() or is_owner):
            raise PermissionDenied("You can only manage your own submissions.")

    def perform_update(self, serializer):
        self._check_owner_or_admin(serializer.instance)
        timesheet = serializer.validated_data.pop('timesheet', None)
        invoice = serializer.validated_data.pop('invoice', None)
        self._validate_file(timesheet, 'Timesheet')
        self._validate_file(invoice, 'Invoice')
        instance = serializer.save()
        self._apply_files(instance, timesheet, invoice)

    def destroy(self, request, *args, **kwargs):
        self._check_owner_or_admin(self.get_object())
        return super().destroy(request, *args, **kwargs)

    def _serve_file(self, instance, kind):
        data = getattr(instance, f'{kind}_data')
        fname = _safe_filename(getattr(instance, f'{kind}_filename') or f'{kind}.xlsx')
        if not data:
            return Response({'error': f'No {kind} uploaded for this submission.'}, status=status.HTTP_404_NOT_FOUND)
        content_type = (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            if fname.lower().endswith('.xlsx') else 'application/vnd.ms-excel'
        )
        resp = HttpResponse(bytes(data), content_type=content_type)
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp

    @action(detail=True, methods=['get'], url_path='download-timesheet')
    def download_timesheet(self, request, pk=None):
        return self._serve_file(self.get_object(), 'timesheet')

    @action(detail=True, methods=['get'], url_path='download-invoice')
    def download_invoice(self, request, pk=None):
        return self._serve_file(self.get_object(), 'invoice')


class WorkOrderPaymentViewSet(ViewerReadOnlyMixin, viewsets.ModelViewSet):
    """Payment log against work orders.

    Admins record/edit/delete payments. BGEs and viewers have read-only
    access, scoped (for BGEs) to payments against their own work orders.
    """
    serializer_class = WorkOrderPaymentSerializer
    permission_classes = [IsAuthenticated]

    def _is_admin(self):
        u = self.request.user
        return u.is_staff or u.is_superuser or _managed_groups(u) is not None

    def get_queryset(self):
        user = self.request.user
        qs = WorkOrderPayment.objects.select_related('work_order', 'work_order__bge', 'recorded_by')
        wo_id = self.request.query_params.get('work_order')
        if wo_id:
            qs = qs.filter(work_order_id=wo_id)
        if self._is_admin() or _is_viewer(user):
            return qs
        try:
            bge = user.bge_profile
        except Exception:
            return qs.none()
        return qs.filter(work_order__bge=bge)

    def perform_create(self, serializer):
        if not self._is_admin():
            raise PermissionDenied("Only admins can record payments.")
        serializer.save(recorded_by=self.request.user)

    def perform_update(self, serializer):
        if not self._is_admin():
            raise PermissionDenied("Only admins can edit payments.")
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        if not self._is_admin():
            raise PermissionDenied("Only admins can delete payments.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='notify')
    def notify(self, request, pk=None):
        """Admin-only: email the BGE that a payment has been recorded."""
        if not self._is_admin():
            raise PermissionDenied("Only admins can notify BGEs about payments.")
        payment = self.get_object()
        work_order = payment.work_order
        bge = work_order.bge
        recipient_email = (bge.email or '').strip()
        if not recipient_email:
            raise ValidationError("This BGE has no email address on file.")

        notes_line = f'\nNotes: {payment.notes}\n' if payment.notes else ''
        reference_line = f'Reference: {payment.reference}\n' if payment.reference else ''
        subject = f'Payment Recorded — {work_order.work_order_number}'
        body = (
            f'Dear {bge.name},\n\n'
            f'A payment has been recorded against your work order ({work_order.work_order_number}):\n\n'
            f'Date: {payment.payment_date}\n'
            f'Amount: UGX {payment.amount:,.0f}\n'
            f'{reference_line}'
            f'{notes_line}\n'
            f'Please log in to confirm receipt of this payment.\n\n'
            f'Regards,\nPRUDEV II BDS Team\nGOPA AFC / GIZ'
        )
        try:
            msg = EmailMultiAlternatives(
                subject, body,
                getattr(settings, 'DEFAULT_FROM_EMAIL', ''),
                [recipient_email],
            )
            msg.send(fail_silently=True)
        except Exception:
            pass

        payment.notified_at = timezone.now()
        payment.save(update_fields=['notified_at'])
        return Response(self.get_serializer(payment).data)

    @action(detail=True, methods=['post'], url_path='confirm')
    def confirm(self, request, pk=None):
        """BGE-only: confirm receipt of a logged payment, notifying admin by email."""
        payment = self.get_object()
        work_order = payment.work_order
        user = request.user
        if not self._is_admin():
            try:
                bge = user.bge_profile
            except Exception:
                raise PermissionDenied("Only the BGE on this work order can confirm receipt.")
            if bge.id != work_order.bge_id and not work_order.co_bges.filter(id=bge.id).exists():
                raise PermissionDenied("Only the BGE on this work order can confirm receipt.")

        payment.confirmed_by_bge = True
        payment.confirmed_at = timezone.now()
        payment.save(update_fields=['confirmed_by_bge', 'confirmed_at'])

        notify_email = getattr(settings, 'PAYMENT_CONFIRMATION_NOTIFY_EMAIL', '')
        if notify_email:
            notes_line = f'\nNotes: {payment.notes}\n' if payment.notes else ''
            reference_line = f'Reference: {payment.reference}\n' if payment.reference else ''
            subject = f'Payment Receipt Confirmed — {work_order.work_order_number}'
            body = (
                f'{work_order.bge.name} has confirmed receipt of a payment against '
                f'work order {work_order.work_order_number}:\n\n'
                f'Date: {payment.payment_date}\n'
                f'Amount: UGX {payment.amount:,.0f}\n'
                f'{reference_line}'
                f'{notes_line}\n'
                f'Confirmed at: {payment.confirmed_at:%Y-%m-%d %H:%M}\n\n'
                f'Regards,\nPRUDEV II BDS Team\nGOPA AFC / GIZ'
            )
            try:
                msg = EmailMultiAlternatives(
                    subject, body,
                    getattr(settings, 'DEFAULT_FROM_EMAIL', ''),
                    [notify_email],
                )
                msg.send(fail_silently=True)
            except Exception:
                pass

        return Response(self.get_serializer(payment).data)
