"""
Management command: Export raw data for the Data Update Exercise in Excel format (.xlsx).

Usage:
    python manage.py export_data_update_excel
    python manage.py export_data_update_excel --output /path/to/custom_file.xlsx
"""

import os
from datetime import datetime, date
from django.core.management.base import BaseCommand
from django.db.models import Count, Max, Min, Q
from portfolio.models import MSME, MSMEGrowthSnapshot, BusinessGrowthExpert, Cohort


def build_data_update_workbook(cohort_name='Cohort 1 (Selected MSMEs)'):
    """
    Builds and returns an openpyxl Workbook containing all sheets for the Data Update Exercise (226 MSMEs).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ── Color Palette ─────────────────────────────────────────────────────────
    NAVY_PRIMARY   = '1A2F4B'
    NAVY_DARK      = '0F1D30'
    GOLD_ACCENT    = 'C89D3C'
    GREEN_SUCCESS  = '2E7D32'
    GREEN_LIGHT    = 'E8F5E9'
    AMBER_LIGHT    = 'FFF8E1'
    GREY_HEADER    = 'F1F5F9'
    GREY_ALT_ROW   = 'F8FAFC'
    BORDER_COLOR   = 'E2E8F0'

    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_bold   = Font(name='Calibri', size=10, bold=True)
    font_normal = Font(name='Calibri', size=10)
    font_small  = Font(name='Calibri', size=9, color='64748B')
    font_title  = Font(name='Calibri', size=14, bold=True, color=NAVY_PRIMARY)

    fill_navy   = PatternFill(start_color=NAVY_PRIMARY, end_color=NAVY_PRIMARY, fill_type='solid')
    fill_dark   = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type='solid')
    fill_alt    = PatternFill(start_color=GREY_ALT_ROW, end_color=GREY_ALT_ROW, fill_type='solid')
    fill_green  = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type='solid')
    fill_amber  = PatternFill(start_color=AMBER_LIGHT, end_color=AMBER_LIGHT, fill_type='solid')

    thin_border_side = Side(border_style='thin', color=BORDER_COLOR)
    cell_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side
    )

    align_left   = Alignment(horizontal='left', vertical='center')
    align_right  = Alignment(horizontal='right', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    align_wrap   = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def _auto_fit_columns(ws, max_col_width=45):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if '\n' in val:
                    val = max(val.split('\n'), key=len)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 11), max_col_width)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1: MSME Master & Latest Data Update
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet(title='MSME Master & Latest Update')
    ws1.views.sheetView[0].showGridLines = True

    if cohort_name:
        msmes = MSME.objects.filter(cohort__name=cohort_name).select_related('assigned_bge', 'cohort').prefetch_related('growth_snapshots').all().order_by('id')
    else:
        msmes = MSME.objects.select_related('assigned_bge', 'cohort').prefetch_related('growth_snapshots').all().order_by('id')

    total_msmes_count = msmes.count()

    # Title Block
    ws1.merge_cells('A1:AV1')
    title_cell = ws1['A1']
    title_cell.value = 'PRUDEV II — BDS Portfolio Master & Data Update Exercise'
    title_cell.font = font_title
    title_cell.alignment = align_left
    ws1.row_dimensions[1].height = 28

    ws1.cell(row=2, column=1, value=f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | Total Selected MSMEs in Exercise: {total_msmes_count}').font = font_small
    ws1.row_dimensions[2].height = 18

    headers_ws1 = [
        'MSME ID', 'MSME Code', 'Business Name', 'Business Type', 'Sector',
        'Business Category', 'Owner Name', 'Gender', 'Phone', 'Email',
        'Alt Contact Name', 'Alt Contact Role', 'Alt Phone', 'Alt Email',
        'District', 'City/Town', 'Physical Address', 'Operational Status',
        'Is Active', 'Latitude', 'Longitude', 'Assigned BGE', 'Cohort',
        'Has Update Snapshot', 'Latest Update Date', 'Annual Turnover (UGX)',
        'Last Month Rev (UGX)', 'FT Employees (M)', 'FT Employees (F)',
        'Total FT Staff', 'PT Employees (M)', 'PT Employees (F)', 'Total PT Staff',
        'Total Employees', 'Refugee Staff (FT+PT)', 'Has TIN', 'TIN Number',
        'Has URSB', 'URSB Reg No', 'Has Bank Account', 'Bank Name', 'Has SACCO',
        'Has MOMO Pay', 'MOMO Pay Code', 'Digital Tools Adopted', 'Training Made Changes',
        'Training Impact Areas', 'BGE Notes & Context'
    ]

    header_row_idx = 4
    ws1.row_dimensions[header_row_idx].height = 26
    for col_idx, h in enumerate(headers_ws1, 1):
        c = ws1.cell(row=header_row_idx, column=col_idx, value=h)
        c.font = font_header
        c.fill = fill_navy
        c.alignment = align_center
        c.border = cell_border

    current_row = 5
    for msme in msmes:
        snaps = list(msme.growth_snapshots.all())
        # Sort snapshots by snapshot_date descending
        snaps.sort(key=lambda s: s.snapshot_date, reverse=True)
        latest_snap = snaps[0] if snaps else None
        has_update = len(snaps) > 1 or (len(snaps) == 1 and snaps[0].source != 'diagnostic')

        # Workforce calculation
        ft_m = latest_snap.employees_ft_male if latest_snap and latest_snap.employees_ft_male is not None else (msme.diag_employees_ft_male or 0)
        ft_f = latest_snap.employees_ft_female if latest_snap and latest_snap.employees_ft_female is not None else (msme.diag_employees_ft_female or 0)
        pt_m = latest_snap.employees_pt_male if latest_snap and latest_snap.employees_pt_male is not None else (msme.diag_employees_pt_male or 0)
        pt_f = latest_snap.employees_pt_female if latest_snap and latest_snap.employees_pt_female is not None else (msme.diag_employees_pt_female or 0)
        tot_ft = ft_m + ft_f
        tot_pt = pt_m + pt_f
        tot_emp = tot_ft + tot_pt
        refugee = ((latest_snap.employees_ft_refugee or 0) + (latest_snap.employees_pt_refugee or 0)) if latest_snap else 0

        # Financials
        turnover = float(latest_snap.annual_turnover) if (latest_snap and latest_snap.annual_turnover is not None) else (float(msme.annual_revenue) if msme.annual_revenue else None)
        last_m_rev = float(latest_snap.last_month_revenue) if (latest_snap and latest_snap.last_month_revenue is not None) else None

        # Compliance
        has_tin = latest_snap.has_tin if (latest_snap and latest_snap.has_tin is not None) else msme.diag_has_tin
        tin_num = latest_snap.tin_number if latest_snap else ''
        has_ursb = latest_snap.has_ursb if (latest_snap and latest_snap.has_ursb is not None) else None
        ursb_num = latest_snap.ursb_reg_number if latest_snap else ''
        has_bank = latest_snap.has_business_bank if (latest_snap and latest_snap.has_business_bank is not None) else msme.diag_has_business_bank
        bank_name = latest_snap.bank_name if latest_snap else ''
        has_sacco = latest_snap.has_sacco if latest_snap else None
        has_momo = latest_snap.has_momo_pay if latest_snap else None
        momo_code = latest_snap.momo_pay_code if latest_snap else ''

        digital_tools = ', '.join(latest_snap.digital_tools) if (latest_snap and latest_snap.digital_tools) else ''
        training_changes = ', '.join(latest_snap.training_changes) if (latest_snap and latest_snap.training_changes) else ''
        notes = latest_snap.notes if latest_snap else (msme.business_description or '')

        row_data = [
            msme.id,
            msme.msme_code or '',
            msme.business_name,
            msme.get_business_type_display() if msme.business_type else '',
            msme.get_sector_display() if msme.sector else '',
            msme.business_category or '',
            msme.owner_name or '',
            msme.get_gender_display() if msme.gender else '',
            msme.phone or '',
            msme.email or '',
            msme.alt_contact_name or '',
            msme.alt_contact_role or '',
            msme.alt_phone or '',
            msme.alt_email or '',
            msme.district or '',
            msme.city or '',
            msme.address or '',
            (msme.status or 'active').upper(),
            'Yes' if msme.is_active else 'No',
            float(msme.latitude) if msme.latitude is not None else '',
            float(msme.longitude) if msme.longitude is not None else '',
            msme.assigned_bge.name if msme.assigned_bge else 'Unassigned',
            msme.cohort.name if msme.cohort else '',
            'Yes' if has_update else 'No',
            str(latest_snap.snapshot_date) if latest_snap else '',
            turnover if turnover is not None else '',
            last_m_rev if last_m_rev is not None else '',
            ft_m,
            ft_f,
            tot_ft,
            pt_m,
            pt_f,
            tot_pt,
            tot_emp,
            refugee,
            'Yes' if has_tin is True else ('No' if has_tin is False else ''),
            tin_num,
            'Yes' if has_ursb is True else ('No' if has_ursb is False else ''),
            ursb_num,
            'Yes' if has_bank is True else ('No' if has_bank is False else ''),
            bank_name,
            'Yes' if has_sacco is True else ('No' if has_sacco is False else ''),
            'Yes' if has_momo is True else ('No' if has_momo is False else ''),
            momo_code,
            digital_tools,
            'Yes' if (latest_snap and latest_snap.training_made_changes is True) else ('No' if (latest_snap and latest_snap.training_made_changes is False) else ''),
            training_changes,
            notes,
        ]

        ws1.row_dimensions[current_row].height = 20
        for col_idx, val in enumerate(row_data, 1):
            c = ws1.cell(row=current_row, column=col_idx, value=val)
            c.font = font_normal
            c.border = cell_border
            if current_row % 2 == 0:
                c.fill = fill_alt

            # Formatting
            if col_idx in [26, 27]: # Revenue
                if val != '':
                    c.number_format = '#,##0'
                c.alignment = align_right
            elif col_idx in [28, 29, 30, 31, 32, 33, 34, 35]: # Counts
                c.number_format = '#,##0'
                c.alignment = align_right
            elif col_idx in [1, 8, 18, 19, 24, 25, 36, 38, 40, 42, 43, 46]:
                c.alignment = align_center
            else:
                c.alignment = align_left

        current_row += 1

    ws1.freeze_panes = 'D5'
    _auto_fit_columns(ws1)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2: Growth Snapshots (Raw Point-in-Time Log)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title='All Snapshots (Raw Log)')
    ws2.views.sheetView[0].showGridLines = True

    all_snapshots = MSMEGrowthSnapshot.objects.filter(msme__in=msmes).select_related('msme', 'collected_by').all().order_by('msme_id', 'snapshot_date')

    ws2.merge_cells('A1:AH1')
    ws2['A1'].value = 'PRUDEV II — Growth Snapshots Raw Data Log'
    ws2['A1'].font = font_title
    ws2.row_dimensions[1].height = 28
    ws2.cell(row=2, column=1, value=f'Total Snapshot Records in Database for Selected MSMEs: {all_snapshots.count()}').font = font_small
    ws2.row_dimensions[2].height = 18

    headers_ws2 = [
        'Snapshot ID', 'MSME ID', 'MSME Code', 'Business Name', 'District', 'Sector',
        'Snapshot Date', 'Source / Stage', 'Collected By BGE', 'Annual Turnover (UGX)',
        'Last Month Revenue (UGX)', 'Total Assets (UGX)', 'FT Male', 'FT Female', 'Total FT',
        'PT Male', 'PT Female', 'Total PT', 'Total Staff', 'Refugee Staff',
        'Has TIN', 'TIN Number', 'Has URSB', 'URSB Reg No', 'Has Business Bank', 'Bank Name',
        'Has SACCO', 'Has MOMO Pay', 'MOMO Code', 'Digital Tools', 'Training Made Changes',
        'Training Impact Areas', 'Notes / Context', 'Created Timestamp'
    ]

    header_row_idx = 4
    ws2.row_dimensions[header_row_idx].height = 26
    for col_idx, h in enumerate(headers_ws2, 1):
        c = ws2.cell(row=header_row_idx, column=col_idx, value=h)
        c.font = font_header
        c.fill = fill_dark
        c.alignment = align_center
        c.border = cell_border

    all_snapshots = MSMEGrowthSnapshot.objects.select_related('msme', 'collected_by').all().order_by('msme_id', 'snapshot_date')

    current_row = 5
    for s in all_snapshots:
        ft_m = s.employees_ft_male or 0
        ft_f = s.employees_ft_female or 0
        pt_m = s.employees_pt_male or 0
        pt_f = s.employees_pt_female or 0
        tot_ft = ft_m + ft_f
        tot_pt = pt_m + pt_f
        tot_staff = tot_ft + tot_pt
        refugee = (s.employees_ft_refugee or 0) + (s.employees_pt_refugee or 0)

        turnover = float(s.annual_turnover) if s.annual_turnover is not None else ''
        last_m_rev = float(s.last_month_revenue) if s.last_month_revenue is not None else ''
        assets = float(s.total_assets) if s.total_assets is not None else ''

        row_data = [
            s.id,
            s.msme_id,
            s.msme.msme_code or '',
            s.msme.business_name,
            s.msme.district or '',
            s.msme.get_sector_display() if s.msme.sector else '',
            str(s.snapshot_date),
            s.get_source_display(),
            s.collected_by.name if s.collected_by else 'System / Baseline',
            turnover,
            last_m_rev,
            assets,
            ft_m,
            ft_f,
            tot_ft,
            pt_m,
            pt_f,
            tot_pt,
            tot_staff,
            refugee,
            'Yes' if s.has_tin is True else ('No' if s.has_tin is False else ''),
            s.tin_number or '',
            'Yes' if s.has_ursb is True else ('No' if s.has_ursb is False else ''),
            s.ursb_reg_number or '',
            'Yes' if s.has_business_bank is True else ('No' if s.has_business_bank is False else ''),
            s.bank_name or '',
            'Yes' if s.has_sacco is True else ('No' if s.has_sacco is False else ''),
            'Yes' if s.has_momo_pay is True else ('No' if s.has_momo_pay is False else ''),
            s.momo_pay_code or '',
            ', '.join(s.digital_tools or []),
            'Yes' if s.training_made_changes is True else ('No' if s.training_made_changes is False else ''),
            ', '.join(s.training_changes or []),
            s.notes or '',
            s.created_at.strftime('%Y-%m-%d %H:%M') if s.created_at else '',
        ]

        ws2.row_dimensions[current_row].height = 20
        for col_idx, val in enumerate(row_data, 1):
            c = ws2.cell(row=current_row, column=col_idx, value=val)
            c.font = font_normal
            c.border = cell_border
            if current_row % 2 == 0:
                c.fill = fill_alt

            if col_idx in [10, 11, 12]:
                if val != '':
                    c.number_format = '#,##0'
                c.alignment = align_right
            elif col_idx in [13, 14, 15, 16, 17, 18, 19, 20]:
                c.number_format = '#,##0'
                c.alignment = align_right
            elif col_idx in [1, 2, 7, 8, 21, 23, 25, 27, 28, 31]:
                c.alignment = align_center
            else:
                c.alignment = align_left

        current_row += 1

    ws2.freeze_panes = 'E5'
    _auto_fit_columns(ws2)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3: Baseline vs Latest Comparative Analysis
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet(title='Baseline vs Latest Comparison')
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells('A1:V1')
    ws3['A1'].value = 'PRUDEV II — Growth Impact & Baseline vs Latest Comparison'
    ws3['A1'].font = font_title
    ws3.row_dimensions[1].height = 28
    ws3.cell(row=2, column=1, value='Side-by-side growth variance analysis per MSME').font = font_small
    ws3.row_dimensions[2].height = 18

    headers_ws3 = [
        'MSME ID', 'MSME Code', 'Business Name', 'District', 'Sector', 'Assigned BGE',
        'Baseline Date', 'Latest Update Date',
        'Baseline Turnover (UGX)', 'Latest Turnover (UGX)', 'Turnover Change (UGX)', 'Turnover Growth %',
        'Baseline Total Staff', 'Latest Total Staff', 'Jobs Created / (Lost)',
        'Baseline TIN', 'Latest TIN',
        'Baseline URSB', 'Latest URSB',
        'Baseline Bank', 'Latest Bank',
        'Digital Tools Deployed'
    ]

    header_row_idx = 4
    ws3.row_dimensions[header_row_idx].height = 26
    for col_idx, h in enumerate(headers_ws3, 1):
        c = ws3.cell(row=header_row_idx, column=col_idx, value=h)
        c.font = font_header
        c.fill = fill_navy
        c.alignment = align_center
        c.border = cell_border

    current_row = 5
    for msme in msmes:
        snaps = list(msme.growth_snapshots.all().order_by('snapshot_date'))
        baseline_snap = next((s for s in snaps if s.source == 'diagnostic'), (snaps[0] if snaps else None))
        latest_snap = snaps[-1] if snaps else None

        base_date = str(baseline_snap.snapshot_date) if baseline_snap else ''
        latest_date = str(latest_snap.snapshot_date) if latest_snap else ''

        base_turnover = float(baseline_snap.annual_turnover) if (baseline_snap and baseline_snap.annual_turnover is not None) else (float(msme.annual_revenue) if msme.annual_revenue else None)
        latest_turnover = float(latest_snap.annual_turnover) if (latest_snap and latest_snap.annual_turnover is not None) else base_turnover

        rev_diff = (latest_turnover - base_turnover) if (latest_turnover is not None and base_turnover is not None) else ''
        rev_pct = f'{round(((latest_turnover - base_turnover) / base_turnover) * 100, 1)}%' if (latest_turnover is not None and base_turnover is not None and base_turnover > 0) else ''

        base_staff = ((baseline_snap.employees_ft_male or 0) + (baseline_snap.employees_ft_female or 0) + (baseline_snap.employees_pt_male or 0) + (baseline_snap.employees_pt_female or 0)) if baseline_snap else (msme.employee_count or 0)
        latest_staff = ((latest_snap.employees_ft_male or 0) + (latest_snap.employees_ft_female or 0) + (latest_snap.employees_pt_male or 0) + (latest_snap.employees_pt_female or 0)) if latest_snap else base_staff
        jobs_diff = latest_staff - base_staff

        base_tin = 'Yes' if (baseline_snap and baseline_snap.has_tin) else ('Yes' if msme.diag_has_tin else 'No')
        latest_tin = 'Yes' if (latest_snap and latest_snap.has_tin) else base_tin

        base_ursb = 'Yes' if (baseline_snap and baseline_snap.has_ursb) else 'No'
        latest_ursb = 'Yes' if (latest_snap and latest_snap.has_ursb) else base_ursb

        base_bank = 'Yes' if (baseline_snap and baseline_snap.has_business_bank) else ('Yes' if msme.diag_has_business_bank else 'No')
        latest_bank = 'Yes' if (latest_snap and latest_snap.has_business_bank) else base_bank

        digital_tools = ', '.join(latest_snap.digital_tools or []) if latest_snap else ''

        row_data = [
            msme.id,
            msme.msme_code or '',
            msme.business_name,
            msme.district or '',
            msme.get_sector_display() if msme.sector else '',
            msme.assigned_bge.name if msme.assigned_bge else 'Unassigned',
            base_date,
            latest_date,
            base_turnover if base_turnover is not None else '',
            latest_turnover if latest_turnover is not None else '',
            rev_diff,
            rev_pct,
            base_staff,
            latest_staff,
            jobs_diff,
            base_tin,
            latest_tin,
            base_ursb,
            latest_ursb,
            base_bank,
            latest_bank,
            digital_tools,
        ]

        ws3.row_dimensions[current_row].height = 20
        for col_idx, val in enumerate(row_data, 1):
            c = ws3.cell(row=current_row, column=col_idx, value=val)
            c.font = font_normal
            c.border = cell_border
            if current_row % 2 == 0:
                c.fill = fill_alt

            if col_idx in [9, 10, 11]:
                if val != '':
                    c.number_format = '#,##0'
                c.alignment = align_right
            elif col_idx in [13, 14, 15]:
                c.number_format = '#,##0'
                c.alignment = align_right
            elif col_idx in [1, 2, 7, 8, 12, 16, 17, 18, 19, 20, 21]:
                c.alignment = align_center
            else:
                c.alignment = align_left

        current_row += 1

    ws3.freeze_panes = 'D5'
    _auto_fit_columns(ws3)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4: BGE Field Collection Scorecard
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet(title='BGE Collection Scorecard')
    ws4.views.sheetView[0].showGridLines = True

    ws4.merge_cells('A1:H1')
    ws4['A1'].value = 'PRUDEV II — BGE Field Collection Progress & Scorecard'
    ws4['A1'].font = font_title
    ws4.row_dimensions[1].height = 28

    headers_ws4 = [
        'BGE Code', 'BGE Name', 'Phone', 'Location', 'Assigned MSMEs',
        'MSMEs with Updates', 'Update Progress %', 'Total Snapshots Collected'
    ]

    header_row_idx = 4
    ws4.row_dimensions[header_row_idx].height = 26
    for col_idx, h in enumerate(headers_ws4, 1):
        c = ws4.cell(row=header_row_idx, column=col_idx, value=h)
        c.font = font_header
        c.fill = fill_navy
        c.alignment = align_center
        c.border = cell_border

    bges = BusinessGrowthExpert.objects.all().order_by('name')
    current_row = 5

    for bge in bges:
        assigned_msmes = MSME.objects.filter(assigned_bge=bge)
        total_assigned = assigned_msmes.count()
        
        # Count MSMEs assigned to this BGE that have at least one snapshot
        updated_msmes = assigned_msmes.filter(growth_snapshots__isnull=False).distinct().count()
        total_snaps = MSMEGrowthSnapshot.objects.filter(Q(collected_by=bge) | Q(msme__assigned_bge=bge)).count()
        pct_progress = f'{round((updated_msmes / total_assigned) * 100, 1)}%' if total_assigned > 0 else '0.0%'

        row_data = [
            bge.bge_code or 'BGE',
            bge.name,
            bge.phone or '',
            bge.location or '',
            total_assigned,
            updated_msmes,
            pct_progress,
            total_snaps,
        ]

        ws4.row_dimensions[current_row].height = 20
        for col_idx, val in enumerate(row_data, 1):
            c = ws4.cell(row=current_row, column=col_idx, value=val)
            c.font = font_normal
            c.border = cell_border
            if current_row % 2 == 0:
                c.fill = fill_alt

            if col_idx in [5, 6, 8]:
                c.number_format = '#,##0'
                c.alignment = align_right
            elif col_idx in [1, 7]:
                c.alignment = align_center
            else:
                c.alignment = align_left

        current_row += 1

    _auto_fit_columns(ws4)

    return wb


def export_csv_files(output_dir):
    """
    Exports CSV files for all 4 sheets to the specified directory.
    """
    import csv

    os.makedirs(output_dir, exist_ok=True)
    wb = build_data_update_workbook()

    exported_files = []
    for sheet in wb.worksheets:
        filename = sheet.title.lower().replace(' ', '_').replace('&', 'and').replace('(', '').replace(')', '').replace('/', '_') + '.csv'
        filepath = os.path.join(output_dir, filename)
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # Skip title / subtitle rows (rows 1-3) and start from header row (row 4)
            for row in sheet.iter_rows(min_row=4, values_only=True):
                if any(row):
                    writer.writerow(row)
        exported_files.append(filepath)

    return exported_files


class Command(BaseCommand):
    help = 'Export all raw data for the Data Update Exercise into a formatted Excel workbook (.xlsx) and optionally CSVs.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output',
            type=str,
            default='data_update_raw_export.xlsx',
            help='Output file path for the generated Excel file',
        )
        parser.add_argument(
            '--csv-dir',
            type=str,
            default=None,
            help='Optional directory path to also export individual CSV files for each sheet',
        )

    def handle(self, *args, **options):
        output_path = options['output']
        csv_dir = options.get('csv_dir')

        self.stdout.write(f'Building Data Update Excel Workbook...')

        wb = build_data_update_workbook()
        wb.save(output_path)

        abs_path = os.path.abspath(output_path)
        self.stdout.write(self.style.SUCCESS(
            f'Successfully exported Data Update Exercise raw data to Excel:\n  {abs_path}'
        ))

        if csv_dir:
            self.stdout.write(f'Exporting CSV sheets to {csv_dir}...')
            csv_files = export_csv_files(csv_dir)
            for f in csv_files:
                self.stdout.write(self.style.SUCCESS(f'  Created CSV: {os.path.abspath(f)}'))

